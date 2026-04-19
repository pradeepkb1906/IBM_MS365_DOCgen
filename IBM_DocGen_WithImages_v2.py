"""# Last synced to OWUI DB: 2026-04-19 20:10 IST (caps 15/15/10+100, Notes + Folder sources, smart merges not..."""
import re
import json
import base64
import uuid
import io
import time
import threading
import traceback
import zipfile
import asyncio
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Optional, Literal, Any
from xml.etree import ElementTree as ET
from urllib.parse import urlparse
import requests
from pydantic import BaseModel, Field
from PIL import Image
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    fitz = None
    HAS_FITZ = False
    print("[DocGen] PyMuPDF (fitz) not installed — PDF extraction disabled. Install with: pip install pymupdf")
from fastapi.responses import HTMLResponse
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
try:
    import os as _os
    _CAIRO_CANDIDATES = [
        "/opt/homebrew/lib/libcairo.2.dylib",
        "/usr/local/lib/libcairo.2.dylib",
        "/opt/local/lib/libcairo.2.dylib",
        "/usr/lib/libcairo.so.2",
    ]
    for _p in _CAIRO_CANDIDATES:
        if _os.path.exists(_p):
            _dir = _os.path.dirname(_p)
            _existing = _os.environ.get("DYLD_FALLBACK_LIBRARY_PATH", "")
            if _dir not in _existing.split(":"):
                _os.environ["DYLD_FALLBACK_LIBRARY_PATH"] = (
                    _dir + (":" + _existing if _existing else "")
                )
            _existing_ld = _os.environ.get("LD_LIBRARY_PATH", "")
            if _dir not in _existing_ld.split(":"):
                _os.environ["LD_LIBRARY_PATH"] = (
                    _dir + (":" + _existing_ld if _existing_ld else "")
                )
            break
    del _os, _CAIRO_CANDIDATES
    import cairosvg
    HAS_CAIROSVG = True
except (ImportError, OSError) as _cairo_err:
    HAS_CAIROSVG = False
    print(f"[DocGen] cairosvg unavailable: {_cairo_err}")
PDF_EXT = {".pdf"}
DOCX_EXT = {".docx"}
PPTX_EXT = {".pptx"}
XLSX_EXT = {".xlsx", ".xlsm"}
IMG_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
SVG_EXT = {".svg"}
ALL_DOC_EXT = PDF_EXT | DOCX_EXT | PPTX_EXT | XLSX_EXT
ALL_IMG_EXT = IMG_EXT | SVG_EXT
NS_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
NS_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NS_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
IBM_BLUE_60 = "#0F62FE"
IBM_BLUE_70 = "#0043CE"
IBM_GRAY_100 = "#161616"
IBM_GRAY_70 = "#525252"
IBM_GRAY_20 = "#E0E0E0"
IBM_GRAY_10 = "#F4F4F4"
IBM_WHITE = "#FFFFFF"
SVG_THEME_CSS = """
:root {
  --color-text-primary:
  --color-text-secondary:
  --color-text-tertiary:
  --color-text-info:
  --color-text-success:
  --color-text-warning:
  --color-text-danger:
  --color-bg-primary:
  --color-bg-secondary:
  --color-bg-tertiary:
  --color-border-tertiary:
  --color-border-secondary:
  --color-border-primary:
  --font-sans: 'IBM Plex Sans', system-ui, -apple-system, sans-serif;
  --font-mono: 'IBM Plex Mono', 'SF Mono', Menlo, Consolas, monospace;
  --primary:
  --accent:
  --ramp-purple-fill:
  --ramp-teal-fill:
  --ramp-coral-fill:
  --ramp-pink-fill:
  --ramp-gray-fill:
  --ramp-blue-fill:
  --ramp-green-fill:
  --ramp-amber-fill:
  --ramp-red-fill:
}
:root[data-theme="dark"] {
  --color-text-primary:
  --color-bg-primary:
  --color-border-tertiary:
  --ramp-purple-fill:
  --ramp-teal-fill:
  --ramp-coral-fill:
  --ramp-pink-fill:
  --ramp-gray-fill:
  --ramp-blue-fill:
  --ramp-green-fill:
  --ramp-amber-fill:
  --ramp-red-fill:
}
"""
SVG_CLASSES_CSS = """
.t  { font: 400 14px/1.4 var(--font-sans); fill: var(--color-text-primary); }
.ts { font: 400 12px/1.4 var(--font-sans); fill: var(--color-text-secondary); }
.th { font: 500 14px/1.4 var(--font-sans); fill: var(--color-text-primary); }
.box    { fill: var(--color-bg-secondary); stroke: var(--color-border-tertiary); stroke-width: 0.5; }
.node   { cursor: pointer; }
.node:hover { opacity: 0.85; }
.arr    { stroke: var(--color-border-secondary); stroke-width: 1.5; fill: none; }
.leader { stroke: var(--color-text-tertiary); stroke-width: 0.5; stroke-dasharray: 3 2; fill: none; }
.c-purple>rect,.c-purple>circle,.c-purple>ellipse{fill:var(--ramp-purple-fill);stroke:var(--ramp-purple-stroke);stroke-width:.5}
.c-purple>.th{fill:var(--ramp-purple-th)!important} .c-purple>.ts{fill:var(--ramp-purple-ts)!important}
.c-teal>rect,.c-teal>circle,.c-teal>ellipse{fill:var(--ramp-teal-fill);stroke:var(--ramp-teal-stroke);stroke-width:.5}
.c-teal>.th{fill:var(--ramp-teal-th)!important} .c-teal>.ts{fill:var(--ramp-teal-ts)!important}
.c-coral>rect,.c-coral>circle,.c-coral>ellipse{fill:var(--ramp-coral-fill);stroke:var(--ramp-coral-stroke);stroke-width:.5}
.c-coral>.th{fill:var(--ramp-coral-th)!important} .c-coral>.ts{fill:var(--ramp-coral-ts)!important}
.c-pink>rect,.c-pink>circle,.c-pink>ellipse{fill:var(--ramp-pink-fill);stroke:var(--ramp-pink-stroke);stroke-width:.5}
.c-pink>.th{fill:var(--ramp-pink-th)!important} .c-pink>.ts{fill:var(--ramp-pink-ts)!important}
.c-gray>rect,.c-gray>circle,.c-gray>ellipse{fill:var(--ramp-gray-fill);stroke:var(--ramp-gray-stroke);stroke-width:.5}
.c-gray>.th{fill:var(--ramp-gray-th)!important} .c-gray>.ts{fill:var(--ramp-gray-ts)!important}
.c-blue>rect,.c-blue>circle,.c-blue>ellipse{fill:var(--ramp-blue-fill);stroke:var(--ramp-blue-stroke);stroke-width:.5}
.c-blue>.th{fill:var(--ramp-blue-th)!important} .c-blue>.ts{fill:var(--ramp-blue-ts)!important}
.c-green>rect,.c-green>circle,.c-green>ellipse{fill:var(--ramp-green-fill);stroke:var(--ramp-green-stroke);stroke-width:.5}
.c-green>.th{fill:var(--ramp-green-th)!important} .c-green>.ts{fill:var(--ramp-green-ts)!important}
.c-amber>rect,.c-amber>circle,.c-amber>ellipse{fill:var(--ramp-amber-fill);stroke:var(--ramp-amber-stroke);stroke-width:.5}
.c-amber>.th{fill:var(--ramp-amber-th)!important} .c-amber>.ts{fill:var(--ramp-amber-ts)!important}
.c-red>rect,.c-red>circle,.c-red>ellipse{fill:var(--ramp-red-fill);stroke:var(--ramp-red-stroke);stroke-width:.5}
.c-red>.th{fill:var(--ramp-red-th)!important} .c-red>.ts{fill:var(--ramp-red-ts)!important}
"""
SVG_BASE_STYLES = """
* { box-sizing: border-box; margin: 0; font-family: var(--font-sans); }
html, body { overflow: hidden; }
body { background: var(--color-bg-primary); color: var(--color-text-primary); line-height: 1.7; padding: 8px; }
svg { overflow: visible; }
svg text { fill: var(--color-text-primary); }
h1 { font-size: 22px; font-weight: 500; margin-bottom: 12px; }
h2 { font-size: 18px; font-weight: 500; margin-bottom: 8px; }
h3 { font-size: 16px; font-weight: 500; margin-bottom: 6px; }
p  { font-size: 14px; color: var(--color-text-secondary); margin-bottom: 8px; }
button { background: transparent; border: 0.5px solid var(--color-border-secondary); border-radius: 8px; padding: 6px 14px; font-size: 13px; color: var(--color-text-primary); cursor: pointer; }
button:hover { background: var(--color-bg-secondary); }
.responsive-container { width: 100%; max-width: 100%; overflow-x: auto; }
"""
SVG_BODY_SCRIPTS = """
<script>
// SVG download bar — auto-injected when an SVG is present.
window.addEventListener('load',function(){document.querySelectorAll('svg[viewBox]').forEach(function(svg,i){
var bar=document.createElement('div');bar.style.cssText='display:flex;gap:6px;padding:6px 0;margin-bottom:8px';
['SVG','PNG','JPG'].forEach(function(f){var b=document.createElement('button');b.textContent='\\u2913 '+f;
b.style.cssText='background:#0f62fe;color:#fff;border:none;border-radius:4px;padding:4px 12px;font-size:11px;cursor:pointer;font-weight:600';
b.onclick=function(){_dlS(svg,f,i)};bar.appendChild(b)});svg.parentNode.insertBefore(bar,svg.nextSibling)})});
function _dlB(b,n){var u=URL.createObjectURL(b),a=document.createElement('a');a.href=u;a.download=n;document.body.appendChild(a);a.click();document.body.removeChild(a);URL.revokeObjectURL(u)}
function _dlS(svg,fmt,idx){var cl=svg.cloneNode(true),vb=svg.viewBox.baseVal,w=vb&&vb.width>0?vb.width:680,h=vb&&vb.height>0?vb.height:400;
cl.setAttribute('xmlns','http://www.w3.org/2000/svg');
var sr=svg.querySelectorAll('rect,circle,ellipse,path,line,polyline,polygon,text'),cr=cl.querySelectorAll('rect,circle,ellipse,path,line,polyline,polygon,text');
for(var i=0;i<sr.length;i++){var cs=getComputedStyle(sr[i]),ce=cr[i];ce.style.fill=ce.style.fill||cs.fill;if(ce.tagName==='text')ce.style.fontSize=ce.style.fontSize||cs.fontSize;else ce.style.stroke=ce.style.stroke||cs.stroke}
var xml=new XMLSerializer().serializeToString(cl),nm='diagram_'+(idx+1);
if(fmt==='SVG'){_dlB(new Blob([xml],{type:'image/svg+xml'}),nm+'.svg');return}
var img=new Image();img.onload=function(){var c=document.createElement('canvas');c.width=w*2;c.height=h*2;var ctx=c.getContext('2d');
if(fmt==='JPG'){ctx.fillStyle='#fff';ctx.fillRect(0,0,c.width,c.height)}ctx.drawImage(img,0,0,c.width,c.height);
c.toBlob(function(bl){_dlB(bl,nm+'.'+(fmt==='JPG'?'jpg':'png'))},fmt==='JPG'?'image/jpeg':'image/png',.95)};
img.src='data:image/svg+xml;base64,'+btoa(unescape(encodeURIComponent(xml)))}
// Iframe height resizer — works when allow-same-origin is on (our patched OWUI default).
(function(){function fit(){try{var fe=window.frameElement;if(fe){var h=Math.max(720, document.documentElement.scrollHeight);fe.style.height=h+'px';fe.style.minHeight=h+'px';fe.style.width='100%';}}catch(e){}}
fit();setTimeout(fit,100);setTimeout(fit,500);window.addEventListener('load',fit);
new MutationObserver(fit).observe(document.documentElement,{attributes:true,childList:true,subtree:true});})();
</script>
"""
_SVG_WRAPPER_TAG_RE = re.compile(
    r"<!DOCTYPE[^>]*>|</?html[^>]*>|</?head[^>]*>|</?body[^>]*>|<meta[^>]*/?>",
    re.IGNORECASE,
)
def _sanitize_svg_content(content: str) -> str:
    if not content: return ""
    content = _SVG_WRAPPER_TAG_RE.sub("", content)
    content = re.sub(r"\n{3,}", "\n\n", content)
    return content.strip()
def _build_svg_shell(content: str, title: str = "Diagram") -> str:
    content = _sanitize_svg_content(content)
    return (
        "<!DOCTYPE html><html><head><meta charset=\"utf-8\">"
        f"<title>{title}</title>"
        f"<style>{SVG_THEME_CSS}\n{SVG_CLASSES_CSS}\n{SVG_BASE_STYLES}</style>"
        "</head><body>"
        f"<div class=\"responsive-container\">{content}</div>"
        f"{SVG_BODY_SCRIPTS}"
        "</body></html>"
    )
def _svg_to_png_bytes(svg_str: str, output_width: int = 1200) -> Optional[bytes]:
    if not HAS_CAIROSVG or not svg_str: return None
    try:
        svg_clean = _sanitize_svg_content(svg_str)
        if not svg_clean.lstrip().startswith("<svg"): return None
        return cairosvg.svg2png(bytestring=svg_clean.encode("utf-8"), output_width=output_width)
    except Exception as e:
        print(f"[DocGen] SVG rasterization failed: {e}")
        return None
class _ImageStore:
    MAX_IMAGES = 600
    MAX_BYTES = 500 * 1024 * 1024
    TTL_SECONDS = 3600
    def __init__(self):
        self._store = {}
        self._lock = threading.Lock()
    def put(self, img_id: str, png_bytes: bytes, metadata: dict):
        with self._lock:
            self._store[img_id] = {
                "png_bytes": png_bytes,
                "metadata": metadata,
                "created_at": time.time(),
                "pinned": False,
            }
            self._evict()
    def pin(self, img_id: str, display_id: Optional[str] = None):
        with self._lock:
            rec = self._store.get(img_id)
            if not rec: return
            rec["pinned"] = True
            rec["created_at"] = time.time()
            if display_id:
                rec["metadata"]["display_id"] = display_id
    def get_bytes(self, img_id: str) -> Optional[bytes]:
        with self._lock:
            rec = self._store.get(img_id)
            return rec["png_bytes"] if rec else None
    def get_metadata(self, img_id: str) -> Optional[dict]:
        with self._lock:
            rec = self._store.get(img_id)
            return rec["metadata"] if rec else None
    def get_by_display_id(self, display_id: str) -> tuple:
        with self._lock:
            for rec in self._store.values():
                if rec["metadata"].get("display_id") == display_id: return rec["png_bytes"], rec["metadata"]
        return None, None
    def _evict(self):
        now = time.time()
        expired = [k for k, v in self._store.items()
                   if not v.get("pinned") and now - v["created_at"] > self.TTL_SECONDS]
        for k in expired:
            del self._store[k]
        def _unpinned_oldest():
            cands = [(k, v["created_at"]) for k, v in self._store.items() if not v.get("pinned")]
            return min(cands, key=lambda p: p[1])[0] if cands else None
        while len(self._store) > self.MAX_IMAGES:
            k = _unpinned_oldest()
            if not k: break
            del self._store[k]
        def _total():
            return sum(len(v["png_bytes"]) for v in self._store.values())
        while _total() > self.MAX_BYTES and self._store:
            k = _unpinned_oldest()
            if not k: break
            del self._store[k]
_IMAGE_STORE = _ImageStore()
class _ExtractionCache:
    TTL_SECONDS = 3600
    MAX_ENTRIES = 64
    def __init__(self):
        self._store = {}
        self._lock = threading.Lock()
    def _hash(self, file_bytes: bytes) -> str:
        h = hashlib.sha1()
        h.update(str(len(file_bytes)).encode())
        h.update(file_bytes[:65536])
        h.update(file_bytes[-65536:])
        return h.hexdigest()
    def get(self, file_bytes: bytes):
        key = self._hash(file_bytes)
        with self._lock:
            rec = self._store.get(key)
            if not rec: return None
            if time.time() - rec["t"] > self.TTL_SECONDS:
                del self._store[key]
                return None
            return rec["text"], rec["images"]
    def put(self, file_bytes: bytes, text_chunks: list, images: list):
        key = self._hash(file_bytes)
        with self._lock:
            self._store[key] = {"t": time.time(), "text": text_chunks, "images": images}
            while len(self._store) > self.MAX_ENTRIES:
                oldest = min(self._store.keys(), key=lambda k: self._store[k]["t"])
                del self._store[oldest]
_EXTRACT_CACHE = _ExtractionCache()
class _DocBuffer:
    TTL_SECONDS = 600
    def __init__(self):
        self._sessions = {}
        self._lock = threading.Lock()
    def ensure(self, session_id: str, title: str, fmt: str):
        with self._lock:
            self._cleanup()
            if session_id not in self._sessions:
                self._sessions[session_id] = {
                    "title": title,
                    "format": fmt,
                    "pages": {},
                    "image_refs": [],
                    "created_at": time.time(),
                }
            return self._sessions[session_id]
    def add_page(self, session_id: str, page_num: int, elements: list):
        with self._lock:
            if session_id in self._sessions:
                self._sessions[session_id]["pages"][page_num] = elements
    def pop(self, session_id: str):
        with self._lock: return self._sessions.pop(session_id, None)
    def get(self, session_id: str):
        with self._lock: return self._sessions.get(session_id)
    def _cleanup(self):
        now = time.time()
        expired = [k for k, v in self._sessions.items() if now - v["created_at"] > self.TTL_SECONDS]
        for k in expired:
            del self._sessions[k]
_DOC_BUFFER = _DocBuffer()
class _MCPClient:
    """Dual-transport MCP client. One instance per MCP server URL."""
    TRANSPORT_SSE = "sse"
    TRANSPORT_STREAMABLE = "stream"
    def __init__(self, url: str, headers: Optional[dict] = None,
                 timeout: int = 60, transport: Optional[str] = None):
        self.url = url.rstrip("/")
        self.headers = dict(headers or {})
        self.headers.setdefault("Accept", "application/json, text/event-stream")
        self.timeout = timeout
        self._transport = transport or self._detect_transport()
        self._lock = threading.Lock()
        self._next_id = 0
        self._initialized = False
        self._session_id: Optional[str] = None
        self._sse_thread: Optional[threading.Thread] = None
        self._sse_response = None
        self._messages_url: Optional[str] = None
        self._pending: dict = {}
        self._results: dict = {}
        self._endpoint_event = threading.Event()
        self._stream_error: Optional[str] = None
    def _detect_transport(self) -> str:
        low = self.url.lower()
        if low.endswith("/sse") or "/sse?" in low: return self.TRANSPORT_SSE
        if low.endswith("/mcp") or "/mcp?" in low: return self.TRANSPORT_STREAMABLE
        return self.TRANSPORT_STREAMABLE
    def _rpc_id(self) -> int:
        with self._lock:
            self._next_id += 1
            return self._next_id
    def _ensure_sse_stream(self):
        if self._sse_thread and self._sse_thread.is_alive() and self._messages_url: return
        if self._stream_error: raise RuntimeError(self._stream_error)
        get_headers = {k: v for k, v in self.headers.items() if k.lower() != "content-type"}
        get_headers["Accept"] = "text/event-stream"
        try:
            response = requests.get(self.url, headers=get_headers,
                                     stream=True, timeout=self.timeout)
        except Exception as e: raise RuntimeError(f"MCP SSE connect failed: {e}")
        if response.status_code != 200: raise RuntimeError(f"MCP SSE HTTP {response.status_code}: {response.text[:300]}")
        self._sse_response = response
        self._endpoint_event.clear()
        def pump():
            try:
                self._pump_sse(response)
            except Exception as e:
                self._stream_error = str(e)
                self._endpoint_event.set()
        self._sse_thread = threading.Thread(target=pump, daemon=True)
        self._sse_thread.start()
        if not self._endpoint_event.wait(timeout=15): raise RuntimeError("MCP SSE: no endpoint event received in 15s")
        if not self._messages_url: raise RuntimeError(f"MCP SSE: endpoint event missing URL ({self._stream_error or ''})")
    def _pump_sse(self, response):
        event_name = "message"
        data_buf = []
        for raw in response.iter_lines(chunk_size=1, decode_unicode=True):
            if raw is None: continue
            if raw == "":
                data_str = "\n".join(data_buf).strip()
                data_buf = []
                if not data_str:
                    event_name = "message"
                    continue
                if event_name == "endpoint":
                    ep = data_str
                    if ep.startswith("/"):
                        parsed = urlparse(self.url)
                        self._messages_url = f"{parsed.scheme}://{parsed.netloc}{ep}"
                    elif ep.startswith("http"):
                        self._messages_url = ep
                    else:
                        base = self.url.rsplit("/", 1)[0]
                        self._messages_url = f"{base}/{ep}"
                    self._endpoint_event.set()
                elif event_name in ("message", "") and data_str.startswith("{"):
                    try:
                        msg = json.loads(data_str)
                    except Exception:
                        event_name = "message"
                        continue
                    mid = msg.get("id")
                    if mid is not None and mid in self._pending:
                        self._results[mid] = msg
                        self._pending[mid].set()
                event_name = "message"
                continue
            if raw.startswith(":"): continue
            if raw.startswith("event:"):
                event_name = raw[6:].strip()
            elif raw.startswith("data:"):
                data_buf.append(raw[5:].lstrip())
            elif raw.startswith("id:") or raw.startswith("retry:"): pass
    def _rpc_via_sse(self, method: str, params: Optional[dict] = None,
                      is_notification: bool = False) -> dict:
        self._ensure_sse_stream()
        rid = None if is_notification else self._rpc_id()
        payload = {"jsonrpc": "2.0", "method": method}
        if rid is not None:
            payload["id"] = rid
        if params is not None:
            payload["params"] = params
        post_headers = dict(self.headers)
        post_headers["Content-Type"] = "application/json"
        post_headers.setdefault("Accept", "application/json, text/event-stream")
        if is_notification:
            try:
                requests.post(self._messages_url, json=payload,
                              headers=post_headers, timeout=10)
            except Exception: pass
            return {}
        event = threading.Event()
        self._pending[rid] = event
        try:
            r = requests.post(self._messages_url, json=payload,
                              headers=post_headers, timeout=self.timeout)
            if r.status_code not in (200, 202): raise RuntimeError(f"MCP SSE POST HTTP {r.status_code}: {r.text[:300]}")
            if r.content:
                ct = (r.headers.get("Content-Type") or "").lower()
                if "application/json" in ct:
                    try:
                        msg = r.json()
                        if isinstance(msg, dict) and msg.get("id") == rid:
                            self._results[rid] = msg
                            event.set()
                    except Exception: pass
            if not event.wait(timeout=self.timeout): raise RuntimeError(f"MCP SSE: no response for {method} within {self.timeout}s")
            msg = self._results.pop(rid, None)
            if msg is None: raise RuntimeError(f"MCP SSE: empty result for {method}")
            if "error" in msg: raise RuntimeError(f"MCP error: {msg['error']}")
            return msg.get("result", {})
        finally:
            self._pending.pop(rid, None)
    def _rpc_via_streamable(self, method: str, params: Optional[dict] = None,
                              is_notification: bool = False) -> dict:
        payload = {"jsonrpc": "2.0", "method": method}
        if not is_notification:
            payload["id"] = self._rpc_id()
        if params is not None:
            payload["params"] = params
        headers = dict(self.headers)
        headers["Content-Type"] = "application/json"
        if self._session_id:
            headers["Mcp-Session-Id"] = self._session_id
        try:
            r = requests.post(self.url, json=payload, headers=headers,
                               timeout=self.timeout, stream=True)
        except Exception as e: raise RuntimeError(f"MCP streamable transport error: {e}")
        if is_notification: return {}
        if r.status_code == 404 and self._session_id:
            self._session_id = None
            self._initialized = False
            self.initialize()
            headers["Mcp-Session-Id"] = self._session_id or ""
            r = requests.post(self.url, json=payload, headers=headers,
                               timeout=self.timeout, stream=True)
        if r.status_code >= 400: raise RuntimeError(f"MCP HTTP {r.status_code}: {r.text[:500]}")
        new_sid = r.headers.get("Mcp-Session-Id") or r.headers.get("mcp-session-id")
        if new_sid:
            self._session_id = new_sid
        ct = (r.headers.get("Content-Type") or "").lower()
        want_id = payload["id"]
        if "text/event-stream" in ct: return self._parse_sse_inline(r, want_id)
        try:
            data = r.json()
        except Exception: raise RuntimeError(f"MCP: non-JSON response: {r.text[:300]}")
        if "error" in data: raise RuntimeError(f"MCP error: {data['error']}")
        return data.get("result", {})
    def _parse_sse_inline(self, response, want_id: int) -> dict:
        data_buf = []
        for raw in response.iter_lines(chunk_size=1, decode_unicode=True):
            if raw is None: continue
            if raw == "":
                data_str = "\n".join(data_buf).strip()
                data_buf = []
                if data_str and data_str.startswith("{"):
                    try:
                        msg = json.loads(data_str)
                    except Exception: continue
                    if msg.get("id") == want_id:
                        if "error" in msg: raise RuntimeError(f"MCP error: {msg['error']}")
                        return msg.get("result", {})
                continue
            if raw.startswith("data:"):
                data_buf.append(raw[5:].lstrip())
        raise RuntimeError("MCP streamable SSE ended without matching response")
    def _rpc(self, method: str, params: Optional[dict] = None,
              is_notification: bool = False) -> dict:
        if self._transport == self.TRANSPORT_SSE: return self._rpc_via_sse(method, params, is_notification)
        return self._rpc_via_streamable(method, params, is_notification)
    def initialize(self) -> dict:
        if self._initialized: return {}
        result = self._rpc("initialize", {
            "protocolVersion": "2025-06-18",
            "capabilities": {"sampling": {}, "roots": {"listChanged": False}},
            "clientInfo": {"name": "ibm-docgen-withimages", "version": "2.1"},
        })
        try:
            self._rpc("notifications/initialized", is_notification=True)
        except Exception: pass
        self._initialized = True
        return result
    def list_tools(self) -> list[dict]:
        self.initialize()
        result = self._rpc("tools/list")
        return result.get("tools", [])
    def call_tool(self, name: str, arguments: Optional[dict] = None) -> dict:
        self.initialize()
        return self._rpc("tools/call", {
            "name": name,
            "arguments": arguments or {},
        })
    def read_resource(self, uri: str) -> dict:
        self.initialize()
        return self._rpc("resources/read", {"uri": uri})
_MCP_CLIENTS: dict = {}
_MCP_CLIENTS_LOCK = threading.Lock()
def _get_mcp_client(url: str, headers: Optional[dict] = None,
                     transport: Optional[str] = None) -> _MCPClient:
    key = url.rstrip("/")
    with _MCP_CLIENTS_LOCK:
        if key not in _MCP_CLIENTS:
            _MCP_CLIENTS[key] = _MCPClient(url, headers=headers, transport=transport)
        return _MCP_CLIENTS[key]
class Tools:
    """Single-file IBM document generator."""
    class Valves(BaseModel):
        owui_base_url: str = Field(default="http://localhost:8080")
        google_api_key: str = Field(
            default="",
            description="Google Programmable Search API key (web search mode). Empty = disabled.",
        )
        google_cx: str = Field(
            default="",
            description="Google Programmable Search engine ID (CX).",
        )
        mcp_servers_json: str = Field(
            default="",
            description=(
                'JSON array of MCP servers to make callable. Each entry: '
                '{"id": "<short-id>", "name": "<display>", "url": "<streamable-http URL>", '
                '"auth_header": "<optional Authorization header value>", "tools": ["tool1","tool2"] (optional filter)}. '
                'Example: [{"id":"ibm_hr","name":"IBM HR","url":"https://contextforge.ibm.com/hr/mcp","auth_header":"Bearer xyz"}]. '
                'Leave empty to disable MCP mode.'
            ),
        )
        mcp_max_image_extract_per_call: int = Field(
            default=3,
            description="Max images to extract from a single MCP tool result (safety cap).",
        )
        mcp_catalog_ttl_seconds: int = Field(
            default=600,
            description="How long to cache each MCP server's tool list before re-listing (seconds).",
        )
        max_text_chunks: int = Field(default=12)
        max_image_candidates: int = Field(default=10)
        min_image_width: int = Field(default=100)
        min_image_height: int = Field(default=100)
        max_image_aspect_ratio: float = Field(default=6.0)
        max_image_bytes: int = Field(default=3_500_000)
        web_image_fetch_timeout: int = Field(
            default=3,
            description="Per-image HTTP download timeout (seconds). 3s forces fast fallthrough on blocked/slow hosts. Good hosts (Wikimedia, CloudFront) return in <1s anyway.",
        )
        request_timeout: int = Field(
            default=30,
            description="Generic HTTP request timeout (seconds). API calls to Google / DuckDuckGo / Wikipedia.",
        )
        image_fetch_parallelism: int = Field(
            default=4,
            description="Parallel workers when downloading images WITHIN one source's candidate list. Lowered from 8 to 4 to avoid tripping Wikimedia 429 rate-limits.",
        )
        enrich_parallelism: int = Field(
            default=4,
            description="Parallel image fetches when enriching a batch of sections at once. Each worker can hit multiple sources serially. Lowered from 6 to 4 to spread Wikimedia load.",
        )
        skip_vision_rank_for_web: bool = Field(
            default=True,
            description="When source is web/DuckDuckGo and the query was targeted (e.g. 'Red Fort Delhi'), skip the vision-ranking round-trip — trust the search. Big latency win.",
        )
        disable_image_enrichment: bool = Field(
            default=True,
            description="DEFAULT TRUE. enrich_sections_with_images returns immediately with text-only sections — no web fetch, no placeholder. Latency win: ~15-40s saved per deck. When the user explicitly asks for images ('with photos', 'include visuals'), the LLM can still call generate_image() for ONE hero image. Flip to False to auto-enrich every deck again.",
        )
        security_level: Literal["strict", "balanced", "none"] = Field(default="strict")
        vision_rank_enabled: bool = Field(
            default=True,
            description="If true, send extracted image thumbnails to the multimodal base model so it captions + scores them. Descriptions improve ranking and image captions in the final doc.",
        )
        vision_rank_model: str = Field(
            default="claude-opus-4-6",
            description="Model ID used for vision captioning/ranking. Must support image input.",
        )
        vision_rank_max_images: int = Field(
            default=24,
            description="Max images to send to the vision model per request (safety cap).",
        )
        vision_rank_thumb_px: int = Field(
            default=512,
            description="Longest-edge px for the thumbnail sent to the vision model. Smaller = cheaper/faster.",
        )
        enable_kb_vision_layout: bool = Field(default=True,
            description="When attachments are present, auto-build 50/50 pages/slides: left = generated text, right = Claude-Vision-matched image (or reference card) from the KB.")
        kb_vision_score_threshold: int = Field(default=95,
            description="Min score 0-100 from Claude Vision to accept a KB match. Below = section falls back to text-only full-width.")
        kb_max_candidates_per_section: int = Field(default=5,
            description="Max candidate images sent to Claude Vision per section (Plan C).")
        kb_plan_b_enabled: bool = Field(default=True,
            description="Try Anthropic native PDF document block first. If Bedrock/litellm rejects it, fall back to Plan C zipfile extraction.")
    def __init__(self):
        self.valves = self.Valves()
        self._logo_png_cache: Optional[bytes] = None
        self._phase: str = ""
        self._phase_started: float = 0.0
        from requests.adapters import HTTPAdapter
        self._http = requests.Session()
        _adapter = HTTPAdapter(pool_connections=16, pool_maxsize=16, max_retries=0)
        self._http.mount("http://", _adapter)
        self._http.mount("https://", _adapter)
    _LOGO_URL = (
        "https://upload.wikimedia.org/wikipedia/commons/thumb/5/51/"
        "IBM_logo.svg/600px-IBM_logo.svg.png"
    )
    _LOGO_FALLBACK_B64 = (
        "iVBORw0KGgoAAAANSUhEUgAAAlgAAADwCAYAAADcthp2AAAEhklEQVR42u3cAYrEIBBE0dSS+1+59goTkNDG9w4gKvTwMTBpewEA"
        "sM6fKwAAEFgAAAILAEBgAQAgsAAABBYAgMACAEBgAQAILACA/d2rF0zir+EZoW2m7MVcYC7MBWfNhRcsAACBBQAgsAAABBYAAAIL"
        "AEBgAQAILAAABBYAgMACABBYAAAILAAAgQUAILAAAAQWAAACCwBgrLR1CwAAlxcsAACBBQBwinv1gkl8c2SEtpmyl93n4sld7nTW"
        "X8/1pd81cwHvzIUXLAAAgQUAILAAAAQWAAACCwBAYAEACCwAAAQWAIDAAgAQWAAACCwAAIEFACCwAAAEFgAAAgsAYKy0dQsAAJcX"
        "LAAAgQUAcIp79YJJfHNkhLaZspfd5+LJXe501l/P9aXfNXMB78yFFywAAIEFACCwAAAEFgAAAgsAQGABAAgsAAAEFgCAwAIAEFgA"
        "AAgsAACBBQAgsAAABBYAAAILAGCstHULAACXFywAAIEFAHCKe/WCSXxzZIS2mbKX3efiyV3udNZfz/Wl3zVzAe/MhRcsAACBBQAg"
        "sAAABBYAAAILAEBgAQAILAAABBYAgMACABBYAAAILAAAgQUAILAAAAQWAAACCwBgrLR1CwAAlxcsAACBBQBwinv1gkl8c2SEtpmy"
        "F3OBuTAXnDUXXrAAAAQWAIDAAgAQWAAACCwAAIEFACCwAAAQWAAAAgsAQGABACCwAAAEFgCAwAIAEFgAAAgsAICx0tYtAABcXrAA"
        "AAQWAMAp7tULJvHNkRHaZspedp+LJ3e501l/PdeXftfMBbwzF16wAAAEFgCAwAIAEFgAAAgsAACBBQAgsAAAEFgAAAILAEBgAQAg"
        "sAAABBYAgMACABBYAAAILACAsdLWLQAAXF6wAAAEFgDAKe7VCybxzZER2mbKXnafiyd3udNZfz3Xl37XzAW8MxdesAAABBYAgMAC"
        "ABBYAAAILAAAgQUAILAAABBYAAACCwBAYAEAILAAAAQWAIDAAgAQWAAACCwAgLHS1i0AAFxesAAABBYAwCnu1Qsm8c2REdpmyl7M"
        "BebCXHDWXHjBAgAQWAAAAgsAQGABACCwAAAEFgCAwAIAQGABAAgsAACBBQCAwAIAEFgAAAILAEBgAQAgsAAAxkpbtwAAcHnBAgAQ"
        "WAAAAgsAAIEFACCwAAAEFgAAAgsAQGABAAgsAAAEFgCAwAIAEFgAAAILAACBBQAgsAAABBYAAAILAEBgAQAILAAABBYAgMACABBY"
        "AAAILAAAgQUAILAAAAQWAAACCwBAYAEACCwAAAQWAIDAAgAQWAAACCwAAIEFACCwAAAQWAAAAgsAQGABAAgsAAAEFgCAwAIAEFgA"
        "AAgsAACBBQAgsAAAEFgAAAILAEBgAQAILAAABBYAgMACABBYAAAILAAAgQUAILAAABBYAACCBQDw1Gv+AwZFpM3uAIqvAAAAAElF"
        "TkSuQmCC"
    )
    def _get_ibm_logo_png(self) -> Optional[bytes]:
        if self._logo_png_cache: return self._logo_png_cache
        try:
            r = requests.get(
                self._LOGO_URL,
                headers={"User-Agent": "IBM-DocGen/2.0 (https://ibm.com; IBM Consulting) python-requests"},
                timeout=self.valves.web_image_fetch_timeout,
            )
            if r.status_code == 200 and r.content:
                raw = r.content
                im = Image.open(io.BytesIO(raw)).convert("RGBA")
                _, _, _, alpha = im.split()
                black = Image.new("RGBA", im.size, (0, 0, 0, 255))
                black.putalpha(alpha)
                out = io.BytesIO()
                black.save(out, format="PNG", optimize=True)
                self._logo_png_cache = out.getvalue()
                return self._logo_png_cache
            print(f"[DocGen] IBM logo remote fetch HTTP {r.status_code}, using offline fallback")
        except Exception as e:
            print(f"[DocGen] IBM logo remote fetch failed: {e} — using offline fallback")
        try:
            self._logo_png_cache = base64.b64decode(self._LOGO_FALLBACK_B64)
            return self._logo_png_cache
        except Exception as e:
            print(f"[DocGen] IBM logo fallback decode failed: {e}")
            return None
    def _get_ibm_logo_dims(self) -> tuple:
        png = self._get_ibm_logo_png()
        if not png: return (600, 240)
        try:
            im = Image.open(io.BytesIO(png))
            return im.size
        except Exception: return (600, 240)
    async def prepare_content_from_knowledge(
        self,
        query: str,
        collection_id: str,
        max_images: int = 10,
        __user__: Optional[dict] = None,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: OWUI Knowledge Collection."""
        try:
            await self._emit(__event_emitter__, "🔍 Retrieving from knowledge collection...")
            auth = self._auth_from_request(__request__)
            text_chunks = self._retrieve_text_from_collection(query, collection_id, auth)
            collection_files = self._list_collection_files(collection_id, auth)
            await self._emit(__event_emitter__,
                f"📚 {len(text_chunks)} chunks · {len(collection_files)} files")
            if not text_chunks and not collection_files: return json.dumps({"error": "No content in collection", "text_chunks": [], "images": []})
            await self._emit(__event_emitter__, "🖼️ Extracting images...")
            all_images = self._extract_from_collection(text_chunks, collection_files, auth)
            all_images = await self._vision_rank_async(query, all_images, auth)
            ranked = self._rank_images(query, all_images)[:max_images]
            for i, img in enumerate(ranked):
                img["display_id"] = f"IMG{i+1}"
                _IMAGE_STORE.pin(img.get("id",""), f"IMG{i+1}")
            await self._emit(__event_emitter__, f"✨ {len(ranked)} images ready", done=True)
            return self._package(query, text_chunks, ranked, source="knowledge_collection")
        except Exception: return json.dumps({"error": traceback.format_exc(), "text_chunks": [], "images": []})
    async def prepare_content_from_notes(
        self,
        query: str,
        note_ids: Optional[list[str]] = None,
        max_chunks: int = 12,
        __user__: Optional[dict] = None,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: OWUI Notes. Reads the user's notes stored in OWUI"""
        try:
            await self._emit(__event_emitter__, f"📝 Reading user notes for: {query[:60]}")
            _hb = self._start_heartbeat(__event_emitter__,
                eta_seconds=self._eta_for('prepare_content_from_attachments'),
                initial_phase='📝 Reading notes')
            auth = self._auth_from_request(__request__)
            user_id = (__user__ or {}).get("id") or ""
            text_chunks = self._read_owui_notes(user_id=user_id, note_ids=note_ids or None)
            if not text_chunks:
                await self._stop_heartbeat(_hb)
                return json.dumps({
                    "error": "No notes found for this user.",
                    "hint": "Create a note in OWUI's Notes panel, or pass note_ids explicitly.",
                    "text_chunks": [], "images": [],
                })
            text_chunks = self._rank_text(query, text_chunks)[:max_chunks]
            await self._emit(__event_emitter__,
                f"📚 {len(text_chunks)} note chunks ready", done=True)
            await self._stop_heartbeat(_hb)
            return self._package(query, text_chunks, [], source="owui_notes")
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            return json.dumps({"error": traceback.format_exc(),
                               "text_chunks": [], "images": []})
    def _read_owui_notes(self, user_id: str = "", note_ids: Optional[list] = None) -> list[dict]:
        import sqlite3 as _sql, os as _os, re as _re
        db_path = None
        try:
            import subprocess
            out = subprocess.check_output(
                "lsof -p $(pgrep -f 'open-webui serve'|head -1)|grep webui.db|awk '{print $NF}'|sort -u|head -1",
                shell=True, timeout=5,
            ).decode().strip()
            if out and _os.path.exists(out):
                db_path = out
        except Exception: pass
        if not db_path:
            db_path = "/Users/pradeepbasavarajappa/.local/share/uv/tools/open-webui/lib/python3.12/site-packages/open_webui/data/webui.db"
        if not _os.path.exists(db_path): return []
        try:
            con = _sql.connect(db_path, timeout=10)
            con.execute("PRAGMA busy_timeout=5000")
            if note_ids:
                placeholders = ",".join(["?"] * len(note_ids))
                rows = con.execute(
                    f"SELECT id, user_id, title, data FROM note WHERE id IN ({placeholders})",
                    note_ids,
                ).fetchall()
            elif user_id:
                rows = con.execute(
                    "SELECT id, user_id, title, data FROM note WHERE user_id=? ORDER BY updated_at DESC LIMIT 50",
                    (user_id,),
                ).fetchall()
            else:
                rows = con.execute(
                    "SELECT id, user_id, title, data FROM note ORDER BY updated_at DESC LIMIT 50"
                ).fetchall()
            con.close()
        except Exception as e:
            print(f"[DocGen] reading OWUI notes failed: {e}")
            return []
        chunks = []
        for nid, uid, title, data in rows:
            body = ""
            if isinstance(data, str):
                try:
                    parsed = json.loads(data) if data else {}
                except Exception:
                    parsed = {}
                body = (parsed.get("content") or parsed.get("text") or parsed.get("body")
                        or (data if len(data) < 50000 else ""))
            body = re.sub(r"<[^>]+>", " ", str(body))
            body = re.sub(r"\s+", " ", body).strip()
            if not body: continue
            for sub in self._chunk_text(body):
                chunks.append({
                    "content": (f"{title}. {sub}" if title else sub),
                    "source": f"owui-note:{nid}",
                    "url": f"#note/{nid}",
                    "page": 0, "doc_type": "note",
                })
        return chunks
    async def prepare_content_from_folder(
        self,
        query: str,
        folder_id: str,
        max_chunks: int = 12,
        __user__: Optional[dict] = None,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: OWUI Folder. Pulls content from everything grouped under"""
        try:
            if not folder_id: return json.dumps({
                    "error": "folder_id is required.",
                    "text_chunks": [], "images": [],
                })
            await self._emit(__event_emitter__, f"📁 Reading folder {folder_id[:8]}... for: {query[:60]}")
            _hb = self._start_heartbeat(__event_emitter__,
                eta_seconds=self._eta_for('prepare_content_from_attachments'),
                initial_phase='📁 Reading folder')
            user_id = (__user__ or {}).get("id") or ""
            text_chunks = self._read_owui_folder(folder_id=folder_id, user_id=user_id)
            if not text_chunks:
                await self._stop_heartbeat(_hb)
                return json.dumps({
                    "error": f"No content found in folder {folder_id}.",
                    "hint": "Make sure the folder has chats, notes, or files under it.",
                    "text_chunks": [], "images": [],
                })
            text_chunks = self._rank_text(query, text_chunks)[:max_chunks]
            await self._emit(__event_emitter__,
                f"📚 {len(text_chunks)} folder chunks ready", done=True)
            await self._stop_heartbeat(_hb)
            return self._package(query, text_chunks, [], source="owui_folder")
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            return json.dumps({"error": traceback.format_exc(),
                               "text_chunks": [], "images": []})
    def _read_owui_folder(self, folder_id: str = "", user_id: str = "") -> list[dict]:
        import sqlite3 as _sql, os as _os
        db_path = None
        try:
            import subprocess
            out = subprocess.check_output(
                "lsof -p $(pgrep -f 'open-webui serve'|head -1)|grep webui.db|awk '{print $NF}'|sort -u|head -1",
                shell=True, timeout=5,
            ).decode().strip()
            if out and _os.path.exists(out):
                db_path = out
        except Exception: pass
        if not db_path:
            db_path = "/Users/pradeepbasavarajappa/.local/share/uv/tools/open-webui/lib/python3.12/site-packages/open_webui/data/webui.db"
        if not _os.path.exists(db_path): return []
        chunks: list[dict] = []
        try:
            con = _sql.connect(db_path, timeout=10)
            con.execute("PRAGMA busy_timeout=5000")
            folder_name = ""
            ref_note_ids: list[str] = []
            ref_file_ids: list[str] = []
            try:
                if user_id:
                    row = con.execute(
                        "SELECT name, items, data FROM folder WHERE id=? AND user_id=?",
                        (folder_id, user_id),
                    ).fetchone()
                else:
                    row = con.execute(
                        "SELECT name, items, data FROM folder WHERE id=?",
                        (folder_id,),
                    ).fetchone()
            except Exception:
                row = None
            if row:
                folder_name = row[0] or ""
                for blob in (row[1], row[2]):
                    if not blob: continue
                    try:
                        parsed = json.loads(blob) if isinstance(blob, str) else blob
                    except Exception:
                        parsed = None
                    if isinstance(parsed, dict):
                        for nid in (parsed.get("note_ids") or parsed.get("notes") or []):
                            if isinstance(nid, str):
                                ref_note_ids.append(nid)
                        for fid in (parsed.get("file_ids") or parsed.get("files") or []):
                            if isinstance(fid, str):
                                ref_file_ids.append(fid)
                    elif isinstance(parsed, list):
                        for entry in parsed:
                            if isinstance(entry, dict):
                                t = (entry.get("type") or "").lower()
                                eid = entry.get("id") or ""
                                if t == "note" and eid:
                                    ref_note_ids.append(eid)
                                elif t == "file" and eid:
                                    ref_file_ids.append(eid)
            try:
                chat_rows = con.execute(
                    "SELECT id, title, chat FROM chat WHERE folder_id=? ORDER BY updated_at DESC LIMIT 50",
                    (folder_id,),
                ).fetchall()
            except Exception:
                chat_rows = []
            for cid, ctitle, cjson in chat_rows:
                try:
                    parsed = json.loads(cjson) if isinstance(cjson, str) else (cjson or {})
                except Exception:
                    parsed = {}
                messages = []
                if isinstance(parsed, dict):
                    messages = parsed.get("messages") or (parsed.get("history") or {}).get("messages") or []
                    if isinstance(messages, dict):
                        messages = list(messages.values())
                body_parts: list[str] = []
                for m in messages or []:
                    if not isinstance(m, dict): continue
                    role = m.get("role") or ""
                    content = m.get("content") or ""
                    if isinstance(content, list):
                        texts = [c.get("text", "") for c in content if isinstance(c, dict)]
                        content = " ".join(t for t in texts if t)
                    if isinstance(content, str) and content.strip():
                        body_parts.append(f"[{role}] {content.strip()}")
                body = re.sub(r"\s+", " ", " ".join(body_parts)).strip()
                if not body: continue
                for sub in self._chunk_text(body):
                    chunks.append({
                        "content": (f"{ctitle}. {sub}" if ctitle else sub),
                        "source": f"owui-chat:{cid}",
                        "url": f"#chat/{cid}",
                        "page": 0, "doc_type": "chat",
                    })
            if ref_note_ids:
                try:
                    note_chunks = self._read_owui_notes(user_id=user_id, note_ids=ref_note_ids)
                    chunks.extend(note_chunks)
                except Exception: pass
            if ref_file_ids:
                placeholders = ",".join(["?"] * len(ref_file_ids))
                try:
                    file_rows = con.execute(
                        f"SELECT id, filename, path, data FROM file WHERE id IN ({placeholders})",
                        ref_file_ids,
                    ).fetchall()
                except Exception:
                    file_rows = []
                for fid, fname, fpath, fdata in file_rows:
                    body = ""
                    if fdata:
                        try:
                            parsed = json.loads(fdata) if isinstance(fdata, str) else fdata
                            if isinstance(parsed, dict):
                                body = parsed.get("content") or parsed.get("text") or ""
                        except Exception: pass
                    if not body and fpath and _os.path.exists(fpath):
                        try:
                            if _os.path.getsize(fpath) < 2_000_000:
                                with open(fpath, "rb") as fh:
                                    raw = fh.read()
                                try:
                                    body = raw.decode("utf-8", errors="ignore")
                                except Exception:
                                    body = ""
                        except Exception: pass
                    body = re.sub(r"\s+", " ", str(body)).strip()
                    if not body: continue
                    for sub in self._chunk_text(body):
                        chunks.append({
                            "content": (f"{fname}. {sub}" if fname else sub),
                            "source": f"owui-file:{fid}",
                            "url": f"#file/{fid}",
                            "page": 0, "doc_type": "file",
                        })
            con.close()
        except Exception as e:
            print(f"[DocGen] reading OWUI folder failed: {e}")
        if folder_name and chunks:
            for c in chunks:
                c["content"] = f"[Folder: {folder_name}] {c['content']}"
        return chunks
    async def prepare_content_from_attachments(
        self,
        query: str,
        attachment_file_ids: Optional[list[str]] = None,
        max_images: int = 10,
        __user__: Optional[dict] = None,
        __request__=None,
        __files__=None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: Chat Attachments."""
        try:
            if not attachment_file_ids and __files__:
                auto_ids = []
                for f in __files__:
                    if isinstance(f, dict):
                        fid = f.get("id") or (f.get("file") or {}).get("id")
                        if fid:
                            auto_ids.append(fid)
                attachment_file_ids = auto_ids
            if not attachment_file_ids: return json.dumps({
                    "error": "No attachment_file_ids provided and no files attached to chat.",
                    "text_chunks": [], "images": []
                })
            await self._emit(__event_emitter__, f"📎 Processing {len(attachment_file_ids)} attachment(s)...")
            _hb = self._start_heartbeat(__event_emitter__, eta_seconds=self._eta_for('prepare_content_from_attachments'), initial_phase='📎 Extracting from attachments')
            auth = self._auth_from_request(__request__)
            text_chunks, _unused_imgs = await asyncio.to_thread(
                self._extract_attachments_parallel, attachment_file_ids, auth, 4
            )
            text_chunks = self._rank_text(query, text_chunks)[:self.valves.max_text_chunks]
            await self._emit(__event_emitter__, f"📚 {len(text_chunks)} text chunks ready", done=True)
            await self._stop_heartbeat(_hb)
            return self._package(query, text_chunks, [], source="chat_attachments")
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            return json.dumps({"error": traceback.format_exc(), "text_chunks": [], "images": []})
    async def prepare_content_from_web_search(
        self,
        query: str,
        num_text_results: int = 6,
        num_image_results: int = 10,
        __user__: Optional[dict] = None,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: Web Search (Google Programmable Search)."""
        try:
            await self._emit(__event_emitter__, f"🌐 Searching: {query}")
            _hb = self._start_heartbeat(__event_emitter__, eta_seconds=self._eta_for('prepare_content_from_web_search'), initial_phase='🌐 Searching the web')
            text_chunks = self._web_search_text(query, num_text_results)
            await self._emit(__event_emitter__, f"📚 {len(text_chunks)} web results", done=True)
            await self._stop_heartbeat(_hb)
            return self._package(query, text_chunks, [], source="web_search")
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            return json.dumps({"error": traceback.format_exc(), "text_chunks": [], "images": []})
    async def list_mcp_tools(
        self,
        server_id: Optional[str] = None,
        __user__: Optional[dict] = None,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """List tools advertised by configured MCP servers (ICA Context Forge, etc.)."""
        try:
            servers = self._load_mcp_servers(__request__)
            if not servers: return json.dumps({
                    "error": (
                        "No MCP servers discovered. In ICA/OWUI, ask your admin to add the "
                        "MCP server under Admin Settings → External Tools (it will then be auto-detected). "
                        "For local/dev use, set the mcp_servers_json valve."
                    ),
                    "servers": [],
                })
            out = []
            for srv in servers:
                if server_id and srv["id"] != server_id: continue
                await self._emit(__event_emitter__, f"🔌 Listing tools on {srv['name']}...")
                try:
                    client = _get_mcp_client(srv["url"], headers=self._mcp_headers(srv),
                                              transport=srv.get("transport"))
                    tools = client.list_tools()
                    allow = set(srv.get("tools", []) or [])
                    if allow:
                        tools = [t for t in tools if t.get("name") in allow]
                    out.append({
                        "server_id": srv["id"],
                        "server_name": srv["name"],
                        "url": srv["url"],
                        "tool_count": len(tools),
                        "tools": [
                            {
                                "name": t.get("name"),
                                "description": t.get("description", ""),
                                "input_schema": t.get("inputSchema", {}),
                                "has_ui_resource": bool(
                                    (t.get("_meta") or {}).get("ui", {}).get("resourceUri")
                                ),
                                "ui_resource_uri": (t.get("_meta") or {}).get("ui", {}).get("resourceUri"),
                            }
                            for t in tools
                        ],
                    })
                except Exception as e:
                    out.append({
                        "server_id": srv["id"],
                        "server_name": srv["name"],
                        "error": str(e),
                    })
            await self._emit(__event_emitter__, "✅ MCP tool listing complete", done=True)
            return json.dumps({"servers": out}, indent=2)
        except Exception: return json.dumps({"error": traceback.format_exc(), "servers": []})
    async def prepare_content_from_mcp(
        self,
        query: str,
        server_id: str,
        tool_name: str,
        tool_arguments: Optional[dict] = None,
        max_images: int = 10,
        __user__: Optional[dict] = None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: MCP server (ICA Context Forge or any Streamable-HTTP MCP)."""
        try:
            await self._emit(__event_emitter__, f"🔌 Calling {server_id}.{tool_name}...")
            servers = {s["id"]: s for s in self._load_mcp_servers()}
            if server_id not in servers: return json.dumps({
                    "error": f"MCP server '{server_id}' not configured.",
                    "available_server_ids": list(servers.keys()),
                    "text_chunks": [], "images": [],
                })
            srv = servers[server_id]
            client = _get_mcp_client(srv["url"], headers=self._mcp_headers(srv),
                                      transport=srv.get("transport"))
            tool_def = None
            try:
                all_tools = client.list_tools()
                tool_def = next((t for t in all_tools if t.get("name") == tool_name), None)
            except Exception as e:
                print(f"[MCP] tools/list failed on {server_id}: {e}")
            result = client.call_tool(tool_name, tool_arguments or {})
            await self._emit(__event_emitter__, "📦 Parsing MCP result...")
            src_meta = {
                "source": f"mcp://{server_id}/{tool_name}",
                "ext": ".mcp",
                "doc_type": "mcp",
                "mcp_server_id": server_id,
                "mcp_tool_name": tool_name,
            }
            text_chunks, all_images = self._parse_mcp_result(
                result=result,
                tool_def=tool_def,
                client=client,
                src_meta=src_meta,
                query=query,
            )
            await self._emit(__event_emitter__,
                f"📚 {len(text_chunks)} text chunks · {len(all_images)} raw images")
            if tool_def and not all_images:
                ui_uri = (tool_def.get("_meta") or {}).get("ui", {}).get("resourceUri")
                if ui_uri:
                    await self._emit(__event_emitter__, f"🎨 Fetching MCP UI resource {ui_uri}...")
                    try:
                        res = client.read_resource(ui_uri)
                        extra_text, extra_img = self._extract_from_mcp_resource_contents(
                            res.get("contents", []), src_meta, client
                        )
                        text_chunks.extend(extra_text)
                        all_images.extend(extra_img)
                    except Exception as e:
                        print(f"[MCP] UI resource fetch failed: {e}")
            ranked = self._rank_images(query, all_images)[:max_images]
            for i, img in enumerate(ranked):
                img["display_id"] = f"IMG{i+1}"
                _IMAGE_STORE.pin(img.get("id",""), f"IMG{i+1}")
            text_chunks = self._rank_text(query, text_chunks)[:self.valves.max_text_chunks]
            await self._emit(__event_emitter__,
                f"✨ {len(ranked)} images from MCP ready", done=True)
            return self._package(query, text_chunks, ranked, source=f"mcp:{server_id}/{tool_name}")
        except Exception: return json.dumps({"error": traceback.format_exc(),
                               "text_chunks": [], "images": []})
    async def prepare_content_mixed(
        self,
        query: str,
        knowledge_collection_id: Optional[str] = None,
        attachment_file_ids: Optional[list[str]] = None,
        mcp_calls: Optional[list[dict]] = None,
        web_search: bool = False,
        max_images: int = 10,
        __user__: Optional[dict] = None,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """Source mode: mix multiple sources in one call. Merges results so sections can draw"""
        try:
            auth = self._auth_from_request(__request__)
            all_text, all_images = [], []
            _hb = self._start_heartbeat(__event_emitter__, eta_seconds=self._eta_for('prepare_content_smart'), initial_phase='📥 Gathering content from all sources')
            if knowledge_collection_id:
                await self._emit(__event_emitter__, "🔍 Knowledge collection...")
                tc = self._retrieve_text_from_collection(query, knowledge_collection_id, auth)
                cf = self._list_collection_files(knowledge_collection_id, auth)
                all_text.extend(tc)
                all_images.extend(self._extract_from_collection(tc, cf, auth))
            if attachment_file_ids:
                await self._emit(__event_emitter__, f"📎 {len(attachment_file_ids)} attachment(s)...")
                tc, ti = await asyncio.to_thread(
                    self._extract_attachments_parallel, attachment_file_ids, auth, 4
                )
                all_text.extend(tc)
                all_images.extend(ti)
            if mcp_calls:
                servers = {s["id"]: s for s in self._load_mcp_servers()}
                for call in mcp_calls:
                    sid = call.get("server_id")
                    tname = call.get("tool_name")
                    args = call.get("arguments", {})
                    if sid not in servers: continue
                    await self._emit(__event_emitter__, f"🔌 MCP {sid}.{tname}...")
                    srv = servers[sid]
                    try:
                        client = _get_mcp_client(srv["url"], headers=self._mcp_headers(srv),
                                                  transport=srv.get("transport"))
                        tool_def = None
                        try:
                            all_tools = client.list_tools()
                            tool_def = next((t for t in all_tools if t.get("name") == tname), None)
                        except Exception: pass
                        result = client.call_tool(tname, args or {})
                        src_meta = {
                            "source": f"mcp://{sid}/{tname}",
                            "ext": ".mcp", "doc_type": "mcp",
                            "mcp_server_id": sid, "mcp_tool_name": tname,
                        }
                        tc, ti = self._parse_mcp_result(result, tool_def, client, src_meta, query)
                        all_text.extend(tc)
                        all_images.extend(ti)
                    except Exception as e:
                        print(f"[MCP mixed] {sid}.{tname} failed: {e}")
            if web_search:
                await self._emit(__event_emitter__, "🌐 Web search (text only)...")
                all_text.extend(self._web_search_text(query, 6))
            if self.valves.vision_rank_enabled and all_images:
                await self._emit(__event_emitter__, f"👁️ Vision-ranking {min(len(all_images), self.valves.vision_rank_max_images)} images...")
                all_images = await self._vision_rank_async(query, all_images, auth)
            ranked = self._rank_images(query, all_images)[:max_images]
            for i, img in enumerate(ranked):
                img["display_id"] = f"IMG{i+1}"
                _IMAGE_STORE.pin(img.get("id",""), f"IMG{i+1}")
            text_chunks = self._rank_text(query, all_text)[:self.valves.max_text_chunks]
            await self._emit(__event_emitter__,
                f"✨ Mixed mode: {len(text_chunks)} chunks, {len(ranked)} images", done=True)
            await self._stop_heartbeat(_hb)
            return self._package(query, text_chunks, ranked, source="mixed")
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            return json.dumps({"error": traceback.format_exc(),
                               "text_chunks": [], "images": []})
    async def prepare_content_auto(
        self,
        query: str,
        preferred_servers: Optional[list[str]] = None,
        max_tools_to_call: int = 3,
        max_images: int = 10,
        __user__: Optional[dict] = None,
        __event_emitter__=None,
    ) -> str:
        """Auto-routing MCP mode: the tool itself picks which MCP tools to call."""
        try:
            await self._emit(__event_emitter__, "🧭 Discovering MCP tools...")
            servers = self._load_mcp_servers()
            if preferred_servers:
                servers = [s for s in servers if s["id"] in preferred_servers]
            if not servers: return json.dumps({
                    "error": "No MCP servers configured (or none match preferred_servers).",
                    "text_chunks": [], "images": [],
                })
            catalog = self._discover_mcp_catalog(servers, __event_emitter__)
            if not catalog: return json.dumps({
                    "error": "Could not discover tools on any configured MCP server.",
                    "text_chunks": [], "images": [],
                })
            await self._emit(__event_emitter__,
                f"🧮 Ranking {len(catalog)} tools against the query...")
            ranked_tools = self._rank_mcp_tools(query, catalog)[:max_tools_to_call]
            if not ranked_tools: return json.dumps({
                    "error": "No MCP tools scored above the relevance floor for this query.",
                    "text_chunks": [], "images": [],
                })
            picks = [f"{t['server_id']}.{t['name']} ({t['_score']:.2f})" for t in ranked_tools]
            await self._emit(__event_emitter__, f"🎯 Picked: {', '.join(picks)}")
            all_text: list[dict] = []
            all_images: list[dict] = []
            invocation_log: list[dict] = []
            servers_by_id = {s["id"]: s for s in servers}
            for t in ranked_tools:
                sid = t["server_id"]
                srv = servers_by_id.get(sid)
                if not srv: continue
                args = self._auto_fill_tool_args(query, t.get("inputSchema") or {})
                await self._emit(__event_emitter__,
                    f"🔌 Calling {sid}.{t['name']} {args or '{}'}...")
                try:
                    client = _get_mcp_client(srv["url"], headers=self._mcp_headers(srv),
                                              transport=srv.get("transport"))
                    result = client.call_tool(t["name"], args)
                    src_meta = {
                        "source": f"mcp://{sid}/{t['name']}",
                        "ext": ".mcp", "doc_type": "mcp",
                        "mcp_server_id": sid, "mcp_tool_name": t["name"],
                    }
                    tc, ti = self._parse_mcp_result(result, t, client, src_meta, query)
                    all_text.extend(tc)
                    all_images.extend(ti)
                    invocation_log.append({
                        "server_id": sid, "tool_name": t["name"],
                        "arguments": args, "score": round(t["_score"], 3),
                        "text_chunks_yielded": len(tc), "images_yielded": len(ti),
                    })
                except Exception as e:
                    invocation_log.append({
                        "server_id": sid, "tool_name": t["name"],
                        "arguments": args, "error": str(e)[:300],
                    })
            ranked_imgs = self._rank_images(query, all_images)[:max_images]
            for i, img in enumerate(ranked_imgs):
                img["display_id"] = f"IMG{i+1}"
                _IMAGE_STORE.pin(img.get("id",""), f"IMG{i+1}")
            text_chunks = self._rank_text(query, all_text)[:self.valves.max_text_chunks]
            await self._emit(__event_emitter__,
                f"✨ Auto-route: {len(text_chunks)} chunks, {len(ranked_imgs)} images from "
                f"{len([e for e in invocation_log if 'error' not in e])}/{len(invocation_log)} tools",
                done=True)
            pkg = json.loads(self._package(query, text_chunks, ranked_imgs, source="mcp_auto"))
            pkg["mcp_invocations"] = invocation_log
            return json.dumps(pkg, indent=2)
        except Exception: return json.dumps({"error": traceback.format_exc(),
                               "text_chunks": [], "images": []})
    async def prepare_content_smart(
        self,
        query: str,
        knowledge_collection_id: Optional[str] = None,
        attachment_file_ids: Optional[list[str]] = None,
        note_ids: Optional[list[str]] = None,
        folder_id: Optional[str] = None,
        use_notes: bool = True,
        use_mcp_auto: bool = True,
        use_web_search: bool = False,
        max_mcp_tools: int = 3,
        max_images: int = 10,
        __user__: Optional[dict] = None,
        __request__=None,
        __files__=None,
        __event_emitter__=None,
    ) -> str:
        """One-call "do the right thing" mode. Tool figures out where to pull from:"""
        try:
            auth = self._auth_from_request(__request__)
            all_text: list[dict] = []
            all_images: list[dict] = []
            source_log: list[str] = []
            _hb = self._start_heartbeat(__event_emitter__, eta_seconds=self._eta_for('prepare_content_smart'), initial_phase='📥 Gathering content from all sources')
            if not attachment_file_ids and __files__:
                auto_ids = []
                for f in __files__:
                    if isinstance(f, dict):
                        fid = f.get("id") or (f.get("file") or {}).get("id")
                        if fid:
                            auto_ids.append(fid)
                if auto_ids:
                    attachment_file_ids = auto_ids
                    await self._emit(__event_emitter__,
                        f"📎 Auto-detected {len(auto_ids)} chat attachment(s)")
            if knowledge_collection_id:
                await self._emit(__event_emitter__, "🔍 Knowledge collection...")
                tc = self._retrieve_text_from_collection(query, knowledge_collection_id, auth)
                cf = self._list_collection_files(knowledge_collection_id, auth)
                all_text.extend(tc)
                all_images.extend(self._extract_from_collection(tc, cf, auth))
                source_log.append(f"knowledge:{knowledge_collection_id}")
            if attachment_file_ids:
                await self._emit(__event_emitter__, f"📎 {len(attachment_file_ids)} attachment(s)...")
                tc, ti = await asyncio.to_thread(
                    self._extract_attachments_parallel, attachment_file_ids, auth, 4
                )
                all_text.extend(tc)
                all_images.extend(ti)
                source_log.append(f"attachments:{len(attachment_file_ids)}")
            if use_notes or note_ids:
                user_id = (__user__ or {}).get("id") or ""
                note_chunks = self._read_owui_notes(user_id=user_id, note_ids=note_ids or None)
                if note_chunks:
                    await self._emit(__event_emitter__,
                        f"📝 Pulled {len(note_chunks)} note chunks")
                    all_text.extend(note_chunks)
                    source_log.append(f"notes:{len(note_chunks)}")
            if folder_id:
                user_id = (__user__ or {}).get("id") or ""
                folder_chunks = self._read_owui_folder(folder_id=folder_id, user_id=user_id)
                if folder_chunks:
                    await self._emit(__event_emitter__,
                        f"📁 Pulled {len(folder_chunks)} folder chunks")
                    all_text.extend(folder_chunks)
                    source_log.append(f"folder:{folder_id}:{len(folder_chunks)}")
            if use_mcp_auto:
                servers = self._load_mcp_servers()
                if servers:
                    catalog = self._discover_mcp_catalog(servers, __event_emitter__)
                    if catalog:
                        ranked_tools = self._rank_mcp_tools(query, catalog)[:max_mcp_tools]
                        servers_by_id = {s["id"]: s for s in servers}
                        for t in ranked_tools:
                            srv = servers_by_id.get(t["server_id"])
                            if not srv: continue
                            args = self._auto_fill_tool_args(query, t.get("inputSchema") or {})
                            await self._emit(__event_emitter__,
                                f"🔌 MCP {t['server_id']}.{t['name']}...")
                            try:
                                client = _get_mcp_client(srv["url"],
                                                          headers=self._mcp_headers(srv),
                                                          transport=srv.get("transport"))
                                result = client.call_tool(t["name"], args)
                                src_meta = {
                                    "source": f"mcp://{t['server_id']}/{t['name']}",
                                    "ext": ".mcp", "doc_type": "mcp",
                                    "mcp_server_id": t["server_id"],
                                    "mcp_tool_name": t["name"],
                                }
                                tc, ti = self._parse_mcp_result(result, t, client,
                                                                  src_meta, query)
                                all_text.extend(tc)
                                all_images.extend(ti)
                                source_log.append(f"mcp:{t['server_id']}.{t['name']}")
                            except Exception as e:
                                print(f"[smart] {t['server_id']}.{t['name']} failed: {e}")
            web_only_images = False
            if use_web_search or (not all_text and not all_images):
                await self._emit(__event_emitter__, "🌐 Web search (text only)...")
                web_texts = await asyncio.to_thread(self._web_search_text, query, 6)
                all_text.extend(web_texts)
                source_log.append("web_search")
            if not all_text and not all_images: return json.dumps({
                    "error": "No content retrieved from any source.",
                    "sources_tried": source_log,
                    "text_chunks": [], "images": [],
                })
            if (self.valves.vision_rank_enabled and all_images
                    and not (web_only_images and self.valves.skip_vision_rank_for_web)):
                await self._emit(__event_emitter__,
                    f"👁️ Vision-ranking {min(len(all_images), self.valves.vision_rank_max_images)} images...")
                all_images = await self._vision_rank_async(query, all_images, auth)
            elif web_only_images:
                await self._emit(__event_emitter__, "⚡ Skipping vision-rank (web-only results)")
            ranked_imgs = self._rank_images(query, all_images)[:max_images]
            for i, img in enumerate(ranked_imgs):
                img["display_id"] = f"IMG{i+1}"
                _IMAGE_STORE.pin(img.get("id",""), f"IMG{i+1}")
            text_chunks = self._rank_text(query, all_text)[:self.valves.max_text_chunks]
            await self._emit(__event_emitter__,
                f"✨ Smart: {len(text_chunks)} chunks, {len(ranked_imgs)} images "
                f"from {len(source_log)} sources", done=True)
            pkg = json.loads(self._package(query, text_chunks, ranked_imgs, source="smart"))
            pkg["sources_used"] = source_log
            await self._stop_heartbeat(_hb)
            return json.dumps(pkg, indent=2)
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            return json.dumps({"error": traceback.format_exc(),
                               "text_chunks": [], "images": []})
    async def assemble_document(
        self,
        session_id: str,
        format: Literal["docx", "pptx", "xlsx"],
        title: str,
        client_name: str,
        sections_json: str,
        workbook_json: Optional[str] = None,
        __user__: Optional[dict] = None,
        __event_emitter__=None,
        __files__=None,
        __request__=None,
    ):
        """Build and render the final DOCX, PPTX or XLSX inline in chat."""
        try:
            await self._emit(__event_emitter__, f"📝 Assembling {format.upper()}...")
            _hb = self._start_heartbeat(__event_emitter__, eta_seconds=self._eta_for('assemble_document'), initial_phase='📝 Assembling document')
            try:
                sections = json.loads(sections_json) if isinstance(sections_json, str) else sections_json
            except json.JSONDecodeError as e: return f"❌ sections_json is not valid JSON: {e}"
            if not isinstance(sections, list) or not sections: return "❌ sections_json must be a non-empty array"
            original_count = len(sections)
            if format == "pptx" and original_count > self.MAX_SLIDES_PPTX:
                sections = sections[: self.MAX_SLIDES_PPTX]
                await self._emit(__event_emitter__,
                    f"⚠️ PPTX capped at {self.MAX_SLIDES_PPTX} content slides "
                    f"(was {original_count}). Showing first {self.MAX_SLIDES_PPTX}.")
            elif format == "docx" and original_count > self.MAX_PAGES_DOCX:
                sections = sections[: self.MAX_PAGES_DOCX]
                await self._emit(__event_emitter__,
                    f"⚠️ DOCX capped at {self.MAX_PAGES_DOCX} content pages "
                    f"(was {original_count}). Showing first {self.MAX_PAGES_DOCX}.")
            resolved_sections = []
            missing_images = []
            for s in sections:
                resolved = dict(s)
                svg_str = s.get("svg")
                if svg_str and not resolved.get("_img_bytes"):
                    png = _svg_to_png_bytes(svg_str, output_width=1600)
                    if png:
                        resolved["_img_bytes"] = png
                        resolved["_img_width"] = 1600
                        resolved["_img_height"] = 1000
                        resolved["_img_source"] = "generated:svg"
                    else:
                        await self._emit(__event_emitter__,
                            "⚠️ SVG present but cairosvg unavailable — diagram not rasterized")
                if s.get("image_id") and not resolved.get("_img_bytes"):
                    img_bytes = _IMAGE_STORE.get_bytes(s["image_id"])
                    img_meta = _IMAGE_STORE.get_metadata(s["image_id"])
                    if not img_bytes:
                        img_bytes, img_meta = _IMAGE_STORE.get_by_display_id(s["image_id"])
                    if img_bytes and img_meta:
                        resolved["_img_bytes"] = img_bytes
                        resolved["_img_width"] = img_meta.get("width", 1200)
                        resolved["_img_height"] = img_meta.get("height", 800)
                        resolved["_img_source"] = img_meta.get("source", "")
                    else:
                        missing_images.append(s["image_id"])
                resolved_sections.append(resolved)
            if missing_images:
                await self._emit(__event_emitter__,
                    f"⚠️ {len(missing_images)} image(s) expired/missing — continuing without them")
            stripped = []
            for s in resolved_sections:
                if isinstance(s, dict):
                    clean = {k: v for k, v in s.items()
                             if k not in ("image_id", "svg",
                                          "_img_width", "_img_height", "_img_source",
                                          "image_caption", "image_hint")}
                    stripped.append(clean)
                else:
                    stripped.append(s)
            resolved_sections = stripped
            if format in ("docx", "pptx", "xlsx"):
                resolved_sections = self._autoinject_charts(resolved_sections)
            resolved_sections = self._enforce_content_caps(resolved_sections, format)
            if format in ("docx", "pptx") and __files__:
                auto_ids = []
                for f in __files__:
                    if isinstance(f, dict):
                        fid = f.get("id") or (f.get("file") or {}).get("id")
                        if fid: auto_ids.append(fid)
                if auto_ids:
                    auth = self._auth_from_request(__request__)
                    resolved_sections = await self._kb_enrich_sections(
                        resolved_sections, auto_ids, auth, __event_emitter__)
            if format == "docx":
                html_response = self._build_and_render_docx(
                    session_id, title, client_name, resolved_sections, __event_emitter__
                )
            elif format == "xlsx":
                wb_spec = None
                if workbook_json:
                    try:
                        wb_spec = json.loads(workbook_json) if isinstance(workbook_json, str) else workbook_json
                    except json.JSONDecodeError as e: return f"❌ workbook_json is not valid JSON: {e}"
                html_response = self._build_and_render_xlsx(
                    session_id, title, client_name, resolved_sections, wb_spec, __event_emitter__
                )
            else:
                html_response = self._build_and_render_pptx(
                    session_id, title, client_name, resolved_sections, __event_emitter__
                )
            await self._emit(__event_emitter__, f"✅ {format.upper()} ready", done=True)
            await self._stop_heartbeat(_hb)
            return html_response
        except Exception:
            try: await self._stop_heartbeat(_hb)
            except Exception: pass
            _tb = traceback.format_exc()
            return "❌ Assembly failed:\n" + _tb
    _IMG_URL_RE = re.compile(
        r"https?://[^\s\"'<>)\}]+?\.(?:png|jpe?g|webp|gif|bmp|tiff?)(?:\?[^\s\"'<>)\}]*)?",
        re.IGNORECASE,
    )
    _PRESIGNED_URL_RE = re.compile(
        r"https?://[^\s\"'<>)\}]+?\?[^\s\"'<>)\}]*(?:AWSAccessKeyId|X-Amz-Signature|Signature=)[^\s\"'<>)\}]*",
        re.IGNORECASE,
    )
    def _extract_images_from_mcp_urls(self, result: dict, src_meta: dict, prompt: str) -> list[dict]:
        out: list[dict] = []
        texts: list[str] = []
        for block in (result.get("content") or []):
            if block.get("type") == "text":
                t = block.get("text") or ""
                if t: texts.append(t)
        sc = result.get("structuredContent")
        if sc:
            try:
                texts.append(json.dumps(sc))
            except Exception: pass
        candidates: list[str] = []
        seen: set = set()
        for txt in texts:
            for m in self._IMG_URL_RE.finditer(txt):
                u = m.group(0)
                if u not in seen:
                    seen.add(u); candidates.append(u)
            for m in self._PRESIGNED_URL_RE.finditer(txt):
                u = m.group(0)
                if u not in seen:
                    seen.add(u); candidates.append(u)
        if not candidates: return out
        for u in candidates[:3]:
            rec = self._ingest_remote_image(u, {
                "url": u,
                "title": src_meta.get("mcp_tool_name", "generated"),
                "snippet": prompt,
                "source_page": f"mcp://{src_meta.get('source','')}",
            })
            if rec:
                out.append(rec)
                break
        return out
    _SOURCE_COOLDOWN_S = 30.0
    _source_breakers: dict = {}
    _breakers_lock = threading.Lock()
    @classmethod
    def _breaker_is_open(cls, source_name: str) -> bool:
        with cls._breakers_lock:
            until = cls._source_breakers.get(source_name, 0)
            return time.time() < until
    @classmethod
    def _breaker_trip(cls, source_name: str, cooldown: Optional[float] = None):
        with cls._breakers_lock:
            cls._source_breakers[source_name] = time.time() + (cooldown or cls._SOURCE_COOLDOWN_S)
    _BLOCKED_IMAGE_HOSTS = frozenset([
        "emvigotech.com",
        "artificall.com",
        "sp-uploads.s3.amazonaws.com",
        "alamy.com", "c8.alamy.com",
        "shutterstock.com", "istockphoto.com", "gettyimages.com",
        "adobe.com", "stock.adobe.com",
        "wallpaperaccess.com",
        "logodix.com",
        "pinterest.com", "pinimg.com",
    ])
    def _prefilter_candidates(self, candidates: list[dict]) -> list[dict]:
        out = []
        for c in candidates:
            url = (c.get("url") or "").lower()
            host = urlparse(url).netloc.lower().lstrip("www.")
            blocked = any(host == h or host.endswith("." + h) for h in self._BLOCKED_IMAGE_HOSTS)
            if not blocked:
                out.append(c)
        return out
    def _load_mcp_servers(self) -> list[dict]:
        raw = (self.valves.mcp_servers_json or "").strip()
        if not raw: return []
        try:
            data = json.loads(raw)
            if isinstance(data, dict):
                if "servers" in data and isinstance(data["servers"], dict):
                    flat = []
                    for sid, entry in data["servers"].items():
                        if not isinstance(entry, dict): continue
                        merged = dict(entry)
                        merged["id"] = sid
                        flat.append(merged)
                    data = flat
                elif "mcpServers" in data and isinstance(data["mcpServers"], dict):
                    flat = []
                    for sid, entry in data["mcpServers"].items():
                        if not isinstance(entry, dict): continue
                        merged = dict(entry)
                        merged["id"] = sid
                        flat.append(merged)
                    data = flat
                else:
                    data = [data]
            if not isinstance(data, list): return []
            out = []
            for entry in data:
                if not isinstance(entry, dict): continue
                if not entry.get("id") or not entry.get("url"): continue
                t = (entry.get("type") or "").lower()
                if t in ("sse",):
                    transport = _MCPClient.TRANSPORT_SSE
                elif t in ("streamable-http", "streamable_http", "http", "mcp"):
                    transport = _MCPClient.TRANSPORT_STREAMABLE
                else:
                    transport = None
                headers = dict(entry.get("headers") or {})
                auth = (entry.get("auth_header") or "").strip()
                if auth and "Authorization" not in headers:
                    headers["Authorization"] = auth
                out.append({
                    "id": entry["id"],
                    "name": entry.get("name", entry["id"]),
                    "url": entry["url"],
                    "transport": transport,
                    "headers": headers,
                    "tools": entry.get("tools", []) or [],
                })
            return out
        except Exception as e:
            print(f"[MCP] mcp_servers_json parse failed: {e}")
            return []
    def _mcp_headers(self, srv: dict) -> dict:
        return dict(srv.get("headers") or {})
    def _discover_mcp_catalog(self, servers: list[dict], emitter=None) -> list[dict]:
        now = time.time()
        if not hasattr(self, "_mcp_catalog_cache"):
            self._mcp_catalog_cache: dict = {}
        ttl = getattr(self.valves, "mcp_catalog_ttl_seconds", 600)
        catalog: list[dict] = []
        for srv in servers:
            cache_key = srv["url"].rstrip("/")
            cached = self._mcp_catalog_cache.get(cache_key)
            if cached and now - cached["at"] < ttl:
                catalog.extend(cached["tools"])
                continue
            try:
                client = _get_mcp_client(srv["url"], headers=self._mcp_headers(srv),
                                          transport=srv.get("transport"))
                tools = client.list_tools()
                allow = set(srv.get("tools", []) or [])
                if allow:
                    tools = [t for t in tools if t.get("name") in allow]
                flat = []
                for t in tools:
                    flat.append({
                        "server_id": srv["id"],
                        "server_name": srv["name"],
                        "name": t.get("name", ""),
                        "description": t.get("description", ""),
                        "inputSchema": t.get("inputSchema", {}),
                        "_meta": t.get("_meta", {}),
                    })
                self._mcp_catalog_cache[cache_key] = {"at": now, "tools": flat}
                catalog.extend(flat)
            except Exception as e:
                print(f"[MCP auto] catalog fetch failed on {srv['id']}: {e}")
        return catalog
    def _rank_mcp_tools(self, query: str, catalog: list[dict]) -> list[dict]:
        q_tokens = set(re.findall(r"\w{3,}", query.lower()))
        if not q_tokens: return []
        action_verbs = {
            "search": {"search", "find", "lookup", "query", "look"},
            "get":    {"get", "fetch", "retrieve", "read", "show", "display"},
            "list":   {"list", "enumerate", "browse"},
            "analyze":{"analyze", "analyse", "evaluate", "assess", "review", "score"},
            "generate":{"generate", "create", "produce", "build", "draft"},
            "summarize":{"summarize","summarise","summary","brief"},
        }
        query_intent = set()
        for intent, synonyms in action_verbs.items():
            if synonyms & q_tokens:
                query_intent.add(intent)
        scored = []
        for t in catalog:
            name = (t.get("name") or "").lower()
            desc = (t.get("description") or "").lower()
            name_tokens = set(re.findall(r"\w{3,}", name))
            desc_tokens = set(re.findall(r"\w{3,}", desc))
            desc_overlap = len(q_tokens & desc_tokens)
            name_overlap = len(q_tokens & name_tokens) * 2.0
            intent_bonus = 0.0
            for intent, synonyms in action_verbs.items():
                if intent not in query_intent: continue
                if synonyms & (name_tokens | desc_tokens):
                    intent_bonus += 1.5
            schema = t.get("inputSchema") or {}
            required = schema.get("required") or []
            props = (schema.get("properties") or {})
            if not required:
                schema_bonus = 0.8
            else:
                can_fill = sum(1 for r in required
                               if self._can_infer_arg(query, r, props.get(r, {})))
                schema_bonus = (can_fill / len(required)) * 1.2
            score = desc_overlap + name_overlap + intent_bonus + schema_bonus
            if score <= 0: continue
            t_copy = dict(t)
            t_copy["_score"] = score
            scored.append(t_copy)
        scored.sort(key=lambda x: x["_score"], reverse=True)
        return scored
    def _can_infer_arg(self, query: str, arg_name: str, arg_schema: dict) -> bool:
        arg_type = (arg_schema.get("type") or "string").lower()
        arg_name_l = arg_name.lower()
        if arg_type == "string":
            if any(k in arg_name_l for k in (
                "query", "question", "text", "content", "topic",
                "keyword", "search", "term", "prompt", "input",
                "url", "link", "title", "name"
            )):
                return True
            if "id" in arg_name_l: return bool(re.search(r"\b[A-Z0-9]{3,}[-_]?\d+\b|\b[a-f0-9]{8,}\b", query))
            return True
        if arg_type in ("integer", "number"): return bool(re.search(r"\b\d+\b", query))
        if arg_type == "boolean": return True
        if arg_type == "array": return True
        return False
    def _auto_fill_tool_args(self, query: str, input_schema: dict) -> dict:
        if not input_schema or not isinstance(input_schema, dict): return {}
        props = input_schema.get("properties") or {}
        required = set(input_schema.get("required") or [])
        if not props: return {}
        out: dict = {}
        for arg_name, schema in props.items():
            arg_type = (schema.get("type") or "string").lower()
            arg_name_l = arg_name.lower()
            default = schema.get("default")
            is_required = arg_name in required
            if not is_required and default is not None:
                out[arg_name] = default
                continue
            if arg_type == "string":
                if any(k in arg_name_l for k in ("url", "link", "href")):
                    m = re.search(r"https?://\S+", query)
                    if m: out[arg_name] = m.group(0); continue
                    if is_required: out[arg_name] = query
                    continue
                if re.search(r"(_|^)id(_|$)", arg_name_l) or arg_name_l.endswith("_id"):
                    m = re.search(r"\b([A-Z]+[-_]?\d{2,}|[a-f0-9]{8,})\b", query)
                    if m: out[arg_name] = m.group(0); continue
                    if is_required: out[arg_name] = ""
                    continue
                if any(k in arg_name_l for k in ("date", "from", "to", "since", "until", "time")):
                    m = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", query)
                    if m: out[arg_name] = m.group(1); continue
                    if is_required:
                        out[arg_name] = time.strftime("%Y-%m-%d")
                    continue
                if "email" in arg_name_l:
                    m = re.search(r"\b[\w.+-]+@[\w-]+\.\w+\b", query)
                    if m: out[arg_name] = m.group(0); continue
                    if is_required: out[arg_name] = ""
                    continue
                if is_required or any(k in arg_name_l for k in (
                    "query", "question", "text", "content", "topic",
                    "keyword", "search", "term", "prompt", "input", "title", "name"
                )):
                    out[arg_name] = query
                continue
            if arg_type in ("integer", "number"):
                m = re.search(r"\b(\d+)\b", query)
                if m:
                    val = int(m.group(1))
                    out[arg_name] = val if arg_type == "integer" else float(val)
                elif default is not None:
                    out[arg_name] = default
                elif is_required:
                    low_name = arg_name_l
                    if any(k in low_name for k in ("limit", "max", "top", "count", "k")):
                        out[arg_name] = 10
                    else:
                        out[arg_name] = 1
                continue
            if arg_type == "boolean":
                if default is not None:
                    out[arg_name] = default
                elif is_required:
                    out[arg_name] = False
                continue
            if arg_type == "array":
                items = schema.get("items", {}) or {}
                item_type = (items.get("type") or "string").lower()
                if item_type == "string":
                    if "," in query:
                        out[arg_name] = [p.strip() for p in query.split(",") if p.strip()]
                    else:
                        out[arg_name] = [query] if is_required else []
                else:
                    out[arg_name] = []
                continue
            if is_required:
                out[arg_name] = {}
        return out
    def _parse_mcp_result(self, result: dict, tool_def: Optional[dict],
                           client: Any, src_meta: dict,
                           query: str) -> tuple[list[dict], list[dict]]:
        text_chunks: list[dict] = []
        images: list[dict] = []
        image_budget = self.valves.mcp_max_image_extract_per_call
        for block in (result.get("content") or []):
            btype = block.get("type")
            if btype == "text":
                txt = (block.get("text") or "").strip()
                if txt:
                    for sub in self._chunk_text(txt):
                        text_chunks.append({
                            "content": sub, "source": src_meta["source"],
                            "page": 0, "doc_type": "mcp",
                        })
            elif btype == "image":
                if len(images) >= image_budget: continue
                rec = self._ingest_mcp_image_block(block, src_meta,
                                                     caption=src_meta.get("mcp_tool_name", ""))
                if rec:
                    images.append(rec)
            elif btype == "resource":
                res = block.get("resource") or {}
                tc, ti = self._extract_from_mcp_resource_contents([res], src_meta, client)
                text_chunks.extend(tc)
                for r in ti:
                    if len(images) < image_budget:
                        images.append(r)
        sc = result.get("structuredContent")
        if sc:
            try:
                sc_text = json.dumps(sc, indent=2)[:4000]
                text_chunks.append({
                    "content": f"Structured output:\n{sc_text}",
                    "source": src_meta["source"],
                    "page": 0, "doc_type": "mcp",
                })
            except Exception: pass
        if tool_def and not images:
            ui_uri = (tool_def.get("_meta") or {}).get("ui", {}).get("resourceUri")
            if ui_uri:
                try:
                    res = client.read_resource(ui_uri)
                    tc, ti = self._extract_from_mcp_resource_contents(
                        res.get("contents", []), src_meta, client
                    )
                    text_chunks.extend(tc)
                    for r in ti:
                        if len(images) < image_budget:
                            images.append(r)
                except Exception as e:
                    print(f"[MCP] UI resource fetch in _parse_mcp_result failed: {e}")
        if len(images) < image_budget:
            url_imgs = self._extract_images_from_mcp_urls(result, src_meta, query)
            for r in url_imgs:
                if len(images) < image_budget:
                    images.append(r)
        return text_chunks, images
    def _extract_from_mcp_resource_contents(
        self, contents: list[dict], src_meta: dict, client: Any
    ) -> tuple[list[dict], list[dict]]:
        text_chunks: list[dict] = []
        images: list[dict] = []
        for part in (contents or []):
            mime = (part.get("mimeType") or "").lower()
            uri = part.get("uri", "")
            is_ui = uri.startswith("ui://") or "mcp-app" in mime or "skybridge" in mime
            if part.get("blob"):
                if mime.startswith("image/"):
                    try:
                        img_bytes = base64.b64decode(part["blob"])
                        rec = self._ingest_raw_image_bytes(
                            img_bytes, src_meta,
                            caption=self._humanize(uri.split("/")[-1] or "mcp_resource"),
                            context=uri,
                            location=uri or "mcp resource",
                        )
                        if rec:
                            images.append(rec)
                    except Exception as e:
                        print(f"[MCP] blob image decode failed: {e}")
            elif part.get("text") is not None:
                txt = part["text"]
                if mime.startswith("text/html") or is_ui:
                    tchunks, iimgs = self._extract_from_html(txt, src_meta, base_uri=uri)
                    text_chunks.extend(tchunks)
                    images.extend(iimgs)
                elif mime == "application/json":
                    try:
                        parsed = json.loads(txt)
                        pretty = json.dumps(parsed, indent=2)[:4000]
                        text_chunks.append({
                            "content": pretty, "source": src_meta["source"],
                            "page": 0, "doc_type": "mcp",
                        })
                    except Exception:
                        text_chunks.append({
                            "content": txt[:4000], "source": src_meta["source"],
                            "page": 0, "doc_type": "mcp",
                        })
                else:
                    for sub in self._chunk_text(txt):
                        text_chunks.append({
                            "content": sub, "source": src_meta["source"],
                            "page": 0, "doc_type": "mcp",
                        })
        return text_chunks, images
    def _extract_from_html(self, html: str, src_meta: dict,
                            base_uri: str = "") -> tuple[list[dict], list[dict]]:
        text_chunks: list[dict] = []
        images: list[dict] = []
        for m in re.finditer(r'<img[^>]+src=["\']([^"\']+)["\'][^>]*>', html, re.I):
            src = m.group(1).strip()
            if not src: continue
            alt_m = re.search(r'alt=["\']([^"\']*)["\']', m.group(0), re.I)
            alt = alt_m.group(1) if alt_m else ""
            rec = self._ingest_html_image_src(src, src_meta, caption=alt, context=base_uri)
            if rec:
                images.append(rec)
        for m in re.finditer(r'<svg[\s>].*?</svg>', html, re.I | re.S):
            svg_text = m.group(0)
            try:
                if HAS_CAIROSVG:
                    png_bytes = cairosvg.svg2png(bytestring=svg_text.encode("utf-8"),
                                                  output_width=1200)
                    rec = self._ingest_raw_image_bytes(
                        png_bytes, src_meta,
                        caption="inline SVG",
                        context=svg_text[:400],
                        location=base_uri or "inline SVG",
                    )
                    if rec:
                        images.append(rec)
            except Exception as e:
                print(f"[MCP] inline SVG rasterize failed: {e}")
        stripped = re.sub(r'<script.*?</script>', ' ', html, flags=re.I | re.S)
        stripped = re.sub(r'<style.*?</style>', ' ', stripped, flags=re.I | re.S)
        stripped = re.sub(r'<[^>]+>', ' ', stripped)
        stripped = re.sub(r'\s+', ' ', stripped).strip()
        if stripped:
            for sub in self._chunk_text(stripped):
                text_chunks.append({
                    "content": sub, "source": src_meta["source"],
                    "page": 0, "doc_type": "mcp",
                })
        return text_chunks, images
    def _ingest_mcp_image_block(self, block: dict, src_meta: dict,
                                  caption: str = "") -> Optional[dict]:
        data = block.get("data")
        if not data: return None
        try:
            img_bytes = base64.b64decode(data)
        except Exception: return None
        return self._ingest_raw_image_bytes(
            img_bytes, src_meta,
            caption=caption or "mcp image",
            context=src_meta.get("mcp_tool_name", ""),
            location=src_meta.get("source", "mcp"),
        )
    def _ingest_raw_image_bytes(self, img_bytes: bytes, src_meta: dict,
                                  caption: str, context: str,
                                  location: str) -> Optional[dict]:
        try:
            pil = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        except Exception: return None
        if not self._passes_filter(pil): return None
        return self._store_image(pil, src_meta, caption, context, location,
                                  tag=src_meta.get("mcp_tool_name", "mcp"))
    def _ingest_html_image_src(self, src: str, src_meta: dict,
                                 caption: str, context: str) -> Optional[dict]:
        src = src.strip()
        if src.startswith("data:"):
            try:
                _, b64_part = src.split(",", 1)
                img_bytes = base64.b64decode(b64_part)
            except Exception: return None
            return self._ingest_raw_image_bytes(
                img_bytes, src_meta, caption=caption or "inline data URI",
                context=context, location=src[:60] + "...",
            )
        if src.startswith(("http://", "https://")):
            try:
                r = requests.get(src, timeout=self.valves.web_image_fetch_timeout,
                                 headers={"User-Agent": "Mozilla/5.0 (ibm-docgen)"})
                r.raise_for_status()
                if not r.content: return None
                return self._ingest_raw_image_bytes(
                    r.content, src_meta,
                    caption=caption or urlparse(src).path.rsplit("/", 1)[-1],
                    context=context or src,
                    location=f"url: {urlparse(src).netloc}",
                )
            except Exception as e:
                print(f"[MCP] HTML image fetch failed for {src}: {e}")
                return None
        return None
    def _auth_from_request(self, request) -> dict:
        headers = {}
        if request and hasattr(request, "headers"):
            auth = request.headers.get("authorization")
            if auth:
                headers["Authorization"] = auth
        return headers
    def _retrieve_text_from_collection(self, query: str, collection_id: str, auth: dict) -> list[dict]:
        url = f"{self.valves.owui_base_url}/api/v1/retrieval/query/collection"
        payload = {
            "collection_names": [collection_id],
            "query": query,
            "k": self.valves.max_text_chunks,
        }
        try:
            r = requests.post(url, json=payload, headers=auth, timeout=self.valves.request_timeout)
            r.raise_for_status()
            data = r.json()
            docs = data.get("documents", [[]])[0] if data.get("documents") else []
            metas = data.get("metadatas", [[]])[0] if data.get("metadatas") else []
            return [
                {
                    "content": d,
                    "source": m.get("source") or m.get("name") or m.get("file_id", "unknown"),
                    "file_id": m.get("file_id"),
                    "page": m.get("page", 0),
                    "doc_type": self._classify(m.get("source", "")),
                }
                for d, m in zip(docs, metas)
            ]
        except Exception as e:
            print(f"[DocGen] Retrieval failed: {e}")
            return []
    def _list_collection_files(self, collection_id: str, auth: dict) -> list[dict]:
        url = f"{self.valves.owui_base_url}/api/v1/knowledge/{collection_id}"
        try:
            r = requests.get(url, headers=auth, timeout=self.valves.request_timeout)
            r.raise_for_status()
            data = r.json()
            files = data.get("files", []) or data.get("data", {}).get("files", [])
            out = []
            for f in files:
                fid = f.get("id") or f.get("file_id")
                name = (
                    f.get("meta", {}).get("name")
                    or f.get("name")
                    or f.get("filename", "")
                )
                if fid and name:
                    out.append({
                        "file_id": fid,
                        "source": name,
                        "ext": self._ext(name),
                        "doc_type": self._classify(name),
                    })
            return out
        except Exception as e:
            print(f"[DocGen] List collection failed: {e}")
            return []
    def _local_upload_path(self, file_id: str) -> Optional[str]:
        try:
            import os, glob
            roots = []
            try:
                import open_webui
                pkg_dir = os.path.dirname(open_webui.__file__)
                roots.append(os.path.join(pkg_dir, "data", "uploads"))
            except Exception: pass
            roots.append(os.path.expanduser("~/.local/share/uv/tools/open-webui/lib/python3.12/site-packages/open_webui/data/uploads"))
            for root in roots:
                if not root or not os.path.isdir(root): continue
                hits = glob.glob(os.path.join(root, f"{file_id}_*"))
                if hits: return hits[0]
            return None
        except Exception as e:
            print(f"[DocGen] local_upload_path failed for {file_id}: {e}")
            return None
    def _fetch_file_metadata(self, file_id: str, auth: dict) -> Optional[dict]:
        local = self._local_upload_path(file_id)
        if local:
            import os
            fname = os.path.basename(local)
            if fname.startswith(f"{file_id}_"):
                fname = fname[len(file_id) + 1:]
            return {"name": fname, "id": file_id}
        url = f"{self.valves.owui_base_url}/api/v1/files/{file_id}"
        try:
            r = requests.get(url, headers=auth, timeout=self.valves.request_timeout)
            r.raise_for_status()
            data = r.json()
            return {
                "name": data.get("filename") or data.get("meta", {}).get("name") or file_id,
                "id": file_id,
            }
        except Exception as e:
            print(f"[DocGen] File metadata fetch failed for {file_id}: {e}")
            return None
    def _fetch_file_bytes(self, file_id: str, auth: dict) -> Optional[bytes]:
        local = self._local_upload_path(file_id)
        if local:
            try:
                with open(local, "rb") as fh: return fh.read()
            except Exception as e:
                print(f"[DocGen] local file read failed for {file_id}: {e}")
        url = f"{self.valves.owui_base_url}/api/v1/files/{file_id}/content"
        try:
            r = requests.get(url, headers=auth, timeout=self.valves.request_timeout)
            r.raise_for_status()
            return r.content
        except Exception as e:
            print(f"[DocGen] File bytes fetch failed for {file_id}: {e}")
            return None
    def _extract_one_attachment(self, fid: str, auth: dict) -> tuple[list[dict], list[dict]]:
        try:
            meta = self._fetch_file_metadata(fid, auth)
            if not meta: return [], []
            fbytes = self._fetch_file_bytes(fid, auth)
            if not fbytes: return [], []
            cached = _EXTRACT_CACHE.get(fbytes)
            if cached is not None: return cached
            src = {
                "file_id": fid,
                "source": meta.get("name", fid),
                "ext": self._ext(meta.get("name", "")),
                "doc_type": self._classify(meta.get("name", "")),
            }
            text_chunks = self._extract_text_from_bytes(fbytes, src)
            images = self._extract_images_from_bytes(fbytes, src)
            _EXTRACT_CACHE.put(fbytes, text_chunks, images)
            return text_chunks, images
        except Exception as e:
            print(f"[DocGen] _extract_one_attachment({fid}) failed: {e}")
            return [], []
    def _extract_attachments_parallel(self, file_ids: list[str], auth: dict,
                                      max_workers: int = 4) -> tuple[list[dict], list[dict]]:
        all_text, all_images = [], []
        if not file_ids: return all_text, all_images
        if len(file_ids) == 1: return self._extract_one_attachment(file_ids[0], auth)
        with ThreadPoolExecutor(max_workers=min(max_workers, len(file_ids))) as ex:
            futures = {ex.submit(self._extract_one_attachment, fid, auth): fid
                       for fid in file_ids}
            for fut in as_completed(futures):
                try:
                    tc, ti = fut.result()
                    all_text.extend(tc)
                    all_images.extend(ti)
                except Exception as e:
                    print(f"[DocGen] parallel extract failed for {futures[fut]}: {e}")
        return all_text, all_images
    def _ingest_images_parallel(self, candidates: list[dict], max_workers: int = 6) -> list[dict]:
        out: list[dict] = []
        if not candidates: return out
        with ThreadPoolExecutor(max_workers=min(max_workers, len(candidates))) as ex:
            futures = [ex.submit(self._ingest_remote_image, c["url"], c) for c in candidates]
            for fut in as_completed(futures):
                try:
                    rec = fut.result()
                    if rec:
                        out.append(rec)
                except Exception as e:
                    print(f"[DocGen] parallel ingest failed: {e}")
        return out
    def _extract_from_collection(self, text_chunks, collection_files, auth) -> list[dict]:
        all_images = []
        referenced_names = {c.get("source") for c in text_chunks}
        for f in collection_files:
            is_ref = f["source"] in referenced_names
            is_standalone = f["ext"] in ALL_IMG_EXT
            is_doc = f["ext"] in ALL_DOC_EXT
            if (is_ref and is_doc) or is_standalone:
                fbytes = self._fetch_file_bytes(f["file_id"], auth)
                if fbytes:
                    all_images.extend(self._extract_images_from_bytes(fbytes, f))
        return all_images
    def _extract_text_from_bytes(self, file_bytes: bytes, src: dict) -> list[dict]:
        ext = src["ext"]
        chunks = []
        try:
            if ext in PDF_EXT:
                if not HAS_FITZ:
                    print("[DocGen] PDF text extraction skipped — PyMuPDF not installed in this environment.")
                    return []
                doc = fitz.open(stream=file_bytes, filetype="pdf")
                try:
                    for page_num, page in enumerate(doc):
                        text = page.get_text("text").strip()
                        if text:
                            for sub in self._chunk_text(text):
                                chunks.append({
                                    "content": sub, "source": src["source"],
                                    "page": page_num + 1, "doc_type": src["doc_type"],
                                })
                finally:
                    doc.close()
            elif ext in DOCX_EXT:
                txt = self._docx_extract_text(file_bytes)
                for sub in self._chunk_text(txt):
                    chunks.append({"content": sub, "source": src["source"], "page": 0, "doc_type": src["doc_type"]})
            elif ext in PPTX_EXT:
                txt = self._pptx_extract_text(file_bytes)
                for sub in self._chunk_text(txt):
                    chunks.append({"content": sub, "source": src["source"], "page": 0, "doc_type": src["doc_type"]})
            elif ext in XLSX_EXT and HAS_XLSX:
                txt = self._xlsx_extract_text(file_bytes)
                for sub in self._chunk_text(txt):
                    chunks.append({"content": sub, "source": src["source"], "page": 0, "doc_type": src["doc_type"]})
        except Exception as e:
            print(f"[DocGen] Text extraction failed for {src['source']}: {e}")
        return chunks
    def _chunk_text(self, text: str, size: int = 1000, overlap: int = 200) -> list[str]:
        words = text.split()
        out, i = [], 0
        while i < len(words):
            out.append(" ".join(words[i:i+size]))
            i += size - overlap
        return [c for c in out if len(c.strip()) > 50]
    def _extract_images_from_bytes(self, file_bytes: bytes, src: dict) -> list[dict]:
        ext = src["ext"]
        images: list[dict] = []
        try:
            if ext in PDF_EXT:
                images.extend(self._extract_pdf_images(file_bytes, src))
                images.extend(self._render_pdf_pages(file_bytes, src))
            elif ext in DOCX_EXT:
                images.extend(self._extract_docx_images(file_bytes, src))
                images.extend(self._render_office_pages(file_bytes, src, ".docx"))
            elif ext in PPTX_EXT:
                images.extend(self._extract_pptx_images(file_bytes, src))
                images.extend(self._render_office_pages(file_bytes, src, ".pptx"))
            elif ext in XLSX_EXT and HAS_XLSX:
                images.extend(self._extract_xlsx_images(file_bytes, src))
            elif ext in IMG_EXT:
                images.extend(self._ingest_standalone_image(file_bytes, src))
            elif ext in SVG_EXT:
                images.extend(self._ingest_svg_image(file_bytes, src))
        except Exception as e:
            print(f"[DocGen] Image extraction dispatch failed: {e}")
        return images
    def _soffice_binary(self) -> Optional[str]:
        import os, shutil
        for cand in (
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            "/opt/homebrew/bin/soffice",
            "/usr/local/bin/soffice",
            shutil.which("soffice") or "",
            shutil.which("libreoffice") or "",
        ):
            if cand and os.path.isfile(cand): return cand
        return None
    def _office_to_pdf(self, file_bytes: bytes, suffix: str) -> Optional[bytes]:
        soffice = self._soffice_binary()
        if not soffice: return None
        import os, subprocess, tempfile, uuid
        with tempfile.TemporaryDirectory(prefix="docgen_") as td:
            in_path = os.path.join(td, f"in_{uuid.uuid4().hex}{suffix}")
            with open(in_path, "wb") as fh:
                fh.write(file_bytes)
            try:
                subprocess.run(
                    [soffice, "--headless", "--norestore", "--convert-to", "pdf",
                     "--outdir", td, in_path],
                    check=True, timeout=180,
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
            except Exception as e:
                print(f"[DocGen] soffice conversion failed: {e}")
                return None
            pdf_path = os.path.splitext(in_path)[0] + ".pdf"
            if not os.path.exists(pdf_path): return None
            with open(pdf_path, "rb") as fh: return fh.read()
    def _render_pdf_pages(self, pdf_bytes: bytes, src: dict,
                          max_pages: int = 40, dpi: int = 110) -> list[dict]:
        try:
            import fitz
        except ImportError: return []
        out: list[dict] = []
        try:
            probe = fitz.open(stream=pdf_bytes, filetype="pdf")
            n = min(len(probe), max_pages)
            probe.close()
        except Exception as e:
            print(f"[DocGen] PDF page render failed: {e}")
            return out
        zoom = dpi / 72.0
        def _render_one(i: int):
            try:
                d = fitz.open(stream=pdf_bytes, filetype="pdf")
                mat = fitz.Matrix(zoom, zoom)
                page = d.load_page(i)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                png = pix.tobytes("png")
                d.close()
                img_id = f"pagerender_{src.get('file_id','x')}_p{i+1}_{uuid.uuid4().hex[:6]}"
                meta = {
                    "source": src.get("source"),
                    "doc_type": src.get("doc_type"),
                    "page": i + 1,
                    "kind": "page_render",
                    "caption": f"{src.get('source','')} — page {i+1}",
                }
                _IMAGE_STORE.put(img_id, png, meta)
                return {"id": img_id, "png_bytes": png, "metadata": meta}
            except Exception as e:
                print(f"[DocGen] page render p{i+1} failed: {e}")
                return None
        with ThreadPoolExecutor(max_workers=min(4, n or 1)) as ex:
            for rec in ex.map(_render_one, range(n)):
                if rec:
                    out.append(rec)
        out.sort(key=lambda r: r.get("metadata", {}).get("page", 0))
        return out
    def _render_office_pages(self, office_bytes: bytes, src: dict,
                             suffix: str) -> list[dict]:
        if suffix.lower() == ".pptx":
            snaps = self._render_pptx_slides_pure_python(office_bytes, src)
            if snaps: return snaps
        pdf = self._office_to_pdf(office_bytes, suffix)
        if pdf: return self._render_pdf_pages(pdf, src)
        return []
    def _extract_pdf_images(self, pdf_bytes: bytes, src: dict) -> list[dict]:
        out = []
        if not HAS_FITZ:
            print("[DocGen] PDF image extraction skipped — PyMuPDF not installed in this environment.")
            return out
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        try:
            for page_num, page in enumerate(doc):
                for img_idx, img in enumerate(page.get_images(full=True)):
                    xref = img[0]
                    try:
                        base = doc.extract_image(xref)
                        pil = Image.open(io.BytesIO(base["image"])).convert("RGB")
                    except Exception: continue
                    if not self._passes_filter(pil): continue
                    caption = self._pdf_caption(page, img)
                    context = page.get_text("text")[:800]
                    rec = self._store_image(pil, src, caption, context,
                                             f"page {page_num+1}", f"p{page_num}_i{img_idx}")
                    if rec: out.append(rec)
        finally:
            doc.close()
        return out
    def _pdf_caption(self, page, img) -> str:
        try:
            bbox = page.get_image_bbox(img)
            blocks = page.get_text("blocks")
            below = sorted([b for b in blocks if b[1] > bbox.y1 and b[1] - bbox.y1 < 80], key=lambda b: b[1])
            if below: return below[0][4].strip().replace("\n", " ")
            above = [b for b in blocks if b[3] < bbox.y0 and bbox.y0 - b[3] < 60]
            if above: return above[-1][4].strip().replace("\n", " ")
        except Exception: pass
        return ""
    def _extract_docx_images(self, docx_bytes: bytes, src: dict) -> list[dict]:
        out = []
        fn_junk = re.compile(r"(logo|icon|bullet|divider|watermark|thumbnail)", re.I)
        try:
            with zipfile.ZipFile(io.BytesIO(docx_bytes)) as zf:
                rel_map = {}
                try:
                    rels_xml = zf.read("word/_rels/document.xml.rels").decode("utf-8", "ignore")
                    rels_root = ET.fromstring(rels_xml)
                    for rel in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                        if "image" in rel.get("Type", ""):
                            rel_map[rel.get("Id")] = rel.get("Target")
                except Exception: pass
                try:
                    doc_xml = zf.read("word/document.xml").decode("utf-8", "ignore")
                    doc_root = ET.fromstring(doc_xml)
                except Exception: return out
                paragraphs = []
                for p in doc_root.iter(f"{NS_W}p"):
                    text = "".join(t.text or "" for t in p.iter(f"{NS_W}t"))
                    rids = [blip.get(f"{NS_R}embed") for blip in p.iter(f"{NS_A}blip") if blip.get(f"{NS_R}embed")]
                    paragraphs.append({"text": text.strip(), "rids": rids})
                for p_idx, para in enumerate(paragraphs):
                    for rid in para["rids"]:
                        target = rel_map.get(rid)
                        if not target: continue
                        media_path = f"word/{target}" if not target.startswith("word/") else target
                        leaf = media_path.rsplit("/", 1)[-1]
                        if fn_junk.search(leaf): continue
                        try: blob = zf.read(media_path)
                        except KeyError:
                            try: blob = zf.read(target)
                            except KeyError: continue
                        try: pil = Image.open(io.BytesIO(blob)).convert("RGB")
                        except Exception: continue
                        if not self._passes_filter(pil): continue
                        caption = para["text"]
                        if not caption:
                            for q in range(p_idx + 1, min(p_idx + 3, len(paragraphs))):
                                if paragraphs[q]["text"]:
                                    caption = paragraphs[q]["text"]; break
                        ctx_parts = [paragraphs[q]["text"]
                                     for q in range(max(0, p_idx-2), min(len(paragraphs), p_idx+3))
                                     if paragraphs[q]["text"]]
                        context = " ".join(ctx_parts)[:800]
                        rec = self._store_image(pil, src, caption[:400], context,
                                                 f"paragraph {p_idx}", f"docx_{rid}")
                        if rec: out.append(rec)
        except Exception as e:
            print(f"[DocGen] DOCX image extraction error: {e}")
        return out
    def _docx_extract_text(self, docx_bytes: bytes) -> str:
        try:
            with zipfile.ZipFile(io.BytesIO(docx_bytes)) as zf:
                doc_xml = zf.read("word/document.xml").decode("utf-8", "ignore")
                root = ET.fromstring(doc_xml)
                return "\n".join(t.text for t in root.iter(f"{NS_W}t") if t.text)
        except Exception as e:
            print(f"[DocGen] DOCX text failed: {e}")
            return ""
    def _extract_pptx_images(self, pptx_bytes: bytes, src: dict) -> list[dict]:
        out: list[dict] = []
        seen_media: set = set()
        fn_junk = re.compile(r"(logo|icon|bullet|divider|watermark|thumbnail)", re.I)
        try:
            with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zf:
                names = zf.namelist()
                slide_files = sorted(
                    [n for n in names if re.match(r"ppt/slides/slide\d+\.xml$", n)],
                    key=lambda n: int(re.search(r"slide(\d+)", n).group(1))
                )
                for slide_idx, slide_name in enumerate(slide_files):
                    try:
                        slide_xml = zf.read(slide_name).decode("utf-8", "ignore")
                        slide_root = ET.fromstring(slide_xml)
                    except Exception: continue
                    texts = [t.text for t in slide_root.iter(f"{NS_A}t") if t.text]
                    slide_text = " | ".join(texts)[:800]
                    title = (texts[0] if texts else "").strip()
                    rels_name = slide_name.replace("ppt/slides/", "ppt/slides/_rels/").replace(".xml", ".xml.rels")
                    rel_map = {}
                    try:
                        rels_xml = zf.read(rels_name).decode("utf-8", "ignore")
                        rels_root = ET.fromstring(rels_xml)
                        for rel in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                            if "image" in rel.get("Type", ""):
                                rel_map[rel.get("Id")] = rel.get("Target")
                    except Exception: pass
                    for shape_idx, blip in enumerate(slide_root.iter(f"{NS_A}blip")):
                        rid = blip.get(f"{NS_R}embed")
                        if not rid or rid not in rel_map: continue
                        target = rel_map[rid]
                        if target.startswith("../"):
                            media_path = "ppt/" + target[3:]
                        elif target.startswith("/"):
                            media_path = target.lstrip("/")
                        else:
                            media_path = f"ppt/slides/{target}"
                        if media_path not in names:
                            leaf = target.rsplit("/", 1)[-1]
                            match = next((n for n in names if n.endswith(leaf) and "media" in n), None)
                            if match:
                                media_path = match
                            else: continue
                        if fn_junk.search(media_path.rsplit("/", 1)[-1]): continue
                        if media_path in seen_media: continue
                        seen_media.add(media_path)
                        try:
                            blob = zf.read(media_path)
                            pil = Image.open(io.BytesIO(blob)).convert("RGB")
                        except Exception: continue
                        if not self._passes_filter(pil): continue
                        caption = title or (slide_text[:160]) or f"Slide {slide_idx+1}"
                        rec = self._store_image(pil, src, caption[:400], slide_text,
                                                 f"slide {slide_idx+1}",
                                                 f"slide{slide_idx}_{shape_idx}")
                        if rec:
                            out.append(rec)
                deck_name = (src.get("source") or "").rsplit(".", 1)[0]
                for n in names:
                    if not n.startswith("ppt/media/"): continue
                    if n in seen_media: continue
                    leaf = n.rsplit("/", 1)[-1]
                    if fn_junk.search(leaf): continue
                    if not re.search(r"\.(png|jpe?g|webp|bmp|tiff?|gif)$", leaf, re.I): continue
                    try:
                        blob = zf.read(n)
                        pil = Image.open(io.BytesIO(blob)).convert("RGB")
                    except Exception: continue
                    if not self._passes_filter(pil): continue
                    seen_media.add(n)
                    caption = f"{deck_name} — {leaf}"
                    rec = self._store_image(pil, src, caption[:400], deck_name,
                                             "deck media", f"media_{leaf}")
                    if rec:
                        out.append(rec)
        except Exception as e:
            print(f"[DocGen] PPTX image extraction error: {e}")
        return out
    def _pptx_extract_text(self, pptx_bytes: bytes) -> str:
        try:
            with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zf:
                parts = []
                for name in sorted(zf.namelist()):
                    if re.match(r"ppt/slides/slide\d+\.xml$", name):
                        try:
                            xml = zf.read(name).decode("utf-8", "ignore")
                            root = ET.fromstring(xml)
                            texts = [t.text for t in root.iter(f"{NS_A}t") if t.text]
                            parts.append(" ".join(texts))
                        except Exception: continue
                return "\n\n".join(parts)
        except Exception as e:
            print(f"[DocGen] PPTX text failed: {e}")
            return ""
    def _extract_xlsx_images(self, xlsx_bytes: bytes, src: dict) -> list[dict]:
        out = []
        if not HAS_XLSX: return out
        try:
            wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheet_text_parts = []
                for row in ws.iter_rows(max_row=30, max_col=15, values_only=True):
                    for v in row:
                        if v is not None and isinstance(v, str) and len(v.strip()) > 2:
                            sheet_text_parts.append(str(v).strip())
                sheet_text = " | ".join(sheet_text_parts)[:600]
                for img_idx, img in enumerate(ws._images):
                    try:
                        blob = img._data()
                        pil = Image.open(io.BytesIO(blob)).convert("RGB")
                    except Exception: continue
                    if not self._passes_filter(pil): continue
                    caption = f"{sheet}: {sheet_text_parts[0] if sheet_text_parts else ''}"[:200]
                    rec = self._store_image(pil, src, caption, sheet_text,
                                             f"sheet '{sheet}'", f"{sheet}_{img_idx}")
                    if rec: out.append(rec)
        except Exception as e:
            print(f"[DocGen] XLSX image extraction error: {e}")
        return out
    def _xlsx_extract_text(self, xlsx_bytes: bytes) -> str:
        try:
            wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                row_parts = []
                for row in ws.iter_rows(max_row=200, max_col=30, values_only=True):
                    cells = [str(v) for v in row if v is not None]
                    if cells: row_parts.append(" | ".join(cells))
                parts.append(f"[Sheet: {sheet}]\n" + "\n".join(row_parts))
            return "\n\n".join(parts)
        except Exception as e:
            print(f"[DocGen] XLSX text failed: {e}")
            return ""
    def _ingest_standalone_image(self, img_bytes: bytes, src: dict) -> list[dict]:
        try: pil = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        except Exception as e:
            print(f"[DocGen] Standalone open failed: {e}"); return []
        if not self._passes_filter(pil): return []
        caption = self._humanize(src["source"].rsplit(".", 1)[0])
        rec = self._store_image(pil, src, caption, "", "standalone image", "standalone")
        return [rec] if rec else []
    def _ingest_svg_image(self, svg_bytes: bytes, src: dict) -> list[dict]:
        svg_text = svg_bytes.decode("utf-8", errors="ignore")[:2000]
        pil = None
        if HAS_CAIROSVG:
            try:
                png_bytes = cairosvg.svg2png(bytestring=svg_bytes, output_width=1200)
                pil = Image.open(io.BytesIO(png_bytes)).convert("RGB")
            except Exception as e:
                print(f"[DocGen] SVG rasterize failed: {e}")
        if pil is None or not self._passes_filter(pil): return []
        caption = self._humanize(src["source"].rsplit(".", 1)[0])
        rec = self._store_image(pil, src, caption, svg_text, "SVG", "svg")
        return [rec] if rec else []
    def _ingest_remote_image(self, url: str, cand: dict) -> Optional[dict]:
        try:
            host = urlparse(url).netloc.lower()
            if "wikimedia.org" in host or "wikipedia.org" in host:
                ua = "IBM-DocGen/2.0 (https://ibm.com; IBM Consulting) python-requests"
            else:
                ua = "Mozilla/5.0 (ibm-docgen)"
            r = self._http.get(url, timeout=self.valves.web_image_fetch_timeout,
                                headers={"User-Agent": ua, "Accept": "image/*,*/*"})
            if r.status_code == 429:
                if "wikimedia.org" in host or "wikipedia.org" in host:
                    self._breaker_trip("wikipedia"); self._breaker_trip("wikimedia")
                print(f"[DocGen] 429 rate-limit from {host} on {url[:80]}")
                return None
            r.raise_for_status()
            if not r.content: return None
            try: pil = Image.open(io.BytesIO(r.content)).convert("RGB")
            except Exception: return None
            if not self._passes_filter(pil): return None
            src = {
                "source": cand.get("source_page") or urlparse(url).netloc,
                "ext": "." + url.rsplit(".", 1)[-1].split("?")[0].lower() if "." in url else ".jpg",
                "doc_type": "web",
            }
            return self._store_image(pil, src, cand.get("title", "")[:400],
                                     cand.get("snippet", "")[:800],
                                     f"web: {urlparse(url).netloc}", "web")
        except Exception as e:
            print(f"[DocGen] Remote image failed for {url}: {e}")
            return None
    def _store_image(self, pil, src, caption, context, location, tag) -> Optional[dict]:
        pil = self._downscale(pil)
        buf = io.BytesIO()
        pil.save(buf, format="PNG", optimize=True)
        png_bytes = buf.getvalue()
        if len(png_bytes) > self.valves.max_image_bytes:
            pil.thumbnail((1024, 1024), Image.Resampling.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format="PNG", optimize=True)
            png_bytes = buf.getvalue()
            if len(png_bytes) > self.valves.max_image_bytes: return None
        stem = (src.get("source", "unknown") or "unknown").rsplit(".", 1)[0]
        safe_stem = re.sub(r"[^a-zA-Z0-9_-]", "_", stem)[:40]
        img_id = f"{safe_stem}_{tag}_{uuid.uuid4().hex[:8]}"
        metadata = {
            "id": img_id, "caption": (caption or "")[:400],
            "context": (context or "")[:800],
            "source": src.get("source", ""), "location": location,
            "doc_type": src.get("doc_type", "general"),
            "source_format": src.get("ext", ""),
            "width": pil.width, "height": pil.height,
            "byte_size": len(png_bytes),
        }
        _IMAGE_STORE.put(img_id, png_bytes, metadata)
        return metadata
    def _downscale(self, pil):
        max_dim = 1600
        if pil.width > max_dim or pil.height > max_dim:
            pil = pil.copy()
            pil.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)
        return pil
    def _passes_filter(self, pil):
        if pil.width < self.valves.min_image_width or pil.height < self.valves.min_image_height: return False
        aspect = pil.width / pil.height
        if aspect > self.valves.max_image_aspect_ratio or aspect < 1/self.valves.max_image_aspect_ratio: return False
        return True
    def _png_thumbnail(self, png_bytes: bytes, max_px: int) -> bytes:
        try:
            im = Image.open(io.BytesIO(png_bytes))
            im.thumbnail((max_px, max_px))
            out = io.BytesIO()
            im.convert("RGB").save(out, format="PNG", optimize=True)
            return out.getvalue()
        except Exception: return png_bytes
    def _vision_rank_sync(self, query: str, images: list, auth: dict) -> list:
        if not self.valves.vision_rank_enabled or not images: return images
        max_n = max(1, int(self.valves.vision_rank_max_images))
        pick = images[:max_n]
        rest = images[max_n:]
        image_parts = []
        for i, img in enumerate(pick):
            png = img.get("png_bytes")
            if not png:
                png = _IMAGE_STORE.get_bytes(img.get("id", "")) if img.get("id") else None
            if not png: continue
            thumb = self._png_thumbnail(png, int(self.valves.vision_rank_thumb_px))
            b64 = base64.b64encode(thumb).decode("ascii")
            image_parts.append((i, b64))
        if not image_parts: return images
        instruction = (
            f'Query: "{query}"\n\n'
            f"You will see {len(image_parts)} numbered images. For EACH image, produce one entry:\n"
            '  {"idx": <int>, "caption": "<<=15 words>>", "score": <0-10 int>}\n'
            "score = visual relevance to the Query (10 = perfect match, 0 = irrelevant/decorative).\n"
            "Return ONLY a JSON array of entries (no prose, no code fences)."
        )
        content = [{"type": "text", "text": instruction}]
        for idx, b64 in image_parts:
            content.append({"type": "text", "text": f"Image {idx}:"})
            content.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{b64}"},
            })
        payload = {
            "model": self.valves.vision_rank_model,
            "messages": [{"role": "user", "content": content}],
            "temperature": 0.0,
            "stream": False,
        }
        url = f"{self.valves.owui_base_url}/api/chat/completions"
        try:
            r = requests.post(
                url,
                headers={**(auth or {}), "Content-Type": "application/json"},
                json=payload,
                timeout=max(60, self.valves.request_timeout + 60),
            )
            r.raise_for_status()
            data = r.json()
            text = (data.get("choices") or [{}])[0].get("message", {}).get("content", "")
            if isinstance(text, list):
                text = "".join(p.get("text", "") for p in text if isinstance(p, dict))
            _fence = chr(96) * 3
            text = re.sub(_fence + r"(?:json)?|" + _fence, "", text, flags=re.I).strip()
            m = re.search(r"\[\s*\{.*\}\s*\]", text, flags=re.S)
            if m:
                text = m.group(0)
            entries = json.loads(text)
        except Exception as e:
            print(f"[DocGen] vision rank request failed: {e}")
            return images
        for entry in entries if isinstance(entries, list) else []:
            try:
                i = int(entry.get("idx", -1))
                if 0 <= i < len(pick):
                    cap = str(entry.get("caption", "")).strip()
                    score = float(entry.get("score", 0) or 0)
                    if cap:
                        pick[i]["caption"] = cap
                        pick[i]["vision_caption"] = cap
                    pick[i]["vision_score"] = max(0.0, min(10.0, score))
            except Exception: continue
        pick.sort(key=lambda x: x.get("vision_score", 0.0), reverse=True)
        return pick + rest
    async def _vision_rank_async(self, query: str, images: list, auth: dict) -> list:
        if not self.valves.vision_rank_enabled or not images: return images
        try:
            return await asyncio.to_thread(self._vision_rank_sync, query, images, auth)
        except Exception as e:
            print(f"[DocGen] vision rank async failed: {e}")
            return images
    def _fetch_attachment_bytes(self, file_ids, auth):
        out = []
        base = self.valves.owui_base_url
        h = {**(auth or {})}
        for fid in file_ids or []:
            try:
                meta_r = requests.get(f"{base}/api/v1/files/{fid}", headers=h, timeout=10)
                fname = (meta_r.json().get("filename") or fid) if meta_r.ok else fid
                ext = ("." + fname.rsplit(".", 1)[-1].lower()) if "." in fname else ""
                if ext not in (".pdf", ".docx", ".pptx", ".xlsx"): continue
                r = requests.get(f"{base}/api/v1/files/{fid}/content", headers=h,
                                 timeout=self.valves.request_timeout)
                if r.status_code == 200 and r.content:
                    out.append((fname, ext, r.content))
            except Exception as e:
                print(f"[DocGen] fetch attachment {fid} failed: {e}")
        return out
    def _extract_office_images(self, office_attachments):
        """Zipfile-extract images + surrounding text context from DOCX/PPTX/XLSX."""
        out = []
        strip_tags = lambda s: re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", s)).strip()
        media_prefix = lambda ext: {".docx": "word/media/", ".pptx": "ppt/media/",
                                     ".xlsx": "xl/media/"}.get(ext, "")
        img_ok = lambda n: any(n.lower().endswith(e) for e in (".png", ".jpg", ".jpeg", ".gif", ".webp"))
        for fname, ext, data in office_attachments:
            try:
                zf = zipfile.ZipFile(io.BytesIO(data))
                parts = []
                for n in zf.namelist():
                    if not n.endswith(".xml"): continue
                    if not any(n.startswith(p) for p in ("word/document", "ppt/slides/", "xl/worksheets/")):
                        continue
                    try:
                        cleaned = strip_tags(zf.read(n).decode("utf-8", errors="ignore"))
                        if cleaned: parts.append(cleaned)
                    except Exception: pass
                ctx = (" ".join(parts))[:12000] or fname
                prefix = media_prefix(ext)
                if not prefix: continue
                for n in zf.namelist():
                    if not n.startswith(prefix) or not img_ok(n): continue
                    b = zf.read(n)
                    if len(b) < 2000: continue
                    out.append({"id": f"{fname}:{n.split('/')[-1]}", "bytes": b,
                                "source_file": fname, "page": 0, "caption_seed": ctx})
            except Exception as e:
                print(f"[DocGen] office extract {fname} failed: {e}")
        return out
    def _bm25_prefilter(self, section, candidates, top_n):
        if not candidates or len(candidates) <= top_n: return list(candidates)
        tok = lambda s: set(re.findall(r"\w{3,}", (s or "").lower()))
        q = tok((section.get("title", "") + " " +
                 " ".join(section.get("bullets", []) or []) + " " +
                 " ".join(section.get("paragraphs", []) or [])))
        scored = sorted(((len(q & tok(c.get("caption_seed", ""))), c) for c in candidates),
                        key=lambda x: x[0], reverse=True)
        return [c for _, c in scored[:top_n]]
    def _claude_api_chat(self, messages, auth, timeout=90):
        """Single-shot Claude call via OWUI's /api/chat/completions. Returns str content or None."""
        payload = {"model": self.valves.vision_rank_model, "messages": messages,
                   "temperature": 0.0, "stream": False}
        url = f"{self.valves.owui_base_url}/api/chat/completions"
        try:
            r = requests.post(url, headers={**(auth or {}), "Content-Type": "application/json"},
                              json=payload, timeout=timeout)
            if r.status_code >= 400:
                print(f"[DocGen] Claude API {r.status_code}: {r.text[:200]}")
                return None
            text = (r.json().get("choices") or [{}])[0].get("message", {}).get("content", "")
            if isinstance(text, list):
                text = "".join(p.get("text", "") for p in text if isinstance(p, dict))
            return text
        except Exception as e:
            print(f"[DocGen] Claude API error: {e}")
            return None
    def _parse_match_json(self, text):
        if not text: return None
        _f = chr(96) * 3
        text = re.sub(_f + r"(?:json)?|" + _f, "", text, flags=re.I).strip()
        m = re.search(r"\{.*\}", text, flags=re.S)
        if m: text = m.group(0)
        try: return json.loads(text)
        except Exception: return None
    def _plan_b_match(self, section, pdf_attachments, auth):
        """Plan B: attach PDFs as native Anthropic document blocks. Return ref-card match or None."""
        if not pdf_attachments: return None
        title = section.get("title", "")
        bullets = " | ".join((section.get("bullets", []) or [])[:5])
        prompt = (f'Section title: "{title}"\nKey points: {bullets}\n\n'
                  "You have the user's attached PDF(s). Identify the ONE page that best illustrates this section. "
                  'Return ONLY JSON: {"source_file":"<fname>","page":<int 1-based>,"score":<0-100>,"caption":"<=15 words"}\n'
                  "95-100 = perfect topical match; <95 = weak.")
        content = [{"type": "text", "text": prompt}]
        for fname, _ext, data in pdf_attachments[:3]:
            content.append({"type": "text", "text": f"--- PDF: {fname} ---"})
            content.append({"type": "document",
                            "source": {"type": "base64", "media_type": "application/pdf",
                                       "data": base64.b64encode(data).decode("ascii")}})
        text = self._claude_api_chat([{"role": "user", "content": content}], auth, timeout=120)
        parsed = self._parse_match_json(text)
        if not parsed: return None
        return {"source_file": str(parsed.get("source_file", ""))[:120],
                "page": int(parsed.get("page", 0) or 0),
                "score": float(parsed.get("score", 0) or 0),
                "caption": str(parsed.get("caption", "")).strip()[:200],
                "type": "reference"}
    def _plan_c_match(self, section, candidates, auth):
        """Plan C: send image candidates to Claude Vision, pick the best."""
        if not candidates: return None
        image_parts = []
        for i, c in enumerate(candidates[:self.valves.kb_max_candidates_per_section]):
            try:
                thumb = self._png_thumbnail(c["bytes"], self.valves.vision_rank_thumb_px)
                b64 = base64.b64encode(thumb).decode("ascii")
                image_parts.append((i, b64, c))
            except Exception: pass
        if not image_parts: return None
        title = section.get("title", "")
        bullets = " | ".join((section.get("bullets", []) or [])[:5])
        prompt = (f'Section title: "{title}"\nKey points: {bullets}\n\n'
                  f"You see {len(image_parts)} candidate images from the user's knowledge base. "
                  "Pick the ONE that best ILLUSTRATES this section (not one that just repeats the text). Score 0-100:\n"
                  '  95-100 = perfectly illustrates; 70-94 = related; <70 = weak\n'
                  'Return ONLY JSON: {"best_idx":<int>,"score":<0-100>,"caption":"<=15 words"}')
        content = [{"type": "text", "text": prompt}]
        for idx, b64, _c in image_parts:
            content.append({"type": "text", "text": f"Image {idx}:"})
            content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}})
        text = self._claude_api_chat([{"role": "user", "content": content}], auth, timeout=90)
        parsed = self._parse_match_json(text)
        if not parsed: return None
        try: idx = int(parsed.get("best_idx", -1))
        except Exception: return None
        hit = next((c for i, _b, c in image_parts if i == idx), None)
        if not hit: return None
        return {"image_bytes": hit["bytes"], "source_file": hit["source_file"],
                "score": float(parsed.get("score", 0) or 0),
                "caption": str(parsed.get("caption", "")).strip()[:200],
                "type": "image"}
    async def _kb_enrich_sections(self, sections, file_ids, auth, __event_emitter__=None):
        """Enrich each section with a KB match (Plan B first, Plan C fallback). 95 threshold."""
        if not file_ids or not self.valves.enable_kb_vision_layout or not sections:
            return sections
        try: await self._emit(__event_emitter__, f"📎 Indexing {len(file_ids)} attachment(s) for 50/50 layout...")
        except Exception: pass
        attachments = await asyncio.to_thread(self._fetch_attachment_bytes, file_ids, auth)
        if not attachments: return sections
        pdfs = [a for a in attachments if a[1] == ".pdf"]
        office = [a for a in attachments if a[1] in (".docx", ".pptx", ".xlsx")]
        office_images = await asyncio.to_thread(self._extract_office_images, office) if office else []
        # Probe Plan B once
        plan_b_ok = False
        if pdfs and self.valves.kb_plan_b_enabled:
            probe = await asyncio.to_thread(self._plan_b_match, sections[0], pdfs, auth)
            plan_b_ok = probe is not None
            try: await self._emit(__event_emitter__,
                ("✅ Plan B active (PDF document blocks)" if plan_b_ok
                 else f"⚠️ Plan B unsupported — falling back to Plan C ({len(office_images)} KB images)"))
            except Exception: pass
        thr = int(self.valves.kb_vision_score_threshold)
        async def match_one(s):
            try:
                if plan_b_ok:
                    m = await asyncio.to_thread(self._plan_b_match, s, pdfs, auth)
                    if m and m["score"] >= thr: return m
                if office_images:
                    top = self._bm25_prefilter(s, office_images, self.valves.kb_max_candidates_per_section)
                    m = await asyncio.to_thread(self._plan_c_match, s, top, auth)
                    if m and m["score"] >= thr: return m
            except Exception as e:
                print(f"[DocGen] KB match error: {e}")
            return None
        results = await asyncio.gather(*[match_one(s) for s in sections], return_exceptions=True)
        for s, m in zip(sections, results):
            if isinstance(m, dict) and m.get("image_bytes"):
                s["_kb_match"] = m
        accepted = sum(1 for s in sections if s.get("_kb_match"))
        try: await self._emit(__event_emitter__,
            f"👁️ KB 50/50 layout applied to {accepted}/{len(sections)} sections with ≥{thr}/100 image match "
            f"(others render as full-width text)")
        except Exception: pass
        return sections
    def _rank_images(self, query, images):
        q_tokens = set(re.findall(r"\w{3,}", query.lower()))
        scored = []
        for img in images:
            hay = " ".join([img.get("caption", ""), img.get("context", ""), img.get("source", "")]).lower()
            tokens = set(re.findall(r"\w{3,}", hay))
            overlap = len(q_tokens & tokens)
            dbst = {"case_study": 1.5, "solution_brief": 1.3, "methodology": 1.2,
                    "capability": 1.1, "web": 1.0, "general": 1.0}.get(img.get("doc_type", "general"), 1.0)
            fbst = {".pdf": 1.1, ".pptx": 1.2, ".docx": 1.0, ".xlsx": 0.9,
                    ".svg": 1.15, ".png": 1.0, ".jpg": 1.0, ".jpeg": 1.0, ".webp": 1.0
                    }.get(img.get("source_format", ""), 1.0)
            lex_score = (overlap + 0.5) * dbst * fbst
            kind = (img.get("metadata", {}) or {}).get("kind") or img.get("kind")
            kind_penalty = 0.4 if kind in ("slide_snapshot", "page_render") else 1.0
            vscore = img.get("vision_score")
            composite = ((float(vscore) * 100.0 if vscore is not None else 0.0) + lex_score) * kind_penalty
            scored.append((composite, img))
        scored.sort(key=lambda x: x[0], reverse=True)
        return [img for _, img in scored]
    def _rank_text(self, query, chunks):
        q_tokens = set(re.findall(r"\w{3,}", query.lower()))
        scored = []
        for c in chunks:
            tokens = set(re.findall(r"\w{3,}", (c.get("content") or "").lower()))
            scored.append((len(q_tokens & tokens), c))
        scored.sort(key=lambda x: x[0], reverse=True)
        return [c for _, c in scored]
    def _google_search_text(self, query, num=6):
        out = []
        try:
            r = requests.get("https://www.googleapis.com/customsearch/v1", params={
                "key": self.valves.google_api_key, "cx": self.valves.google_cx,
                "q": query, "num": max(1, min(num, 10)),
            }, timeout=self.valves.request_timeout)
            r.raise_for_status()
            for item in r.json().get("items", []):
                out.append({
                    "content": f"{item.get('title', '')}. {item.get('snippet', '')}",
                    "source": item.get("displayLink", ""),
                    "url": item.get("link", ""), "page": 0, "doc_type": "web",
                })
        except Exception as e:
            print(f"[DocGen] Google text search failed: {e}")
        return out
    def _google_search_images(self, query, num=10):
        out = []
        try:
            r = requests.get("https://www.googleapis.com/customsearch/v1", params={
                "key": self.valves.google_api_key, "cx": self.valves.google_cx,
                "q": query, "searchType": "image", "num": max(1, min(num, 10)),
                "safe": "active", "imgSize": "large",
            }, timeout=self.valves.request_timeout)
            r.raise_for_status()
            for item in r.json().get("items", []):
                out.append({
                    "url": item.get("link", ""),
                    "title": item.get("title", ""),
                    "snippet": item.get("snippet", ""),
                    "source_page": item.get("image", {}).get("contextLink", ""),
                })
        except Exception as e:
            print(f"[DocGen] Google image search failed: {e}")
        return out
    def _wikipedia_search_text(self, query, num=6):
        out = []
        try:
            r = self._http.get(
                "https://en.wikipedia.org/w/api.php",
                params={
                    "action": "query", "format": "json",
                    "generator": "search",
                    "gsrsearch": query,
                    "gsrlimit": max(1, min(num, 8)),
                    "gsrnamespace": 0,
                    "prop": "extracts|info",
                    "exintro": 1, "explaintext": 1, "exchars": 1500,
                    "inprop": "url",
                    "redirects": 1,
                },
                headers={"User-Agent": self._WIKI_UA, "Accept": "application/json"},
                timeout=self.valves.request_timeout,
            )
            r.raise_for_status()
            pages = r.json().get("query", {}).get("pages", {})
            ordered = sorted(pages.values(), key=lambda p: p.get("index", 999))
            for page in ordered:
                title = page.get("title", "")
                extract = (page.get("extract") or "").strip()
                url = page.get("fullurl") or f"https://en.wikipedia.org/wiki/{title.replace(' ', '_')}"
                if not extract: continue
                out.append({
                    "content": f"{title}\n\n{extract}",
                    "source": "en.wikipedia.org",
                    "url": url, "page": 0, "doc_type": "web",
                })
        except Exception as e:
            print(f"[DocGen] Wikipedia text search failed: {e}")
        return out
    _WIKI_UA = "IBM-DocGen/2.1 (https://ibm.com; IBM Consulting) python-requests"
    def _wikipedia_lead_images(self, query, num=5):
        out = []
        try:
            r = requests.get(
                "https://en.wikipedia.org/w/api.php",
                params={
                    "action": "query", "format": "json",
                    "generator": "search",
                    "gsrsearch": query,
                    "gsrlimit": max(1, min(num, 10)),
                    "gsrnamespace": 0,
                    "prop": "pageimages|pageprops|info|extracts",
                    "piprop": "original|thumbnail",
                    "pithumbsize": 1200,
                    "pilimit": max(1, min(num, 10)),
                    "exintro": 1, "explaintext": 1, "exchars": 400,
                    "inprop": "url",
                    "redirects": 1,
                },
                headers={"User-Agent": self._WIKI_UA, "Accept": "application/json"},
                timeout=self.valves.request_timeout,
            )
            r.raise_for_status()
            pages = r.json().get("query", {}).get("pages", {})
            ordered = sorted(
                pages.values(),
                key=lambda p: p.get("index", 999),
            )
            for page in ordered:
                img = page.get("original") or page.get("thumbnail") or {}
                url = img.get("source")
                if not url: continue
                out.append({
                    "url": url,
                    "title": page.get("title", "")[:200],
                    "snippet": (page.get("extract") or "")[:400],
                    "source_page": page.get("fullurl", f"https://en.wikipedia.org/wiki/{page.get('title','').replace(' ','_')}"),
                })
        except Exception as e:
            print(f"[DocGen] Wikipedia lead-image search failed: {e}")
        return out
    def _wikimedia_search_images(self, query, num=10):
        out = []
        try:
            r = requests.get(
                "https://commons.wikimedia.org/w/api.php",
                params={
                    "action": "query", "format": "json", "generator": "search",
                    "gsrsearch": f"filetype:bitmap {query}",
                    "gsrnamespace": 6,
                    "gsrlimit": max(1, min(num, 20)),
                    "prop": "imageinfo",
                    "iiprop": "url|size|mime|extmetadata",
                    "iiurlwidth": 1200,
                },
                headers={"User-Agent": self._WIKI_UA, "Accept": "application/json"},
                timeout=self.valves.request_timeout,
            )
            r.raise_for_status()
            pages = r.json().get("query", {}).get("pages", {})
            for _, page in pages.items():
                info = (page.get("imageinfo") or [{}])[0]
                mime = info.get("mime", "")
                if not mime.startswith("image/"): continue
                if mime == "image/svg+xml": continue
                if info.get("size", 0) > 8_000_000: continue
                url = info.get("thumburl") or info.get("url", "")
                if not url: continue
                meta = info.get("extmetadata", {})
                title = page.get("title", "").replace("File:", "")
                desc = (meta.get("ImageDescription", {}) or {}).get("value", "")
                desc = re.sub(r"<[^>]+>", "", desc)[:400]
                out.append({
                    "url": url,
                    "title": title[:200],
                    "snippet": desc,
                    "source_page": page.get("fullurl", "commons.wikimedia.org"),
                })
        except Exception as e:
            print(f"[DocGen] Wikimedia image search failed: {e}")
        return out
    def _web_search_text(self, query, num=6):
        if self.valves.google_api_key and self.valves.google_cx:
            results = self._google_search_text(query, num)
            if results: return results
        return self._wikipedia_search_text(query, num)
    def _duckduckgo_search_images(self, query, num=10):
        out = []
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
                "Accept": "text/html,application/json",
                "Referer": "https://duckduckgo.com/",
            }
            sess = requests.Session()
            sess.headers.update(headers)
            token_resp = sess.get(
                "https://duckduckgo.com/",
                params={"q": query, "iax": "images", "ia": "images"},
                timeout=self.valves.request_timeout,
            )
            token_resp.raise_for_status()
            m = re.search(r"vqd=[\"']?([\d-]+)[\"']?", token_resp.text)
            if not m:
                m = re.search(r'"vqd":"([\d-]+)"', token_resp.text)
            if not m:
                print(f"[DocGen] DuckDuckGo: no vqd token in response")
                return out
            vqd = m.group(1)
            img_resp = sess.get(
                "https://duckduckgo.com/i.js",
                params={
                    "l": "us-en",
                    "o": "json",
                    "q": query,
                    "vqd": vqd,
                    "f": ",,,,,,",
                    "p": "1",
                },
                timeout=self.valves.request_timeout,
            )
            img_resp.raise_for_status()
            data = img_resp.json()
            results = data.get("results") or []
            for item in results[:max(1, min(num, 20))]:
                url = item.get("image") or item.get("thumbnail")
                if not url: continue
                if url.startswith("data:"): continue
                out.append({
                    "url": url,
                    "title": (item.get("title") or "")[:200],
                    "snippet": (item.get("source") or "")[:400],
                    "source_page": item.get("url") or "duckduckgo.com",
                })
        except Exception as e:
            print(f"[DocGen] DuckDuckGo image search failed: {e}")
        return out
    def _package(self, query, text_chunks, images, source):
        return json.dumps({
            "source_mode": source,
            "query": query,
            "text_chunks": [
                {
                    "id": f"T{i+1}",
                    "content": c.get("content", ""),
                    "source": c.get("source", ""),
                    "page": c.get("page", 0),
                    "doc_type": c.get("doc_type", "general"),
                    "url": c.get("url", ""),
                }
                for i, c in enumerate(text_chunks)
            ],
            "images": [
                {
                    "id": img["id"],
                    "display_id": img.get("display_id", f"IMG{i+1}"),
                    "caption": img.get("caption", ""),
                    "context": img.get("context", "")[:300],
                    "source": img.get("source", ""),
                    "location": img.get("location", ""),
                    "width": img.get("width"),
                    "height": img.get("height"),
                }
                for i, img in enumerate(images)
            ],
            "next_step": (
                "Now build a sections array (see assemble_document schema). "
                "For each section that deserves an image, set image_id to the "
                "display_id value of the chosen image (e.g. 'IMG1', 'IMG2') — "
                "use the display_id field, not the id field. "
                "Then call assemble_document(session_id, format, title, client_name, sections_json)."
            ),
        }, indent=2)
    def _build_and_render_docx(self, session_id, title, client_name, sections, emitter):
        doc_parts = []
        media_files = []
        rel_entries = []
        chart_parts = []
        chart_overrides = []
        def esc(s):
            return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
        def run_xml(text, size=22, bold=False, italic=False, color="161616"):
            return (
                f'<w:r><w:rPr>'
                f'<w:rFonts w:ascii="IBM Plex Sans" w:hAnsi="IBM Plex Sans"/>'
                f'<w:sz w:val="{size}"/>'
                f'<w:color w:val="{color}"/>'
                f'{"<w:b/>" if bold else ""}'
                f'{"<w:i/>" if italic else ""}'
                f'</w:rPr><w:t xml:space="preserve">{esc(text)}</w:t></w:r>'
            )
        def para_xml(runs, align="left", after=120, before=0):
            return (
                f'<w:p><w:pPr>'
                f'<w:jc w:val="{align}"/>'
                f'<w:spacing w:after="{after}" w:before="{before}" w:line="360" w:lineRule="auto"/>'
                f'</w:pPr>{runs}</w:p>'
            )
        def heading_xml(text, level=1, color="0F62FE"):
            sizes = {1: 40, 2: 32, 3: 26, 4: 22}
            sz = sizes.get(level, 22)
            return (
                f'<w:p><w:pPr><w:spacing w:before="360" w:after="160"/></w:pPr>'
                f'{run_xml(text, size=sz, bold=True, color=color)}</w:p>'
            )
        def table_xml(headers, rows, hdr_bg="0F62FE"):
            out = '<w:tbl><w:tblPr><w:tblW w:w="5000" w:type="pct"/><w:tblBorders>'
            for side in ("top", "left", "bottom", "right", "insideH", "insideV"):
                out += f'<w:{side} w:val="single" w:sz="4" w:color="E0E0E0"/>'
            out += '</w:tblBorders></w:tblPr>'
            if headers:
                out += '<w:tr>'
                for h in headers:
                    out += (
                        f'<w:tc><w:tcPr><w:shd w:val="clear" w:color="auto" w:fill="{hdr_bg}"/></w:tcPr>'
                        f'<w:p><w:pPr><w:jc w:val="left"/></w:pPr>'
                        f'{run_xml(h, size=20, bold=True, color="FFFFFF")}</w:p></w:tc>'
                    )
                out += '</w:tr>'
            for ri, row in enumerate(rows):
                bg = "F4F4F4" if ri % 2 == 1 else "FFFFFF"
                out += '<w:tr>'
                for cell in row:
                    out += (
                        f'<w:tc><w:tcPr><w:shd w:val="clear" w:color="auto" w:fill="{bg}"/></w:tcPr>'
                        f'<w:p>{run_xml(str(cell), size=20)}</w:p></w:tc>'
                    )
                out += '</w:tr>'
            out += '</w:tbl>'
            return out
        def add_image_xml(png_bytes, width_px, height_px, caption=None):
            idx = len(media_files)
            fname = f"image{idx+1}.png"
            media_files.append((fname, png_bytes))
            rid = f"rIdImg{idx}"
            rel_entries.append(
                f'<Relationship Id="{rid}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                f'Target="media/{fname}"/>'
            )
            display_px = 500
            aspect = height_px / width_px if width_px else 0.65
            w_emu = int(display_px * 9525)
            h_emu = int(w_emu * aspect)
            doc_parts.append(
                '<w:p><w:pPr><w:jc w:val="center"/><w:spacing w:before="240" w:after="60"/></w:pPr>'
                '<w:r><w:drawing>'
                '<wp:inline xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<wp:extent cx="{w_emu}" cy="{h_emu}"/>'
                f'<wp:docPr id="{idx+100}" name="Image {idx+1}"/>'
                '<wp:cNvGraphicFramePr/>'
                '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
                '<pic:pic>'
                f'<pic:nvPicPr><pic:cNvPr id="{idx+100}" name="{fname}"/><pic:cNvPicPr/></pic:nvPicPr>'
                f'<pic:blipFill><a:blip r:embed="{rid}"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
                '<pic:spPr>'
                f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{w_emu}" cy="{h_emu}"/></a:xfrm>'
                '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                '</pic:spPr></pic:pic>'
                '</a:graphicData></a:graphic>'
                '</wp:inline></w:drawing></w:r></w:p>'
            )
            if caption:
                cap_run = run_xml(f"Figure — {caption}", size=18, italic=True, color="525252")
                doc_parts.append(para_xml(cap_run, align="center", after=240))
        def add_chart_xml(spec, caption=None):
            chart_idx = len(chart_parts) + 1
            chart_xml_bytes = self._ooxml_chart_part_xml(spec)
            chart_parts.append((f"word/charts/chart{chart_idx}.xml", chart_xml_bytes))
            chart_overrides.append(
                f'<Override PartName="/word/charts/chart{chart_idx}.xml" '
                f'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
            )
            rid = f"rIdChart{chart_idx}"
            rel_entries.append(
                f'<Relationship Id="{rid}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
                f'Target="charts/chart{chart_idx}.xml"/>'
            )
            w_emu = 5486400
            h_emu = 3017520
            doc_parts.append(
                '<w:p><w:pPr><w:jc w:val="center"/><w:spacing w:before="240" w:after="60"/></w:pPr>'
                '<w:r><w:drawing>'
                '<wp:inline xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<wp:extent cx="{w_emu}" cy="{h_emu}"/>'
                f'<wp:docPr id="{1000 + chart_idx}" name="Chart {chart_idx}"/>'
                '<wp:cNvGraphicFramePr/>'
                '<a:graphic>'
                '<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
                f'<c:chart r:id="{rid}"/>'
                '</a:graphicData></a:graphic>'
                '</wp:inline></w:drawing></w:r></w:p>'
            )
            if caption:
                cap_run = run_xml(f"Chart — {caption}", size=18, italic=True, color="525252")
                doc_parts.append(para_xml(cap_run, align="center", after=240))
        doc_parts.append(para_xml(run_xml(title, size=56, bold=True, color="0F62FE"),
                                   align="left", after=240, before=1200))
        doc_parts.append(para_xml(
            run_xml(f"IBM Consulting  |  Prepared for {client_name}", size=24, color="525252"),
            align="left", after=480
        ))
        doc_parts.append(para_xml(
            run_xml(time.strftime("%B %Y"), size=20, color="525252"),
            align="left", after=120
        ))
        doc_parts.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')
        def kb_right_cell_xml(kb):
            """Build the inner XML of the right cell: embedded image + caption."""
            parts = []
            idx = len(media_files)
            fn = f"kb_image_{idx+1}.png"
            media_files.append((fn, kb["image_bytes"]))
            rid = f"rIdKbImg{idx}"
            rel_entries.append(
                f'<Relationship Id="{rid}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                f'Target="media/{fn}"/>'
            )
            w_emu, h_emu = 3200000, 2400000
            parts.append(
                '<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
                '<w:r><w:drawing>'
                '<wp:inline xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<wp:extent cx="{w_emu}" cy="{h_emu}"/>'
                f'<wp:docPr id="{idx+2000}" name="KB Image {idx+1}"/>'
                '<wp:cNvGraphicFramePr/>'
                '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
                '<pic:pic>'
                f'<pic:nvPicPr><pic:cNvPr id="{idx+2000}" name="{fn}"/><pic:cNvPicPr/></pic:nvPicPr>'
                f'<pic:blipFill><a:blip r:embed="{rid}"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
                '<pic:spPr>'
                f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{w_emu}" cy="{h_emu}"/></a:xfrm>'
                '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                '</pic:spPr></pic:pic></a:graphicData></a:graphic>'
                '</wp:inline></w:drawing></w:r></w:p>'
            )
            cap = f"{kb.get('caption','')}  ·  {kb.get('source_file','KB')}"[:160]
            parts.append(para_xml(run_xml(cap, size=16, italic=True, color="525252"),
                                   align="center", after=60))
            return "".join(parts)
        def kb_2col_section_xml(section, kb):
            """Whole section wrapped as 50/50 w:tbl: left = text, right = image/card."""
            left = []
            left.append(heading_xml(section.get("title", ""), level=1))
            for p in section.get("paragraphs", []) or []:
                left.append(para_xml(run_xml(p, size=22), align="both"))
            for b in section.get("bullets", []) or []:
                left.append(
                    f'<w:p><w:pPr><w:numPr><w:ilvl w:val="0"/><w:numId w:val="1"/></w:numPr>'
                    f'<w:ind w:left="360"/></w:pPr>{run_xml("• " + str(b), size=22)}</w:p>'
                )
            if section.get("table"):
                t = section["table"]
                left.append(table_xml(t.get("headers", []), t.get("rows", [])))
            return (
                '<w:tbl>'
                '<w:tblPr><w:tblW w:w="5000" w:type="pct"/>'
                '<w:tblBorders><w:top w:val="nil"/><w:bottom w:val="nil"/><w:left w:val="nil"/>'
                '<w:right w:val="nil"/><w:insideH w:val="nil"/><w:insideV w:val="nil"/></w:tblBorders>'
                '</w:tblPr>'
                '<w:tblGrid><w:gridCol w:w="4680"/><w:gridCol w:w="4680"/></w:tblGrid>'
                '<w:tr>'
                '<w:tc><w:tcPr><w:tcW w:w="2500" w:type="pct"/></w:tcPr>'
                + "".join(left) +
                '</w:tc>'
                '<w:tc><w:tcPr><w:tcW w:w="2500" w:type="pct"/></w:tcPr>'
                + kb_right_cell_xml(kb) +
                '</w:tc>'
                '</w:tr></w:tbl>'
            )
        for idx, section in enumerate(sections, start=1):
            if section.get("_kb_match"):
                doc_parts.append(kb_2col_section_xml(section, section["_kb_match"]))
                doc_parts.append(para_xml("", after=240))
                continue
            sec_title = section.get("title", f"Section {idx}")
            doc_parts.append(heading_xml(sec_title, level=1))
            for para in section.get("paragraphs", []) or []:
                doc_parts.append(para_xml(run_xml(para, size=22), align="both"))
            bullets = section.get("bullets", []) or []
            for b in bullets:
                doc_parts.append(
                    f'<w:p><w:pPr><w:numPr><w:ilvl w:val="0"/><w:numId w:val="1"/></w:numPr>'
                    f'<w:ind w:left="360"/></w:pPr>{run_xml("• " + str(b), size=22)}</w:p>'
                )
            if section.get("table"):
                t = section["table"]
                doc_parts.append(table_xml(t.get("headers", []), t.get("rows", [])))
                doc_parts.append(para_xml("", after=120))
            if section.get("_img_bytes"):
                add_image_xml(
                    section["_img_bytes"],
                    section.get("_img_width", 1200),
                    section.get("_img_height", 800),
                    section.get("image_caption") or section.get("title", ""),
                )
            elif section.get("_chart_spec"):
                add_chart_xml(
                    section["_chart_spec"],
                    section.get("image_caption") or section.get("title", ""),
                )
        body_xml = "".join(doc_parts)
        logo_png = self._get_ibm_logo_png()
        footer_ref_xml = ""
        footer_xml = None
        footer_rels_xml = None
        if logo_png:
            lw, lh = self._get_ibm_logo_dims()
            footer_h_emu = 144000
            footer_w_emu = int(footer_h_emu * (lw / max(lh, 1)))
            rel_entries.append(
                '<Relationship Id="rIdFooter" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer" '
                'Target="footer1.xml"/>'
            )
            footer_rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdFooterLogo" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                'Target="media/ibm_logo_black.png"/>'
                '</Relationships>'
            )
            footer_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
                '<w:p>'
                '<w:pPr><w:tabs><w:tab w:val="right" w:pos="9360"/></w:tabs></w:pPr>'
                '<w:r><w:drawing>'
                f'<wp:inline distT="0" distB="0" distL="0" distR="0">'
                f'<wp:extent cx="{footer_w_emu}" cy="{footer_h_emu}"/>'
                '<wp:effectExtent l="0" t="0" r="0" b="0"/>'
                '<wp:docPr id="9999" name="IBM Logo"/>'
                '<wp:cNvGraphicFramePr><a:graphicFrameLocks xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" noChangeAspect="1"/></wp:cNvGraphicFramePr>'
                '<a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                '<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
                '<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
                '<pic:nvPicPr><pic:cNvPr id="9999" name="IBM Logo"/><pic:cNvPicPr/></pic:nvPicPr>'
                '<pic:blipFill><a:blip r:embed="rIdFooterLogo"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
                '<pic:spPr><a:xfrm><a:off x="0" y="0"/>'
                f'<a:ext cx="{footer_w_emu}" cy="{footer_h_emu}"/></a:xfrm>'
                '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
                '</pic:pic></a:graphicData></a:graphic></wp:inline></w:drawing></w:r>'
                '<w:r><w:rPr><w:rFonts w:ascii="IBM Plex Sans" w:hAnsi="IBM Plex Sans"/>'
                '<w:sz w:val="18"/><w:color w:val="525252"/></w:rPr>'
                '<w:t xml:space="preserve">  |  IBM Consulting 2026</w:t></w:r>'
                '<w:r><w:tab/></w:r>'
                '<w:r><w:rPr><w:rFonts w:ascii="IBM Plex Sans" w:hAnsi="IBM Plex Sans"/>'
                '<w:sz w:val="18"/><w:color w:val="525252"/></w:rPr>'
                '<w:t xml:space="preserve">Page </w:t></w:r>'
                '<w:fldSimple w:instr=" PAGE ">'
                '<w:r><w:rPr><w:rFonts w:ascii="IBM Plex Sans" w:hAnsi="IBM Plex Sans"/>'
                '<w:sz w:val="18"/><w:color w:val="525252"/></w:rPr>'
                '<w:t>1</w:t></w:r></w:fldSimple>'
                '<w:r><w:rPr><w:rFonts w:ascii="IBM Plex Sans" w:hAnsi="IBM Plex Sans"/>'
                '<w:sz w:val="18"/><w:color w:val="525252"/></w:rPr>'
                '<w:t xml:space="preserve"> of </w:t></w:r>'
                '<w:fldSimple w:instr=" NUMPAGES ">'
                '<w:r><w:rPr><w:rFonts w:ascii="IBM Plex Sans" w:hAnsi="IBM Plex Sans"/>'
                '<w:sz w:val="18"/><w:color w:val="525252"/></w:rPr>'
                '<w:t>1</w:t></w:r></w:fldSimple>'
                '</w:p>'
                '</w:ftr>'
            )
            footer_ref_xml = '<w:footerReference w:type="default" r:id="rIdFooter"/>'
            media_files.append(("ibm_logo_black.png", logo_png))
        doc_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
            f'<w:body>{body_xml}'
            '<w:sectPr>'
            f'{footer_ref_xml}'
            '<w:pgSz w:w="12240" w:h="15840"/>'
            '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="720" w:footer="720" w:gutter="0"/>'
            '</w:sectPr></w:body></w:document>'
        )
        footer_override = (
            '<Override PartName="/word/footer1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.footer+xml"/>'
            if footer_xml else ''
        )
        ct_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Default Extension="png" ContentType="image/png"/>'
            '<Default Extension="jpeg" ContentType="image/jpeg"/>'
            '<Default Extension="jpg" ContentType="image/jpeg"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            f'{footer_override}'
            + "".join(chart_overrides)
            + '</Types>'
        )
        rels_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>'
        )
        doc_rels = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + "".join(rel_entries)
            + '</Relationships>'
        )
        docx_buf = io.BytesIO()
        with zipfile.ZipFile(docx_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", ct_xml)
            zf.writestr("_rels/.rels", rels_xml)
            zf.writestr("word/_rels/document.xml.rels", doc_rels)
            if footer_xml:
                zf.writestr("word/footer1.xml", footer_xml)
                zf.writestr("word/_rels/footer1.xml.rels", footer_rels_xml)
            zf.writestr("word/document.xml", doc_xml)
            for part_name, xml_bytes in chart_parts:
                zf.writestr(part_name, xml_bytes)
            for fname, fbytes in media_files:
                zf.writestr(f"word/media/{fname}", fbytes)
        docx_bytes = docx_buf.getvalue()
        docx_b64 = base64.b64encode(docx_bytes).decode()
        data_uri = f"data:application/vnd.openxmlformats-officedocument.wordprocessingml.document;base64,{docx_b64}"
        return self._render_docx_preview(title, client_name, sections, data_uri)
    def _render_docx_preview(self, title, client_name, sections, data_uri):
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:50] or "document"
        logo_png = self._get_ibm_logo_png()
        logo_img_tag = ""
        if logo_png:
            logo_b64 = base64.b64encode(logo_png).decode()
            logo_img_tag = f'<img src="data:image/png;base64,{logo_b64}" style="height:11px;width:auto;vertical-align:middle" alt="IBM"/>'
        def footer_html(page_num: int, total_pages: int) -> str:
            return (
                '<div style="position:absolute;bottom:24px;left:60px;right:60px;'
                'display:flex;align-items:center;justify-content:space-between;'
                'padding-top:12px;border-top:1px solid #E0E0E0;'
                'font-family:\\"IBM Plex Sans\\",Calibri,sans-serif;'
                'font-size:10px;color:#525252">'
                f'<div style="display:flex;align-items:center;gap:8px">{logo_img_tag}'
                '<span>|&nbsp;&nbsp;IBM Consulting 2026</span></div>'
                f'<div>Page {page_num} of {total_pages}</div>'
                '</div>'
            )
        total_pages = 1 + len(sections)
        page_parts = []
        page_parts.append(
            f'<div class="pg" style="display:block;padding:80px 60px 80px;background:#fff;min-height:9in;position:relative">'
            f'<div style="font-size:36px;font-weight:700;color:{IBM_BLUE_60};margin-bottom:24px;font-family:\\"IBM Plex Sans\\",Calibri,sans-serif">{self._html_esc(title)}</div>'
            f'<div style="font-size:18px;color:{IBM_GRAY_70};margin-bottom:8px;font-family:\\"IBM Plex Sans\\",Calibri,sans-serif">IBM Consulting  |  Prepared for {self._html_esc(client_name)}</div>'
            f'<div style="font-size:14px;color:{IBM_GRAY_70};font-family:\\"IBM Plex Sans\\",Calibri,sans-serif">{time.strftime("%B %Y")}</div>'
            f'{footer_html(1, total_pages)}'
            f'</div>'
        )
        for idx, section in enumerate(sections, start=1):
            parts = []
            parts.append(
                f'<h1 style="font-size:28px;color:{IBM_BLUE_60};font-weight:700;margin:0 0 16px;'
                f'font-family:\"IBM Plex Sans\",Calibri,sans-serif">{self._html_esc(section.get("title", ""))}</h1>'
            )
            for para in section.get("paragraphs", []) or []:
                parts.append(
                    f'<p style="font-size:12px;color:{IBM_GRAY_100};line-height:1.7;margin:8px 0;text-align:justify;'
                    f'font-family:\"IBM Plex Sans\",Calibri,sans-serif">{self._html_esc(para)}</p>'
                )
            bullets = section.get("bullets", []) or []
            if bullets:
                lis = "".join(
                    f'<li style="font-size:12px;color:{IBM_GRAY_100};margin:4px 0;font-family:\"IBM Plex Sans\",Calibri,sans-serif">{self._html_esc(b)}</li>'
                    for b in bullets
                )
                parts.append(f'<ul style="padding-left:24px;margin:8px 0">{lis}</ul>')
            if section.get("table"):
                t = section["table"]
                tbl = (
                    f'<table style="width:100%;border-collapse:collapse;margin:12px 0;font-size:11px;font-family:\"IBM Plex Sans\",Calibri,sans-serif">'
                )
                if t.get("headers"):
                    tbl += f'<tr style="background:{IBM_BLUE_60}">'
                    tbl += "".join(
                        f'<th style="padding:8px 10px;color:#fff;text-align:left;border:1px solid #ddd">{self._html_esc(h)}</th>'
                        for h in t["headers"]
                    )
                    tbl += '</tr>'
                for ri, row in enumerate(t.get("rows", [])):
                    bg = IBM_GRAY_10 if ri % 2 == 1 else "#FFFFFF"
                    tbl += f'<tr style="background:{bg}">'
                    tbl += "".join(
                        f'<td style="padding:6px 10px;color:#333;border:1px solid #ddd">{self._html_esc(str(c))}</td>'
                        for c in row
                    )
                    tbl += '</tr>'
                tbl += '</table>'
                parts.append(tbl)
            if section.get("_chart_spec"):
                svg = self._svg_chart_from_spec(section["_chart_spec"], width=640, height=320)
                parts.append(
                    f'<div style="text-align:center;margin:16px 0">'
                    f'{svg}'
                    + (
                        f'<div style="font-size:10px;color:{IBM_GRAY_70};font-style:italic;margin-top:6px">'
                        f'Chart — {self._html_esc(section.get("image_caption") or "")}</div>'
                        if section.get("image_caption") else ""
                    )
                    + '</div>'
                )
            elif section.get("_img_bytes"):
                img_b64 = base64.b64encode(section["_img_bytes"]).decode()
                parts.append(
                    f'<div style="text-align:center;margin:16px 0">'
                    f'<img src="data:image/png;base64,{img_b64}" '
                    f'style="max-width:100%;width:480px;height:auto;border-radius:4px"/>'
                    + (
                        f'<div style="font-size:10px;color:{IBM_GRAY_70};font-style:italic;margin-top:6px">'
                        f'Figure — {self._html_esc(section.get("image_caption") or "")}</div>'
                        if section.get("image_caption") else ""
                    )
                    + '</div>'
                )
            page_num = idx + 1
            kb = section.get("_kb_match")
            if kb and kb.get("image_bytes"):
                img_b64 = base64.b64encode(kb["image_bytes"]).decode()
                right = (f'<img src="data:image/png;base64,{img_b64}" '
                         f'style="max-width:100%;border-radius:4px;box-shadow:0 2px 8px rgba(0,0,0,0.1)"/>'
                         f'<div style="font-size:10px;color:{IBM_GRAY_70};font-style:italic;margin-top:8px;text-align:center">'
                         f'📎 {self._html_esc(kb.get("caption",""))}  ·  {self._html_esc(kb.get("source_file",""))}</div>')
                page_body = (f'<div style="display:flex;gap:24px">'
                             f'<div style="flex:1;min-width:0">{"".join(parts)}</div>'
                             f'<div style="flex:1;min-width:0;display:flex;flex-direction:column;justify-content:center">{right}</div>'
                             f'</div>')
            else:
                page_body = "".join(parts)
            page_parts.append(
                f'<div class="pg" style="display:none;padding:60px 60px 80px;background:#fff;min-height:9in;'
                f'position:relative;page-break-after:always">'
                f'{page_body}'
                f'{footer_html(page_num, total_pages)}'
                f'</div>'
            )
        total = len(page_parts)
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&display=swap">
<style>
*{{box-sizing:border-box;margin:0}}
html,body{{height:720px;min-height:720px}}
body{{font-family:\"IBM Plex Sans\",Calibri,system-ui,sans-serif;background:#F0F2F5;padding:12px;display:flex;align-items:stretch;justify-content:center}}
.dk{{border:3px solid {IBM_BLUE_60};border-radius:10px;overflow:hidden;width:100%;max-width:1280px;height:696px;margin:0 auto;background:#FFFFFF;display:flex;flex-direction:column;box-shadow:0 4px 16px rgba(15,98,254,0.15)}}
.tb{{display:flex;align-items:center;gap:8px;padding:10px 14px;background:{IBM_BLUE_60};flex-wrap:wrap;flex-shrink:0}}
.b{{border:none;border-radius:4px;padding:6px 14px;font-size:12px;cursor:pointer;font-family:\"IBM Plex Sans\",Calibri,sans-serif;font-weight:600;text-decoration:none;display:inline-block;background:{IBM_BLUE_60};color:#FFFFFF;border:1px solid #FFFFFF}}
.b:hover{{background:{IBM_BLUE_70};color:#FFFFFF}}
.sn{{color:#FFFFFF;font-size:12px;min-width:90px;text-align:center;font-weight:600}}
.sp{{flex:1}}
.sw{{background:{IBM_GRAY_10};padding:20px;overflow:auto;flex:1;min-height:0}}
.pg{{max-width:8.5in;margin:0 auto 16px;box-shadow:0 2px 8px rgba(0,0,0,0.1);min-height:9in;background:#FFFFFF;transform-origin:top center;transition:transform .15s ease}}
</style></head><body>
<div class="dk">
  <div class="tb">
    <button class="b" onclick="nav(-1)">← Prev</button>
    <span class="sn" id="sn">Page 1 / {total}</span>
    <button class="b" onclick="nav(1)">Next →</button>
    <button class="b" onclick="zoomStep(-0.1)" title="Zoom out">🔍−</button>
    <button class="b" onclick="zoomReset()" title="Reset zoom"><span id="zl">100%</span></button>
    <button class="b" onclick="zoomStep(0.1)" title="Zoom in">🔍+</button>
    <span class="sp"></span>
    <a class="b" href="{data_uri}" download="{safe_name}.docx">⬇ Download DOCX</a>
  </div>
  <div class="sw">{"".join(page_parts)}</div>
</div>
<script>
(function(){{
  function fit(){{
    try{{
      var fe=window.frameElement;
      if(fe){{
        fe.style.height='720px';
        fe.style.minHeight='720px';
        fe.style.width='100%';
        fe.setAttribute('height','720');
      }}
    }}catch(e){{}}
    try{{window.parent.postMessage({{type:'ibm-docgen-resize',height:720}},'*');}}catch(e){{}}
  }}
  fit();setTimeout(fit,100);setTimeout(fit,500);setTimeout(fit,1500);
  window.addEventListener('load',fit);
  new MutationObserver(fit).observe(document.documentElement,{{attributes:true,childList:true,subtree:false}});
}})();
var cur=0,sl=document.querySelectorAll(".pg"),tot=sl.length,zm=1;
function applyZoom(){{sl.forEach(function(p){{p.style.transform='scale('+zm+')'}});var el=document.getElementById('zl');if(el)el.textContent=Math.round(zm*100)+'%'}}
function zoomStep(d){{zm=Math.max(0.5,Math.min(2,zm+d));applyZoom()}}
function zoomReset(){{zm=1;applyZoom()}}
function nav(d){{sl[cur].style.display="none";cur=Math.max(0,Math.min(tot-1,cur+d));
sl[cur].style.display="block";document.getElementById("sn").textContent="Page "+(cur+1)+" / "+tot;
var sw=document.querySelector(".sw");if(sw)sw.scrollTop=0}}
document.addEventListener("keydown",function(e){{
if(e.key==="ArrowLeft")nav(-1);if(e.key==="ArrowRight")nav(1);
if((e.ctrlKey||e.metaKey)&&e.key==="=")e.preventDefault()||zoomStep(0.1);
if((e.ctrlKey||e.metaKey)&&e.key==="-")e.preventDefault()||zoomStep(-0.1);
if((e.ctrlKey||e.metaKey)&&e.key==="0")e.preventDefault()||zoomReset()}});
</script></body></html>"""
        return HTMLResponse(content=html, headers={"Content-Disposition": "inline"})
    def _build_and_render_xlsx(self, session_id, title, client_name, sections, workbook_spec, emitter):
        if not HAS_XLSX: return ("❌ openpyxl is not installed in the Open WebUI Python environment. "
                    "Install it with:\n"
                    "    /Users/pradeepbasavarajappa/.local/share/uv/tools/open-webui/bin/python -m pip install openpyxl\n"
                    "Then restart Open WebUI.")
        sheets = []
        if workbook_spec and isinstance(workbook_spec, dict) and workbook_spec.get("sheets"):
            raw_sheets_list = workbook_spec["sheets"][: self.MAX_SHEETS_XLSX]
            if len(workbook_spec["sheets"]) > self.MAX_SHEETS_XLSX:
                print(f"[DocGen] XLSX capped at {self.MAX_SHEETS_XLSX} sheets (was {len(workbook_spec['sheets'])}).")
            for sh in raw_sheets_list:
                if not isinstance(sh, dict): continue
                cols = []
                if sh.get("columns") and isinstance(sh["columns"], list):
                    for col in sh["columns"]:
                        if isinstance(col, dict):
                            cols.append({"header": str(col.get("header", "")), "width": col.get("width")})
                        else:
                            cols.append({"header": str(col), "width": None})
                elif sh.get("headers"):
                    for h in sh["headers"]:
                        cols.append({"header": str(h), "width": None})
                raw_rows = [list(r) for r in (sh.get("rows") or [])]
                capped_rows = raw_rows[:self.MAX_ROWS_XLSX]
                sheets.append({
                    "title": str(sh.get("title") or sh.get("sheet_name") or "Sheet"),
                    "columns": cols,
                    "headers": [c["header"] for c in cols],
                    "rows": capped_rows,
                    "notes": str(sh.get("notes") or "") + (
                        f"  (truncated from {len(raw_rows)} to {self.MAX_ROWS_XLSX} rows)"
                        if len(raw_rows) > self.MAX_ROWS_XLSX else ""
                    ),
                    "styles": sh.get("styles") or {},
                })
        else:
            summary_rows = []
            for idx, s in enumerate(sections, start=1):
                paras = s.get("paragraphs") or []
                bullets = s.get("bullets") or []
                bullet_preview = " • ".join(bullets[:3]) if bullets else ""
                para_preview = (paras[0] if paras else "")[:300]
                summary_rows.append([idx, s.get("title", ""), para_preview, bullet_preview])
            summary_cols = [
                {"header": "#", "width": 6},
                {"header": "Section", "width": 32},
                {"header": "Overview", "width": 60},
                {"header": "Key Points", "width": 50},
            ]
            sheets.append({
                "title": "Summary",
                "columns": summary_cols,
                "headers": [c["header"] for c in summary_cols],
                "rows": summary_rows,
                "notes": f"Prepared for {client_name} — {time.strftime('%B %Y')}",
                "styles": {},
            })
            for idx, s in enumerate(sections, start=1):
                sec_title = s.get("title", f"Section {idx}")
                tbl = s.get("table") or None
                if tbl and tbl.get("headers") and tbl.get("rows"):
                    cols = [{"header": str(h), "width": None} for h in tbl["headers"]]
                    sheets.append({
                        "title": self._sanitize_sheet_name(sec_title, idx),
                        "columns": cols,
                        "headers": list(tbl["headers"]),
                        "rows": [list(r) for r in tbl["rows"]],
                        "notes": "",
                        "styles": {},
                    })
                else:
                    rows = []
                    for p in s.get("paragraphs") or []:
                        rows.append(["Paragraph", p])
                    for b in s.get("bullets") or []:
                        rows.append(["Bullet", b])
                    if not rows:
                        rows.append(["", ""])
                    sheets.append({
                        "title": self._sanitize_sheet_name(sec_title, idx),
                        "columns": [{"header": "Type", "width": 14}, {"header": "Content", "width": 80}],
                        "headers": ["Type", "Content"],
                        "rows": rows,
                        "notes": sec_title,
                        "styles": {},
                    })
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        GANTT_BLOCKS = frozenset("\u2588\u2593\u2592\u2591\u25a0\u25aa\u25fc\u2b1b")
        RAG_MAP = {"RED": "DA1E28", "AMBER": "F1C21B", "GREEN": "24A148",
                   "RAG:RED": "DA1E28", "RAG:AMBER": "F1C21B", "RAG:GREEN": "24A148"}
        note_font = Font(name="IBM Plex Sans", size=10, italic=True, color="525252")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wrap = Alignment(wrap_text=True, vertical="top")
        used_names = set()
        for sh in sheets:
            name = self._unique_sheet_name(sh["title"], used_names)
            used_names.add(name)
            ws = wb.create_sheet(title=name)
            cols = sh.get("columns") or [{"header": h, "width": None} for h in sh.get("headers", [])]
            headers = [c["header"] for c in cols]
            col_count = max(1, len(headers))
            styles = sh.get("styles") or {}
            hdr_bg = (styles.get("header_bg") or "#0F62FE").lstrip("#")
            hdr_fg = (styles.get("header_fg") or "#FFFFFF").lstrip("#")
            alt_bg = (styles.get("alt_row_bg") or "#F4F4F4").lstrip("#")
            hdr_fill = PatternFill(start_color=hdr_bg, end_color=hdr_bg, fill_type="solid")
            hdr_font = Font(name="IBM Plex Sans", size=11, bold=True, color=hdr_fg)
            body_font = Font(name="IBM Plex Sans", size=11, color="161616")
            alt_fill = PatternFill(start_color=alt_bg, end_color=alt_bg, fill_type="solid")
            row_cursor = 1
            ws.cell(row=row_cursor, column=1, value=str(sh["title"]))
            ws.cell(row=row_cursor, column=1).font = Font(name="IBM Plex Sans", size=16, bold=True, color="0F62FE")
            if col_count > 1:
                ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=col_count)
            ws.row_dimensions[row_cursor].height = 26
            row_cursor += 1
            ws.cell(row=row_cursor, column=1, value=f"IBM Consulting  |  Prepared for {client_name}")
            ws.cell(row=row_cursor, column=1).font = note_font
            if col_count > 1:
                ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=col_count)
            row_cursor += 2
            hdr_row = row_cursor
            for ci, h in enumerate(headers, start=1):
                c = ws.cell(row=hdr_row, column=ci, value=str(h))
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = border
            ws.row_dimensions[hdr_row].height = 28
            row_cursor += 1
            first_data_row = row_cursor
            for ri, row in enumerate(sh["rows"]):
                r = row_cursor + ri
                is_alt = ri % 2 == 1
                for ci in range(col_count):
                    val = row[ci] if ci < len(row) else ""
                    cell_str = str(val) if val is not None else ""
                    stripped_upper = cell_str.strip().upper()
                    is_gantt = bool(cell_str.strip()) and all(ch in GANTT_BLOCKS for ch in cell_str.strip())
                    rag_key = stripped_upper if stripped_upper in RAG_MAP else None
                    c = ws.cell(row=r, column=ci + 1)
                    c.border = border
                    c.font = body_font
                    c.alignment = wrap
                    if is_gantt:
                        c.value = ""
                        c.fill = PatternFill(start_color="24A148", end_color="24A148", fill_type="solid")
                    elif rag_key:
                        c.value = cell_str.strip()
                        c.fill = PatternFill(start_color=RAG_MAP[rag_key], end_color=RAG_MAP[rag_key], fill_type="solid")
                        c.font = Font(name="IBM Plex Sans", size=11, bold=True, color="FFFFFF")
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        coerced = val
                        if isinstance(val, str):
                            s2 = val.strip()
                            if s2 and re.fullmatch(r"-?\d+", s2):
                                try: coerced = int(s2)
                                except Exception: pass
                            elif s2 and re.fullmatch(r"-?\d+\.\d+", s2):
                                try: coerced = float(s2)
                                except Exception: pass
                        c.value = coerced
                        if is_alt:
                            c.fill = alt_fill
            last_data_row = row_cursor + max(0, len(sh["rows"]) - 1)
            if sh.get("notes"):
                notes_row = last_data_row + 2
                ws.cell(row=notes_row, column=1, value=str(sh["notes"]))
                ws.cell(row=notes_row, column=1).font = note_font
                if col_count > 1:
                    ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=col_count)
            for ci in range(1, col_count + 1):
                letter = get_column_letter(ci)
                spec_w = cols[ci - 1].get("width") if ci - 1 < len(cols) else None
                if spec_w:
                    ws.column_dimensions[letter].width = float(spec_w)
                else:
                    max_len = 10
                    if ci - 1 < len(headers):
                        max_len = max(max_len, len(str(headers[ci - 1])))
                    for row in sh["rows"]:
                        if ci - 1 < len(row):
                            cell_str = str(row[ci - 1]) if row[ci - 1] is not None else ""
                            max_len = max(max_len, min(60, len(cell_str)))
                    ws.column_dimensions[letter].width = min(60, max(12, int(max_len * 1.1)))
            ws.freeze_panes = ws.cell(row=first_data_row, column=1).coordinate
            if col_count > 0 and len(sh["rows"]) > 0:
                last_col_letter = get_column_letter(col_count)
                ws.auto_filter.ref = f"A{hdr_row}:{last_col_letter}{last_data_row}"
        if not wb.sheetnames:
            ws = wb.create_sheet(title="Sheet1")
            ws["A1"] = title
            ws["A2"] = f"Prepared for {client_name}"
        try:
            from openpyxl.drawing.image import Image as _XLImg
            logo_png = self._get_ibm_logo_png()
            if logo_png:
                for sheet_idx, sn in enumerate(wb.sheetnames, start=1):
                    ws = wb[sn]
                    try:
                        img_buf = io.BytesIO(logo_png)
                        img_buf.name = "ibm_logo.png"
                        xlimg = _XLImg(img_buf)
                        xlimg.width = 38
                        xlimg.height = 38
                        xlimg.anchor = "A1"
                        ws.add_image(xlimg)
                    except Exception as e:
                        print(f"[DocGen] XLSX logo insert failed for '{sn}': {e}")
                    try:
                        from openpyxl.styles import Font as _XLFont
                        total = len(wb.sheetnames)
                        label_cell = ws.cell(row=1, column=26, value=f"Sheet {sheet_idx} / {total}")
                        label_cell.font = _XLFont(name="IBM Plex Sans", size=9, italic=True, color="525252")
                        label_cell.alignment = __import__("openpyxl").styles.Alignment(horizontal="right")
                    except Exception: pass
        except Exception as e:
            print(f"[DocGen] XLSX logo feature failed: {e}")
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        b64 = base64.b64encode(xlsx_bytes).decode()
        data_uri = (
            "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64," + b64
        )
        return self._render_xlsx_preview(title, client_name, sheets, data_uri)
    MAX_WORDS_PPTX = 100
    MAX_WORDS_DOCX = 300
    MAX_ROWS_XLSX = 100
    MAX_SLIDES_PPTX = 15
    MAX_PAGES_DOCX = 15
    MAX_SHEETS_XLSX = 10
    def _table_has_numeric_column(self, table: dict) -> Optional[int]:
        if not isinstance(table, dict): return None
        rows = table.get("rows") or []
        headers = table.get("headers") or []
        if not rows or not headers or len(headers) < 2: return None
        for col_idx in range(1, len(headers)):
            numeric_count = 0
            for row in rows:
                if col_idx >= len(row): continue
                v = row[col_idx]
                try:
                    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "").replace("₹", "").replace("€", "")
                    float(s)
                    numeric_count += 1
                except (ValueError, TypeError): pass
            if numeric_count >= max(2, int(len(rows) * 0.7)): return col_idx
        return None
    _CHART_PALETTE = ["0F62FE", "8A3FFC", "007D79", "FA4D56", "FF832B",
                      "24A148", "4589FF", "D02670", "161616"]
    def _chart_spec_from_table(self, table: dict, section_title: str = "",
                                chart_type: str = "auto") -> Optional[dict]:
        numeric_col_idx = self._table_has_numeric_column(table)
        if numeric_col_idx is None: return None
        headers = table.get("headers") or []
        rows = table.get("rows") or []
        labels, values = [], []
        for row in rows[:20]:
            if len(row) <= numeric_col_idx: continue
            label = str(row[0])[:22]
            try:
                s = (str(row[numeric_col_idx]).strip()
                     .replace(",", "").replace("$", "")
                     .replace("%", "").replace("₹", "").replace("€", ""))
                values.append(float(s))
                labels.append(label)
            except (ValueError, TypeError): continue
        if len(values) < 2: return None
        if chart_type == "auto":
            col_header = (headers[numeric_col_idx] if numeric_col_idx < len(headers) else "").lower()
            is_share = any(k in col_header for k in
                           ("share", "%", "percent", "proportion", "ratio",
                            "distribution", "mix", "split"))
            total = sum(abs(v) for v in values)
            x_header = (headers[0] if headers else "").lower()
            is_time = any(k in x_header for k in
                          ("year", "month", "quarter", "q1", "q2", "q3", "q4",
                           "date", "period", "fy"))
            if is_time:
                chart_type = "line"
            elif is_share or (len(values) <= 6 and total > 0 and all(v >= 0 for v in values)):
                chart_type = "pie"
            else:
                chart_type = "bar"
        y_label = headers[numeric_col_idx] if numeric_col_idx < len(headers) else "Value"
        x_label = headers[0] if headers else ""
        chart_title = (section_title or y_label or "Chart")[:60]
        return {
            "type": chart_type,
            "title": chart_title,
            "labels": labels,
            "values": values,
            "x_label": x_label,
            "y_label": y_label,
            "series_name": y_label,
        }
    @staticmethod
    def _xml_escape(s) -> str:
        s = str(s)
        return (s.replace("&", "&amp;").replace("<", "&lt;")
                 .replace(">", "&gt;").replace('"', "&quot;"))
    def _ooxml_chart_part_xml(self, spec: dict) -> bytes:
        ctype = spec.get("type", "bar")
        labels = [self._xml_escape(l) for l in spec.get("labels", [])]
        values = list(spec.get("values", []))
        n = len(values)
        title = self._xml_escape(spec.get("title", "Chart"))
        x_label = self._xml_escape(spec.get("x_label", ""))
        y_label = self._xml_escape(spec.get("y_label", ""))
        series_name = self._xml_escape(spec.get("series_name", "Series 1"))
        palette = self._CHART_PALETTE
        cat_pts = "".join(
            f'<c:pt idx="{i}"><c:v>{labels[i]}</c:v></c:pt>' for i in range(n)
        )
        cat_xml = (
            f'<c:cat><c:strRef><c:f>Sheet1!$A$2:$A${n+1}</c:f>'
            f'<c:strCache><c:ptCount val="{n}"/>{cat_pts}</c:strCache>'
            f'</c:strRef></c:cat>'
        )
        val_pts = "".join(
            f'<c:pt idx="{i}"><c:v>{values[i]}</c:v></c:pt>' for i in range(n)
        )
        val_xml = (
            f'<c:val><c:numRef><c:f>Sheet1!$B$2:$B${n+1}</c:f>'
            f'<c:numCache><c:formatCode>General</c:formatCode>'
            f'<c:ptCount val="{n}"/>{val_pts}</c:numCache>'
            f'</c:numRef></c:val>'
        )
        title_xml = (
            '<c:title><c:tx><c:rich>'
            '<a:bodyPr rot="0" spcFirstLastPara="1" vertOverflow="ellipsis" wrap="square" anchor="ctr" anchorCtr="1"/>'
            '<a:lstStyle/>'
            '<a:p><a:pPr><a:defRPr sz="1400" b="1">'
            '<a:solidFill><a:srgbClr val="161616"/></a:solidFill>'
            '<a:latin typeface="IBM Plex Sans"/>'
            '</a:defRPr></a:pPr>'
            f'<a:r><a:rPr lang="en-US" sz="1400" b="1"><a:latin typeface="IBM Plex Sans"/></a:rPr>'
            f'<a:t>{title}</a:t></a:r></a:p></c:rich></c:tx>'
            '<c:overlay val="0"/></c:title>'
        )
        if ctype == "pie":
            dpts = "".join(
                f'<c:dPt><c:idx val="{i}"/><c:bubble3D val="0"/>'
                f'<c:spPr><a:solidFill><a:srgbClr val="{palette[i % len(palette)]}"/>'
                f'</a:solidFill><a:ln w="12700"><a:solidFill><a:srgbClr val="FFFFFF"/>'
                f'</a:solidFill></a:ln></c:spPr></c:dPt>'
                for i in range(n)
            )
            ser_xml = (
                '<c:ser><c:idx val="0"/><c:order val="0"/>'
                f'<c:tx><c:v>{series_name}</c:v></c:tx>{dpts}'
                '<c:dLbls><c:spPr><a:noFill/><a:ln><a:noFill/></a:ln></c:spPr>'
                '<c:txPr><a:bodyPr/><a:lstStyle/>'
                '<a:p><a:pPr><a:defRPr sz="900" b="1">'
                '<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
                '<a:latin typeface="IBM Plex Sans"/></a:defRPr></a:pPr><a:endParaRPr lang="en-US"/></a:p></c:txPr>'
                '<c:showLegendKey val="0"/><c:showVal val="0"/>'
                '<c:showCatName val="0"/><c:showSerName val="0"/>'
                '<c:showPercent val="1"/><c:showBubbleSize val="0"/></c:dLbls>'
                f'{cat_xml}{val_xml}</c:ser>'
            )
            plot_xml = (
                '<c:plotArea><c:layout/>'
                f'<c:pieChart><c:varyColors val="1"/>{ser_xml}'
                '<c:firstSliceAng val="0"/></c:pieChart></c:plotArea>'
            )
        elif ctype == "line":
            ser_xml = (
                '<c:ser><c:idx val="0"/><c:order val="0"/>'
                f'<c:tx><c:v>{series_name}</c:v></c:tx>'
                f'<c:spPr><a:ln w="28575" cap="rnd"><a:solidFill>'
                f'<a:srgbClr val="{palette[0]}"/></a:solidFill><a:round/></a:ln>'
                '</c:spPr>'
                '<c:marker><c:symbol val="circle"/><c:size val="7"/>'
                f'<c:spPr><a:solidFill><a:srgbClr val="{palette[0]}"/></a:solidFill>'
                '<a:ln w="9525"><a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill></a:ln>'
                '</c:spPr></c:marker>'
                f'{cat_xml}{val_xml}<c:smooth val="0"/></c:ser>'
            )
            plot_xml = (
                '<c:plotArea><c:layout/>'
                '<c:lineChart><c:grouping val="standard"/><c:varyColors val="0"/>'
                f'{ser_xml}'
                '<c:marker val="1"/><c:axId val="1"/><c:axId val="2"/></c:lineChart>'
                + self._ooxml_cat_val_axes(x_label, y_label)
                + '</c:plotArea>'
            )
        else:
            dpts = "".join(
                f'<c:dPt><c:idx val="{i}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                f'<c:spPr><a:solidFill><a:srgbClr val="{palette[i % len(palette)]}"/>'
                f'</a:solidFill></c:spPr></c:dPt>'
                for i in range(n)
            )
            ser_xml = (
                '<c:ser><c:idx val="0"/><c:order val="0"/>'
                f'<c:tx><c:v>{series_name}</c:v></c:tx>'
                '<c:invertIfNegative val="0"/>'
                f'{dpts}{cat_xml}{val_xml}</c:ser>'
            )
            plot_xml = (
                '<c:plotArea><c:layout/>'
                '<c:barChart><c:barDir val="col"/><c:grouping val="clustered"/>'
                '<c:varyColors val="1"/>'
                f'{ser_xml}'
                '<c:gapWidth val="80"/><c:axId val="1"/><c:axId val="2"/></c:barChart>'
                + self._ooxml_cat_val_axes(x_label, y_label)
                + '</c:plotArea>'
            )
        chart_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
            'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<c:chart>'
            + title_xml
            + '<c:autoTitleDeleted val="0"/>'
            + plot_xml
            + '<c:plotVisOnly val="1"/><c:dispBlanksAs val="gap"/>'
            '</c:chart>'
            '<c:txPr><a:bodyPr/><a:lstStyle/><a:p><a:pPr>'
            '<a:defRPr sz="1000"><a:latin typeface="IBM Plex Sans"/></a:defRPr>'
            '</a:pPr><a:endParaRPr lang="en-US"/></a:p></c:txPr>'
            '</c:chartSpace>'
        )
        return chart_xml.encode("utf-8")
    @staticmethod
    def _ooxml_cat_val_axes(x_label: str, y_label: str) -> str:
        return (
            '<c:catAx><c:axId val="1"/>'
            '<c:scaling><c:orientation val="minMax"/></c:scaling>'
            '<c:delete val="0"/><c:axPos val="b"/>'
            f'<c:title><c:tx><c:rich><a:bodyPr/><a:lstStyle/>'
            f'<a:p><a:pPr><a:defRPr sz="900"><a:solidFill>'
            f'<a:srgbClr val="525252"/></a:solidFill>'
            f'<a:latin typeface="IBM Plex Sans"/></a:defRPr></a:pPr>'
            f'<a:r><a:rPr lang="en-US" sz="900"/><a:t>{x_label}</a:t></a:r>'
            f'</a:p></c:rich></c:tx><c:overlay val="0"/></c:title>'
            '<c:crossAx val="2"/><c:crosses val="autoZero"/>'
            '<c:auto val="1"/><c:lblAlgn val="ctr"/><c:lblOffset val="100"/>'
            '<c:noMultiLvlLbl val="0"/></c:catAx>'
            '<c:valAx><c:axId val="2"/>'
            '<c:scaling><c:orientation val="minMax"/></c:scaling>'
            '<c:delete val="0"/><c:axPos val="l"/>'
            '<c:majorGridlines><c:spPr><a:ln w="3175">'
            '<a:solidFill><a:srgbClr val="E0E0E0"/></a:solidFill>'
            '<a:prstDash val="dash"/></a:ln></c:spPr></c:majorGridlines>'
            f'<c:title><c:tx><c:rich><a:bodyPr rot="-5400000"/><a:lstStyle/>'
            f'<a:p><a:pPr><a:defRPr sz="900"><a:solidFill>'
            f'<a:srgbClr val="525252"/></a:solidFill>'
            f'<a:latin typeface="IBM Plex Sans"/></a:defRPr></a:pPr>'
            f'<a:r><a:rPr lang="en-US" sz="900"/><a:t>{y_label}</a:t></a:r>'
            f'</a:p></c:rich></c:tx><c:overlay val="0"/></c:title>'
            '<c:numFmt formatCode="General" sourceLinked="0"/>'
            '<c:crossAx val="1"/><c:crosses val="autoZero"/>'
            '<c:crossBetween val="between"/></c:valAx>'
        )
    def _svg_chart_from_spec(self, spec: dict, width: int = 720, height: int = 360) -> str:
        ctype = spec.get("type", "bar")
        labels = spec.get("labels", []) or []
        values = spec.get("values", []) or []
        title = self._xml_escape(spec.get("title", ""))
        palette = self._CHART_PALETTE
        if not values: return ""
        esc = self._xml_escape
        if ctype == "pie":
            import math
            cx, cy, r = width // 2, height // 2 + 10, min(width, height) // 2 - 40
            total = sum(max(0, v) for v in values) or 1.0
            angle0 = -math.pi / 2
            slices = []
            legend = []
            for i, (lab, v) in enumerate(zip(labels, values)):
                pct = max(0, v) / total
                angle1 = angle0 + pct * 2 * math.pi
                x1, y1 = cx + r * math.cos(angle0), cy + r * math.sin(angle0)
                x2, y2 = cx + r * math.cos(angle1), cy + r * math.sin(angle1)
                large = 1 if (angle1 - angle0) > math.pi else 0
                color = palette[i % len(palette)]
                slices.append(
                    f'<path d="M{cx},{cy} L{x1:.1f},{y1:.1f} A{r},{r} 0 {large} 1 {x2:.1f},{y2:.1f} Z" '
                    f'fill="#{color}" stroke="#FFFFFF" stroke-width="1.5"/>'
                )
                mid = (angle0 + angle1) / 2
                tx, ty = cx + (r * 0.62) * math.cos(mid), cy + (r * 0.62) * math.sin(mid)
                slices.append(
                    f'<text x="{tx:.0f}" y="{ty:.0f}" fill="#FFFFFF" font-size="11" '
                    f'font-weight="700" text-anchor="middle" dy="4" '
                    f'font-family="IBM Plex Sans, Arial">{pct*100:.1f}%</text>'
                )
                legend.append((lab, color))
                angle0 = angle1
            legend_xml = ""
            ly = 40
            for lab, color in legend[:8]:
                legend_xml += (
                    f'<rect x="{width-170}" y="{ly-10}" width="12" height="12" fill="#{color}"/>'
                    f'<text x="{width-150}" y="{ly}" font-size="11" fill="#161616" '
                    f'font-family="IBM Plex Sans, Arial">{esc(lab)[:22]}</text>'
                )
                ly += 20
            body = "".join(slices) + legend_xml
        elif ctype == "line":
            pad_l, pad_r, pad_t, pad_b = 60, 30, 50, 50
            w, h = width - pad_l - pad_r, height - pad_t - pad_b
            vmax = max(values) or 1
            vmin = min(min(values), 0)
            rng = (vmax - vmin) or 1
            def xpt(i): return pad_l + (i / max(1, len(values) - 1)) * w
            def ypt(v): return pad_t + h - ((v - vmin) / rng) * h
            pts = " ".join(f"{xpt(i):.1f},{ypt(v):.1f}" for i, v in enumerate(values))
            area_pts = f"{pad_l},{pad_t + h} " + pts + f" {pad_l + w},{pad_t + h}"
            axis_y = []
            for i in range(5):
                gy = pad_t + h - (i / 4) * h
                gv = vmin + (i / 4) * rng
                axis_y.append(
                    f'<line x1="{pad_l}" y1="{gy:.0f}" x2="{pad_l + w}" y2="{gy:.0f}" '
                    f'stroke="#E0E0E0" stroke-dasharray="3,3"/>'
                    f'<text x="{pad_l - 8}" y="{gy + 4:.0f}" text-anchor="end" '
                    f'font-size="10" fill="#525252" font-family="IBM Plex Sans, Arial">{gv:.0f}</text>'
                )
            x_labels = []
            step = max(1, len(labels) // 8)
            for i, lab in enumerate(labels):
                if i % step != 0 and i != len(labels) - 1: continue
                x_labels.append(
                    f'<text x="{xpt(i):.0f}" y="{pad_t + h + 18:.0f}" text-anchor="middle" '
                    f'font-size="10" fill="#525252" font-family="IBM Plex Sans, Arial">{esc(lab)[:12]}</text>'
                )
            markers = "".join(
                f'<circle cx="{xpt(i):.1f}" cy="{ypt(v):.1f}" r="4" fill="#{palette[0]}" stroke="#FFFFFF" stroke-width="2"/>'
                for i, v in enumerate(values)
            )
            body = (
                "".join(axis_y)
                + f'<polygon points="{area_pts}" fill="#{palette[0]}" fill-opacity="0.15"/>'
                + f'<polyline points="{pts}" fill="none" stroke="#{palette[0]}" stroke-width="2.5"/>'
                + markers
                + "".join(x_labels)
            )
        else:
            pad_l, pad_r, pad_t, pad_b = 60, 30, 50, 60
            w, h = width - pad_l - pad_r, height - pad_t - pad_b
            vmax = max(values) or 1
            vmin = min(min(values), 0)
            rng = (vmax - vmin) or 1
            bar_w = w / max(1, len(values)) * 0.75
            gap = (w / max(1, len(values))) - bar_w
            bars = []
            axis_y = []
            for i in range(5):
                gy = pad_t + h - (i / 4) * h
                gv = vmin + (i / 4) * rng
                axis_y.append(
                    f'<line x1="{pad_l}" y1="{gy:.0f}" x2="{pad_l + w}" y2="{gy:.0f}" '
                    f'stroke="#E0E0E0" stroke-dasharray="3,3"/>'
                    f'<text x="{pad_l - 8}" y="{gy + 4:.0f}" text-anchor="end" '
                    f'font-size="10" fill="#525252" font-family="IBM Plex Sans, Arial">{gv:.0f}</text>'
                )
            for i, (lab, v) in enumerate(zip(labels, values)):
                bx = pad_l + gap / 2 + i * (bar_w + gap)
                bh = ((v - vmin) / rng) * h
                by = pad_t + h - bh
                color = palette[i % len(palette)]
                bars.append(
                    f'<rect x="{bx:.1f}" y="{by:.1f}" width="{bar_w:.1f}" height="{bh:.1f}" fill="#{color}"/>'
                    f'<text x="{bx + bar_w/2:.1f}" y="{by - 4:.0f}" text-anchor="middle" '
                    f'font-size="10" fill="#161616" font-family="IBM Plex Sans, Arial">{v:,.0f}</text>'
                    f'<text x="{bx + bar_w/2:.1f}" y="{pad_t + h + 18:.0f}" text-anchor="middle" '
                    f'font-size="10" fill="#525252" font-family="IBM Plex Sans, Arial">{esc(lab)[:12]}</text>'
                )
            body = "".join(axis_y) + "".join(bars)
        return (
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" '
            f'width="100%" style="max-width:{width}px;background:#FFFFFF;">'
            f'<text x="{width//2}" y="24" text-anchor="middle" font-size="14" '
            f'font-weight="700" fill="#161616" font-family="IBM Plex Sans, Arial">{title}</text>'
            f'{body}</svg>'
        )
    def _autoinject_charts(self, sections: list) -> list:
        out = []
        for s in sections:
            if not isinstance(s, dict):
                out.append(s); continue
            ns = dict(s)
            tbl = ns.get("table") or ns.get("chart_data")
            chart_type = ns.get("chart_type", "auto")
            if (tbl and not ns.get("_img_bytes") and not ns.get("image_id")
                    and not ns.get("svg") and not ns.get("_chart_spec")):
                spec = self._chart_spec_from_table(tbl, ns.get("title", ""), chart_type=chart_type)
                if spec:
                    ns["_chart_spec"] = spec
                    ns["_img_source"] = f"generated:ooxml-chart:{spec['type']}"
                    ns["image_caption"] = ns.get("image_caption") or f"Chart — {ns.get('title','data')}"
            out.append(ns)
        return out
    _BULLET_NUMBER_PREFIX = re.compile(r'^\s*(?:\d{1,3}[\.\)\-:]|\d{2}\s|\u2022\s|[-\*]\s|[a-zA-Z][\.\)])\s*')
    def _strip_bullet_numbering(self, bullet: str) -> str:
        if not isinstance(bullet, str): return bullet
        cleaned = self._BULLET_NUMBER_PREFIX.sub('', bullet).strip()
        return cleaned or bullet
    def _enforce_content_caps(self, sections: list, fmt: str) -> list:
        out = []
        if fmt == "pptx":
            limit = self.MAX_WORDS_PPTX
            for s in sections:
                if not isinstance(s, dict): out.append(s); continue
                ns = dict(s)
                title = ns.get("title") or ""
                paras = list(ns.get("paragraphs") or [])
                bullets = [self._strip_bullet_numbering(b) for b in (ns.get("bullets") or [])]
                used = len(title.split())
                new_paras, new_bullets = [], []
                for p in paras:
                    w = (p or "").split()
                    if used + len(w) <= limit:
                        new_paras.append(p); used += len(w)
                    else:
                        remaining = max(0, limit - used)
                        if remaining > 3:
                            new_paras.append(" ".join(w[:remaining]) + "…")
                            used = limit
                        break
                for b in bullets:
                    w = (b or "").split()
                    if used + len(w) <= limit:
                        new_bullets.append(b); used += len(w)
                    else:
                        remaining = max(0, limit - used)
                        if remaining > 2:
                            new_bullets.append(" ".join(w[:remaining]) + "…")
                            used = limit
                        break
                ns["paragraphs"] = new_paras
                ns["bullets"] = new_bullets
                out.append(ns)
        elif fmt == "docx":
            limit = self.MAX_WORDS_DOCX
            for s in sections:
                if not isinstance(s, dict): out.append(s); continue
                ns = dict(s)
                title = ns.get("title") or ""
                paras = list(ns.get("paragraphs") or [])
                bullets = [self._strip_bullet_numbering(b) for b in (ns.get("bullets") or [])]
                used = len(title.split())
                new_paras, new_bullets = [], []
                for p in paras:
                    w = (p or "").split()
                    if used + len(w) <= limit:
                        new_paras.append(p); used += len(w)
                    else:
                        remaining = max(0, limit - used)
                        if remaining > 5:
                            new_paras.append(" ".join(w[:remaining]) + "…")
                            used = limit
                        break
                for b in bullets:
                    w = (b or "").split()
                    if used + len(w) <= limit:
                        new_bullets.append(b); used += len(w)
                    else:
                        remaining = max(0, limit - used)
                        if remaining > 3:
                            new_bullets.append(" ".join(w[:remaining]) + "…")
                            used = limit
                        break
                ns["paragraphs"] = new_paras
                ns["bullets"] = new_bullets
                out.append(ns)
        elif fmt == "xlsx":
            limit = self.MAX_ROWS_XLSX
            for s in sections:
                if not isinstance(s, dict): out.append(s); continue
                ns = dict(s)
                tbl = ns.get("table")
                if isinstance(tbl, dict) and tbl.get("rows"):
                    rows = tbl["rows"]
                    if len(rows) > limit:
                        ns["table"] = {**tbl, "rows": rows[:limit]}
                out.append(ns)
        else: return list(sections)
        return out
    def _sanitize_sheet_name(self, name: str, idx: int = 0) -> str:
        cleaned = re.sub(r"[:\\/\?\*\[\]]", " ", str(name or "")).strip()
        if not cleaned:
            cleaned = f"Sheet{idx}"
        return cleaned[:31]
    def _unique_sheet_name(self, name: str, used: set) -> str:
        base = self._sanitize_sheet_name(name, 1)
        if base not in used: return base
        i = 2
        while True:
            cand = f"{base[:28]} ({i})"
            if cand not in used: return cand
            i += 1
    def _render_xlsx_preview(self, title, client_name, sheets, data_uri):
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:50] or "workbook"
        tab_buttons = []
        panels = []
        for i, sh in enumerate(sheets):
            active = " active" if i == 0 else ""
            display = "block" if i == 0 else "none"
            tab_buttons.append(
                f'<button class="tab{active}" onclick="showTab({i})" data-i="{i}">'
                f'{self._html_esc(sh["title"])}</button>'
            )
            headers = sh.get("headers") or []
            rows = sh.get("rows") or []
            notes = sh.get("notes") or ""
            thead = ""
            if headers:
                thead = "<tr>" + "".join(
                    f'<th style="background:{IBM_BLUE_60};color:#fff;padding:8px 12px;'
                    f'text-align:left;border:1px solid #ddd;font-weight:600;font-size:12px">'
                    f'{self._html_esc(h)}</th>' for h in headers
                ) + "</tr>"
            tbody_rows = []
            for ri, row in enumerate(rows):
                bg = IBM_GRAY_10 if ri % 2 == 1 else "#fff"
                tds = "".join(
                    f'<td style="padding:6px 12px;border:1px solid #e0e0e0;'
                    f'font-size:12px;color:{IBM_GRAY_100};vertical-align:top">'
                    f'{self._html_esc(str(c) if c is not None else "")}</td>'
                    for c in row
                )
                tbody_rows.append(f'<tr style="background:{bg}">{tds}</tr>')
            table_html = (
                f'<table style="border-collapse:collapse;width:100%;font-family:\"IBM Plex Sans\",Calibri,sans-serif">'
                f'{thead}{"".join(tbody_rows)}</table>'
            )
            note_html = (
                f'<div style="font-size:11px;color:{IBM_GRAY_70};font-style:italic;margin-top:12px">'
                f'{self._html_esc(notes)}</div>' if notes else ""
            )
            panels.append(
                f'<div class="panel" data-i="{i}" style="display:{display}">'
                f'<div style="font-size:20px;font-weight:700;color:{IBM_BLUE_60};margin-bottom:4px">'
                f'{self._html_esc(sh["title"])}</div>'
                f'<div style="font-size:12px;color:{IBM_GRAY_70};margin-bottom:16px">'
                f'IBM Consulting  |  Prepared for {self._html_esc(client_name)}</div>'
                f'{table_html}{note_html}</div>'
            )
        total = len(sheets)
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&display=swap">
<style>
*{{box-sizing:border-box;margin:0}}
html,body{{height:720px;min-height:720px}}
body{{font-family:\"IBM Plex Sans\",Calibri,system-ui,sans-serif;background:#F0F2F5;padding:12px;display:flex;align-items:stretch;justify-content:center}}
.wk{{border:3px solid {IBM_BLUE_60};border-radius:10px;overflow:hidden;width:100%;max-width:1280px;height:696px;margin:0 auto;background:#FFFFFF;display:flex;flex-direction:column;box-shadow:0 4px 16px rgba(15,98,254,0.15)}}
.tb{{display:flex;align-items:center;gap:8px;padding:10px 14px;background:{IBM_BLUE_60};flex-wrap:wrap;flex-shrink:0}}
.b{{border:none;border-radius:4px;padding:6px 14px;font-size:12px;cursor:pointer;font-family:\"IBM Plex Sans\",Calibri,sans-serif;font-weight:600;text-decoration:none;display:inline-block;background:{IBM_BLUE_60};color:#FFFFFF;border:1px solid #FFFFFF}}
.b:hover{{background:{IBM_BLUE_70};color:#FFFFFF}}
.sp{{flex:1}}
.title{{color:#FFFFFF;font-size:13px;font-weight:600}}
.tabs{{display:flex;gap:2px;padding:8px 14px 0;background:{IBM_GRAY_10};flex-wrap:wrap;border-bottom:1px solid {IBM_GRAY_20};flex-shrink:0}}
.tab{{border:1px solid {IBM_GRAY_20};border-bottom:none;border-radius:6px 6px 0 0;padding:8px 14px;font-size:12px;cursor:pointer;font-family:\"IBM Plex Sans\",Calibri,sans-serif;font-weight:600;background:#FFFFFF;color:{IBM_GRAY_70}}}
.tab.active{{background:{IBM_BLUE_60};color:#FFFFFF;border-color:{IBM_BLUE_60}}}
.sw{{background:{IBM_GRAY_10};padding:20px;overflow:auto;flex:1;min-height:0}}
.panel{{background:#FFFFFF;padding:24px;border-radius:6px;box-shadow:0 2px 8px rgba(0,0,0,0.08);transform-origin:top center;transition:transform .15s ease}}
</style></head><body>
<div class="wk">
  <div class="tb">
    <span class="title">📊 {self._html_esc(title)}</span>
    <span class="sp"></span>
    <span class="title">{total} sheet{"s" if total != 1 else ""}</span>
    <button class="b" onclick="zoomStep(-0.1)" title="Zoom out">🔍−</button>
    <button class="b" onclick="zoomReset()" title="Reset zoom"><span id="zl">100%</span></button>
    <button class="b" onclick="zoomStep(0.1)" title="Zoom in">🔍+</button>
    <a class="b" href="{data_uri}" download="{safe_name}.xlsx">⬇ Download XLSX</a>
  </div>
  <div class="tabs">{"".join(tab_buttons)}</div>
  <div class="sw">{"".join(panels)}</div>
</div>
<script>
(function(){{
  function fit(){{
    try{{
      var fe=window.frameElement;
      if(fe){{
        fe.style.height='720px';
        fe.style.minHeight='720px';
        fe.style.width='100%';
        fe.setAttribute('height','720');
      }}
    }}catch(e){{}}
    try{{window.parent.postMessage({{type:'ibm-docgen-resize',height:720}},'*');}}catch(e){{}}
  }}
  fit();setTimeout(fit,100);setTimeout(fit,500);setTimeout(fit,1500);
  window.addEventListener('load',fit);
  new MutationObserver(fit).observe(document.documentElement,{{attributes:true,childList:true,subtree:false}});
}})();
var zm=1;
function applyZoom(){{document.querySelectorAll(".panel").forEach(function(p){{p.style.transform='scale('+zm+')'}});var el=document.getElementById('zl');if(el)el.textContent=Math.round(zm*100)+'%'}}
function zoomStep(d){{zm=Math.max(0.5,Math.min(2,zm+d));applyZoom()}}
function zoomReset(){{zm=1;applyZoom()}}
function showTab(i){{
  document.querySelectorAll(".tab").forEach(function(t){{
    t.classList.toggle("active", parseInt(t.dataset.i,10)===i);
  }});
  document.querySelectorAll(".panel").forEach(function(p){{
    p.style.display = (parseInt(p.dataset.i,10)===i) ? "block" : "none";
  }});
}}
document.addEventListener("keydown",function(e){{
if((e.ctrlKey||e.metaKey)&&e.key==="=")e.preventDefault()||zoomStep(0.1);
if((e.ctrlKey||e.metaKey)&&e.key==="-")e.preventDefault()||zoomStep(-0.1);
if((e.ctrlKey||e.metaKey)&&e.key==="0")e.preventDefault()||zoomReset()}});
</script></body></html>"""
        return HTMLResponse(content=html, headers={"Content-Disposition": "inline"})
    def _build_and_render_pptx(self, session_id, title, client_name, sections, emitter):
        SLIDE_W_EMU = 12192000
        SLIDE_H_EMU = 6858000
        def esc(s):
            return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        def txt_run(text, size=1800, bold=False, italic=False, color="161616"):
            b_attr = 'b="1" ' if bold else ''
            i_attr = 'i="1" ' if italic else ''
            return (
                f'<a:r><a:rPr lang="en-US" sz="{size}" '
                f'{b_attr}{i_attr}>'
                f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
                f'<a:latin typeface="IBM Plex Sans"/>'
                f'</a:rPr><a:t>{esc(text)}</a:t></a:r>'
            )
        def txt_box(x, y, w, h, text, size=1800, bold=False, color="161616", align="l"):
            return (
                f'<p:sp>'
                f'<p:nvSpPr><p:cNvPr id="{uuid.uuid4().int % 10000}" name="TxtBox"/>'
                f'<p:cNvSpPr txBox="1"/><p:nvPr/></p:nvSpPr>'
                f'<p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{w}" cy="{h}"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/></p:spPr>'
                f'<p:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
                f'<a:p><a:pPr algn="{align}"><a:lnSpc><a:spcPct val="150000"/></a:lnSpc><a:spcBef><a:spcPts val="300"/></a:spcBef><a:spcAft><a:spcPts val="300"/></a:spcAft></a:pPr>'
                + txt_run(text, size=size, bold=bold, color=color)
                + '</a:p></p:txBody></p:sp>'
            )
        def bullet_box(x, y, w, h, bullets, size=1400):
            body = ""
            for b in bullets:
                body += (
                    f'<a:p><a:pPr marL="285750" indent="-285750"><a:lnSpc><a:spcPct val="150000"/></a:lnSpc><a:spcBef><a:spcPts val="400"/></a:spcBef><a:spcAft><a:spcPts val="400"/></a:spcAft><a:buChar char="•"/></a:pPr>'
                    + txt_run(str(b), size=size) + '</a:p>'
                )
            return (
                f'<p:sp><p:nvSpPr><p:cNvPr id="{uuid.uuid4().int % 10000}" name="Bullets"/>'
                f'<p:cNvSpPr txBox="1"/><p:nvPr/></p:nvSpPr>'
                f'<p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{w}" cy="{h}"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/></p:spPr>'
                f'<p:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>{body}</p:txBody></p:sp>'
            )
        def solid_bg(color):
            return (
                f'<p:sp><p:nvSpPr><p:cNvPr id="9999" name="bg"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>'
                f'<p:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{SLIDE_W_EMU}" cy="{SLIDE_H_EMU}"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill><a:ln><a:noFill/></a:ln></p:spPr></p:sp>'
            )
        def accent_bar():
            return (
                f'<p:sp><p:nvSpPr><p:cNvPr id="9998" name="bar"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>'
                f'<p:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="228600" cy="{SLIDE_H_EMU}"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                f'<a:solidFill><a:srgbClr val="0F62FE"/></a:solidFill><a:ln><a:noFill/></a:ln></p:spPr></p:sp>'
            )
        def image_xml(rid, x, y, w_emu, h_emu):
            return (
                f'<p:pic><p:nvPicPr><p:cNvPr id="{uuid.uuid4().int % 10000}" name="Image"/>'
                f'<p:cNvPicPr/><p:nvPr/></p:nvPicPr>'
                f'<p:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:embed="{rid}"/>'
                f'<a:stretch><a:fillRect/></a:stretch></p:blipFill>'
                f'<p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{w_emu}" cy="{h_emu}"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></p:spPr></p:pic>'
            )
        slides_xml = []
        slide_rels = []
        media_files = []
        chart_parts = []
        chart_overrides = []
        def chart_graphic_frame(spec, slide_rel_entries, x_emu, y_emu, w_emu, h_emu, shape_id):
            chart_idx = len(chart_parts) + 1
            chart_xml_bytes = self._ooxml_chart_part_xml(spec)
            chart_parts.append((f"ppt/charts/chart{chart_idx}.xml", chart_xml_bytes))
            chart_overrides.append(
                f'<Override PartName="/ppt/charts/chart{chart_idx}.xml" '
                f'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
            )
            rid = f"rId{len(slide_rel_entries) + 2}"
            slide_rel_entries.append(
                f'<Relationship Id="{rid}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
                f'Target="../charts/chart{chart_idx}.xml"/>'
            )
            return (
                '<p:graphicFrame>'
                f'<p:nvGraphicFramePr><p:cNvPr id="{shape_id}" name="Chart {chart_idx}"/>'
                '<p:cNvGraphicFramePr/><p:nvPr/></p:nvGraphicFramePr>'
                f'<p:xfrm><a:off x="{x_emu}" y="{y_emu}"/>'
                f'<a:ext cx="{w_emu}" cy="{h_emu}"/></p:xfrm>'
                '<a:graphic>'
                '<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
                f'<c:chart xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart" '
                f'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                f'r:id="{rid}"/>'
                '</a:graphicData></a:graphic></p:graphicFrame>'
            )
        logo_png = self._get_ibm_logo_png()
        logo_fname = None
        if logo_png:
            logo_fname = "ibm_logo_black.png"
            media_files.append((logo_fname, logo_png))
            lw, lh = self._get_ibm_logo_dims()
            logo_h_emu = 144000
            logo_w_emu = int(logo_h_emu * (lw / max(lh, 1)))
            logo_x_emu = 228600
            logo_y_emu = SLIDE_H_EMU - logo_h_emu - 114300
        def logo_shape_and_rel(slide_rel_entries: list) -> str:
            if not logo_fname: return ""
            rid = f"rId{len(slide_rel_entries)+2}"
            slide_rel_entries.append(
                f'<Relationship Id="{rid}" '
                f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                f'Target="../media/{logo_fname}"/>'
            )
            return (
                f'<p:pic><p:nvPicPr><p:cNvPr id="{uuid.uuid4().int % 10000}" name="IBM Logo"/>'
                f'<p:cNvPicPr><a:picLocks noChangeAspect="1"/></p:cNvPicPr><p:nvPr/></p:nvPicPr>'
                f'<p:blipFill><a:blip r:embed="{rid}"/><a:stretch><a:fillRect/></a:stretch></p:blipFill>'
                f'<p:spPr><a:xfrm><a:off x="{logo_x_emu}" y="{logo_y_emu}"/>'
                f'<a:ext cx="{logo_w_emu}" cy="{logo_h_emu}"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></p:spPr></p:pic>'
            )
        cover_xml = (
            '<p:sp><p:nvSpPr><p:cNvPr id="1" name="CoverBG"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr>'
            f'<p:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{SLIDE_W_EMU}" cy="{SLIDE_H_EMU}"/></a:xfrm>'
            '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            '<a:solidFill><a:srgbClr val="0F62FE"/></a:solidFill><a:ln><a:noFill/></a:ln></p:spPr></p:sp>'
            + txt_box(685800, 2286000, 10820400, 1143000, title, size=3600, bold=True, color="FFFFFF")
            + txt_box(685800, 3657600, 10820400, 914400,
                      f"IBM Consulting  |  Prepared for {client_name}", size=1800, color="FFFFFF")
            + txt_box(685800, 5486400, 10820400, 457200,
                      time.strftime("%B %Y"), size=1200, color="FFFFFF")
        )
        _layout_rel = (
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideLayout" '
            'Target="../slideLayouts/slideLayout1.xml"/>'
        )
        slides_xml.append(("slide1", cover_xml, [_layout_rel]))
        for idx, section in enumerate(sections, start=1):
            slide_num = idx + 1
            slide_rel_entries = [_layout_rel]
            shapes = []
            shapes.append(accent_bar())
            shapes.append(
                txt_box(457200, 304800, 11277600, 685800,
                        section.get("title", f"Section {idx}"),
                        size=2600, bold=True, color="161616")
            )
            has_kb = bool(section.get("_kb_match"))
            has_chart = bool(section.get("_chart_spec"))
            has_image = bool(section.get("_img_bytes")) or has_chart or has_kb
            text_x = 457200
            text_y = 1143000
            text_w = 5943600 if has_image else 11277600
            text_h = 5334000
            paras = section.get("paragraphs", []) or []
            bullets = section.get("bullets", []) or []
            if paras or bullets:
                body = ""
                for p in paras:
                    body += f'<a:p><a:pPr algn="l"><a:lnSpc><a:spcPct val="150000"/></a:lnSpc><a:spcBef><a:spcPts val="300"/></a:spcBef><a:spcAft><a:spcPts val="300"/></a:spcAft></a:pPr>' + txt_run(p, size=1400) + '</a:p>'
                for b in bullets:
                    body += (
                        f'<a:p><a:pPr marL="285750" indent="-285750" algn="l">'
                        f'<a:lnSpc><a:spcPct val="150000"/></a:lnSpc>'
                        f'<a:spcBef><a:spcPts val="400"/></a:spcBef>'
                        f'<a:spcAft><a:spcPts val="400"/></a:spcAft>'
                        f'<a:buChar char="•"/></a:pPr>' + txt_run(str(b), size=1400) + '</a:p>'
                    )
                shapes.append(
                    f'<p:sp><p:nvSpPr><p:cNvPr id="{idx*100+1}" name="Body"/>'
                    f'<p:cNvSpPr txBox="1"/><p:nvPr/></p:nvSpPr>'
                    f'<p:spPr><a:xfrm><a:off x="{text_x}" y="{text_y}"/><a:ext cx="{text_w}" cy="{text_h}"/></a:xfrm>'
                    f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/></p:spPr>'
                    f'<p:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>{body}</p:txBody></p:sp>'
                )
            if has_kb:
                kb = section["_kb_match"]
                ch_x, ch_y, ch_w = 6629400, 1143000, 5105400
                media_idx = len(media_files)
                fname_kb = f"kb_image_{media_idx+1}.png"
                media_files.append((fname_kb, kb["image_bytes"]))
                rid = f"rId{len(slide_rel_entries)+2}"
                slide_rel_entries.append(
                    f'<Relationship Id="{rid}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                    f'Target="../media/{fname_kb}"/>'
                )
                shapes.append(image_xml(rid, ch_x, ch_y, ch_w, 3600000))
                cap = (f"📎 {kb.get('caption','')}  ·  from {kb.get('source_file','KB')}")[:120]
                shapes.append(txt_box(ch_x, ch_y + 3600000 + 50000, ch_w, 400000,
                                       cap, size=900, color="525252"))
            elif has_chart:
                ch_x = 6629400
                ch_y = 1143000
                ch_w = 5105400
                ch_h = 3500000
                shapes.append(chart_graphic_frame(
                    section["_chart_spec"], slide_rel_entries,
                    ch_x, ch_y, ch_w, ch_h, shape_id=idx * 100 + 5,
                ))
                if section.get("image_caption"):
                    shapes.append(
                        txt_box(ch_x, ch_y + ch_h + 50000, ch_w, 400000,
                                f"Chart — {section['image_caption']}",
                                size=900, color="525252")
                    )
            elif section.get("_img_bytes"):
                media_idx = len(media_files)
                fname = f"image{media_idx+1}.png"
                media_files.append((fname, section["_img_bytes"]))
                rid = f"rId{len(slide_rel_entries)+2}"
                slide_rel_entries.append(
                    f'<Relationship Id="{rid}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
                    f'Target="../media/{fname}"/>'
                )
                img_x = 6629400
                img_y = 1143000
                img_max_w = 5105400
                img_max_h = 4572000
                aspect = section.get("_img_height", 800) / max(section.get("_img_width", 1200), 1)
                img_w = img_max_w
                img_h = int(img_w * aspect)
                if img_h > img_max_h:
                    img_h = img_max_h
                    img_w = int(img_h / aspect)
                shapes.append(image_xml(rid, img_x, img_y, img_w, img_h))
                if section.get("image_caption"):
                    shapes.append(
                        txt_box(img_x, img_y + img_h + 50000, img_w, 400000,
                                f"Figure — {section['image_caption']}",
                                size=900, color="525252")
                    )
            logo_xml = logo_shape_and_rel(slide_rel_entries)
            if logo_xml:
                shapes.append(logo_xml)
            if logo_fname:
                label_x = logo_x_emu + logo_w_emu + 50800
                label_y = logo_y_emu - 15000
                label_w = 2400000
                label_h = logo_h_emu + 30000
                shapes.append(txt_box(
                    label_x, label_y, label_w, label_h,
                    "|  IBM Consulting 2026", size=900, color="525252", align="l",
                ))
            page_num_text = f"{slide_num} / {len(sections) + 1}"
            page_num_x = SLIDE_W_EMU - 914400 - 228600
            page_num_y = SLIDE_H_EMU - 457200 - 114300
            shapes.append(txt_box(
                page_num_x, page_num_y, 914400, 457200,
                page_num_text, size=900, color="8D8D8D", align="r",
            ))
            slide_content = "".join(shapes)
            slides_xml.append((f"slide{slide_num}", slide_content, slide_rel_entries))
        def wrap_slide(content):
            return (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
                '<p:cSld><p:spTree>'
                '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
                '<p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/>'
                '<a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>'
                f'{content}'
                '</p:spTree></p:cSld></p:sld>'
            )
        pptx_buf = io.BytesIO()
        with zipfile.ZipFile(pptx_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            overrides = ''.join(
                f'<Override PartName="/ppt/slides/slide{i+1}.xml" '
                f'ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
                for i in range(len(slides_xml))
            )
            ct_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                '<Default Extension="xml" ContentType="application/xml"/>'
                '<Default Extension="png" ContentType="image/png"/>'
                '<Default Extension="jpeg" ContentType="image/jpeg"/>'
                '<Default Extension="jpg" ContentType="image/jpeg"/>'
                '<Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/>'
                '<Override PartName="/ppt/theme/theme1.xml" ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>'
                '<Override PartName="/ppt/slideMasters/slideMaster1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideMaster+xml"/>'
                '<Override PartName="/ppt/slideLayouts/slideLayout1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideLayout+xml"/>'
                f'{overrides}'
                + "".join(chart_overrides)
                + '</Types>'
            )
            zf.writestr("[Content_Types].xml", ct_xml)
            zf.writestr("_rels/.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="ppt/presentation.xml"/>'
                '</Relationships>'
            )
            slide_id_list = ''.join(
                f'<p:sldId id="{256+i}" r:id="rIdSl{i+1}"/>'
                for i in range(len(slides_xml))
            )
            pres_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
                '<p:sldMasterIdLst><p:sldMasterId id="2147483648" r:id="rIdMaster"/></p:sldMasterIdLst>'
                f'<p:sldIdLst>{slide_id_list}</p:sldIdLst>'
                f'<p:sldSz cx="{SLIDE_W_EMU}" cy="{SLIDE_H_EMU}"/>'
                f'<p:notesSz cx="6858000" cy="9144000"/>'
                '</p:presentation>'
            )
            zf.writestr("ppt/presentation.xml", pres_xml)
            pres_rels = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdMaster" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideMaster" '
                'Target="slideMasters/slideMaster1.xml"/>'
                '<Relationship Id="rIdTheme" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme" '
                'Target="theme/theme1.xml"/>'
            )
            for i in range(len(slides_xml)):
                pres_rels += f'<Relationship Id="rIdSl{i+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{i+1}.xml"/>'
            pres_rels += '</Relationships>'
            zf.writestr("ppt/_rels/presentation.xml.rels", pres_rels)
            theme_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="IBM">'
                '<a:themeElements>'
                '<a:clrScheme name="IBM">'
                '<a:dk1><a:srgbClr val="161616"/></a:dk1>'
                '<a:lt1><a:srgbClr val="FFFFFF"/></a:lt1>'
                '<a:dk2><a:srgbClr val="0F62FE"/></a:dk2>'
                '<a:lt2><a:srgbClr val="F4F4F4"/></a:lt2>'
                '<a:accent1><a:srgbClr val="0F62FE"/></a:accent1>'
                '<a:accent2><a:srgbClr val="8A3FFC"/></a:accent2>'
                '<a:accent3><a:srgbClr val="007D79"/></a:accent3>'
                '<a:accent4><a:srgbClr val="FF7EB6"/></a:accent4>'
                '<a:accent5><a:srgbClr val="FA4D56"/></a:accent5>'
                '<a:accent6><a:srgbClr val="FFC857"/></a:accent6>'
                '<a:hlink><a:srgbClr val="0F62FE"/></a:hlink>'
                '<a:folHlink><a:srgbClr val="8A3FFC"/></a:folHlink>'
                '</a:clrScheme>'
                '<a:fontScheme name="IBM">'
                '<a:majorFont><a:latin typeface="IBM Plex Sans"/><a:ea typeface=""/><a:cs typeface=""/></a:majorFont>'
                '<a:minorFont><a:latin typeface="IBM Plex Sans"/><a:ea typeface=""/><a:cs typeface=""/></a:minorFont>'
                '</a:fontScheme>'
                '<a:fmtScheme name="Office">'
                '<a:fillStyleLst>'
                '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
                '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
                '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
                '</a:fillStyleLst>'
                '<a:lnStyleLst>'
                '<a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>'
                '<a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>'
                '<a:ln w="9525" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>'
                '</a:lnStyleLst>'
                '<a:effectStyleLst>'
                '<a:effectStyle><a:effectLst/></a:effectStyle>'
                '<a:effectStyle><a:effectLst/></a:effectStyle>'
                '<a:effectStyle><a:effectLst/></a:effectStyle>'
                '</a:effectStyleLst>'
                '<a:bgFillStyleLst>'
                '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
                '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
                '<a:solidFill><a:schemeClr val="phClr"/></a:solidFill>'
                '</a:bgFillStyleLst>'
                '</a:fmtScheme>'
                '</a:themeElements>'
                '</a:theme>'
            )
            zf.writestr("ppt/theme/theme1.xml", theme_xml)
            slide_master_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<p:sldMaster xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
                '<p:cSld><p:bg><p:bgRef idx="1001"><a:schemeClr val="bg1"/></p:bgRef></p:bg>'
                '<p:spTree>'
                '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
                '<p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/>'
                '<a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>'
                '</p:spTree></p:cSld>'
                '<p:clrMap bg1="lt1" tx1="dk1" bg2="lt2" tx2="dk2" accent1="accent1" '
                'accent2="accent2" accent3="accent3" accent4="accent4" accent5="accent5" '
                'accent6="accent6" hlink="hlink" folHlink="folHlink"/>'
                '<p:sldLayoutIdLst><p:sldLayoutId id="2147483649" r:id="rIdLayout1"/></p:sldLayoutIdLst>'
                '</p:sldMaster>'
            )
            zf.writestr("ppt/slideMasters/slideMaster1.xml", slide_master_xml)
            zf.writestr(
                "ppt/slideMasters/_rels/slideMaster1.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdLayout1" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideLayout" '
                'Target="../slideLayouts/slideLayout1.xml"/>'
                '<Relationship Id="rIdTheme" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme" '
                'Target="../theme/theme1.xml"/>'
                '</Relationships>'
            )
            slide_layout_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<p:sldLayout xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" '
                'type="blank" preserve="1">'
                '<p:cSld name="Blank"><p:spTree>'
                '<p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr>'
                '<p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/>'
                '<a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>'
                '</p:spTree></p:cSld>'
                '</p:sldLayout>'
            )
            zf.writestr("ppt/slideLayouts/slideLayout1.xml", slide_layout_xml)
            zf.writestr(
                "ppt/slideLayouts/_rels/slideLayout1.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                '<Relationship Id="rIdMaster" '
                'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideMaster" '
                'Target="../slideMasters/slideMaster1.xml"/>'
                '</Relationships>'
            )
            for i, (name, content, rel_entries) in enumerate(slides_xml):
                zf.writestr(f"ppt/slides/slide{i+1}.xml", wrap_slide(content))
                if rel_entries:
                    slide_rels_xml = (
                        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                        + "".join(rel_entries) +
                        '</Relationships>'
                    )
                    zf.writestr(f"ppt/slides/_rels/slide{i+1}.xml.rels", slide_rels_xml)
            for fname, fbytes in media_files:
                zf.writestr(f"ppt/media/{fname}", fbytes)
            for part_name, xml_bytes in chart_parts:
                zf.writestr(part_name, xml_bytes)
        pptx_bytes = pptx_buf.getvalue()
        pptx_b64 = base64.b64encode(pptx_bytes).decode()
        data_uri = f"data:application/vnd.openxmlformats-officedocument.presentationml.presentation;base64,{pptx_b64}"
        return self._render_pptx_preview(title, client_name, sections, data_uri)
    def _render_pptx_preview(self, title, client_name, sections, data_uri):
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", title)[:50] or "deck"
        slide_parts = []
        slide_parts.append(
            f'<div class="sl" style="display:block;aspect-ratio:16/9;background:{IBM_BLUE_60};'
            f'padding:60px;color:#fff;font-family:\"IBM Plex Sans\",Calibri,sans-serif;position:relative">'
            f'<div style="font-size:36px;font-weight:700;margin-top:120px">{self._html_esc(title)}</div>'
            f'<div style="font-size:18px;margin-top:24px;opacity:0.9">IBM Consulting  |  Prepared for {self._html_esc(client_name)}</div>'
            f'<div style="font-size:14px;margin-top:12px;opacity:0.8">{time.strftime("%B %Y")}</div>'
            f'</div>'
        )
        for idx, section in enumerate(sections, start=1):
            has_kb = bool(section.get("_kb_match"))
            has_chart = bool(section.get("_chart_spec"))
            has_img = bool(section.get("_img_bytes")) or has_chart or has_kb
            text_col_w = "50%" if has_img else "100%"
            text_html = f'<h2 style="font-size:26px;color:{IBM_GRAY_100};font-weight:700;margin:0 0 16px">{self._html_esc(section.get("title", ""))}</h2>'
            for p in section.get("paragraphs", []) or []:
                text_html += f'<p style="font-size:13px;color:{IBM_GRAY_100};margin:8px 0;line-height:1.7">{self._html_esc(p)}</p>'
            bullets = section.get("bullets", []) or []
            if bullets:
                lis = "".join(
                    f'<li style="font-size:13px;color:{IBM_GRAY_100};margin:4px 0">{self._html_esc(b)}</li>'
                    for b in bullets
                )
                text_html += f'<ul style="padding-left:20px;margin:8px 0">{lis}</ul>'
            img_html = ""
            if has_kb:
                kb = section["_kb_match"]
                b64 = base64.b64encode(kb["image_bytes"]).decode()
                inner = (f'<img src="data:image/png;base64,{b64}" '
                         f'style="max-width:100%;max-height:55vh;border-radius:4px;object-fit:contain"/>'
                         f'<div style="font-size:10px;color:{IBM_GRAY_70};font-style:italic;margin-top:8px;text-align:center">'
                         f'📎 {self._html_esc(kb.get("caption",""))}  ·  {self._html_esc(kb.get("source_file",""))}</div>')
                img_html = (f'<div style="flex:0 0 46%;padding-left:20px;display:flex;flex-direction:column;justify-content:center">'
                            f'{inner}</div>')
            elif has_chart:
                svg = self._svg_chart_from_spec(section["_chart_spec"], width=520, height=300)
                img_html = (
                    f'<div style="flex:0 0 46%;padding-left:20px;display:flex;flex-direction:column;justify-content:center">'
                    f'{svg}'
                    + (
                        f'<div style="font-size:10px;color:{IBM_GRAY_70};font-style:italic;margin-top:8px;text-align:center">'
                        f'Chart — {self._html_esc(section.get("image_caption") or "")}</div>'
                        if section.get("image_caption") else ""
                    )
                    + '</div>'
                )
            elif section.get("_img_bytes"):
                img_b64 = base64.b64encode(section["_img_bytes"]).decode()
                img_html = (
                    f'<div style="flex:0 0 46%;padding-left:20px;display:flex;flex-direction:column;justify-content:center">'
                    f'<img src="data:image/png;base64,{img_b64}" '
                    f'style="max-width:100%;max-height:60vh;height:auto;border-radius:4px;object-fit:contain"/>'
                    + (
                        f'<div style="font-size:10px;color:{IBM_GRAY_70};font-style:italic;margin-top:8px;text-align:center">'
                        f'Figure — {self._html_esc(section.get("image_caption") or "")}</div>'
                        if section.get("image_caption") else ""
                    )
                    + '</div>'
                )
            slide_parts.append(
                f'<div class="sl" style="display:none;aspect-ratio:16/9;background:#fff;'
                f'border-left:8px solid {IBM_BLUE_60};padding:40px 48px;font-family:\"IBM Plex Sans\",Calibri,sans-serif;'
                f'position:relative">'
                f'<div style="display:flex;height:100%">'
                f'<div style="flex:1 1 {text_col_w};padding-right:16px">{text_html}</div>'
                f'{img_html}'
                f'</div></div>'
            )
        total = len(slide_parts)
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&display=swap">
<style>
*{{box-sizing:border-box;margin:0}}
html,body{{height:720px;min-height:720px}}
body{{font-family:\"IBM Plex Sans\",Calibri,system-ui,sans-serif;background:#F0F2F5;padding:12px;display:flex;align-items:stretch;justify-content:center}}
.dk{{border:3px solid {IBM_BLUE_60};border-radius:10px;overflow:hidden;width:100%;max-width:1280px;height:696px;margin:0 auto;background:#FFFFFF;display:flex;flex-direction:column;box-shadow:0 4px 16px rgba(15,98,254,0.15)}}
.tb{{display:flex;align-items:center;gap:8px;padding:10px 14px;background:{IBM_BLUE_60};flex-wrap:wrap;flex-shrink:0}}
.b{{border:none;border-radius:4px;padding:6px 14px;font-size:12px;cursor:pointer;font-family:\"IBM Plex Sans\",Calibri,sans-serif;font-weight:600;text-decoration:none;display:inline-block;background:{IBM_BLUE_60};color:#FFFFFF;border:1px solid #FFFFFF}}
.b:hover{{background:{IBM_BLUE_70};color:#FFFFFF}}
.sn{{color:#FFFFFF;font-size:12px;min-width:90px;text-align:center;font-weight:600}}
.sp{{flex:1}}
.sw{{background:{IBM_GRAY_10};padding:20px;overflow:auto;flex:1;min-height:0}}
.sl{{max-width:100%;width:100%;margin:0 auto 16px;box-shadow:0 2px 8px rgba(0,0,0,0.12);border-radius:4px;overflow:hidden;transform-origin:top center;transition:transform .15s ease}}
</style></head><body>
<div class="dk">
  <div class="tb">
    <button class="b" onclick="nav(-1)">← Prev</button>
    <span class="sn" id="sn">Slide 1 / {total}</span>
    <button class="b" onclick="nav(1)">Next →</button>
    <button class="b" onclick="zoomStep(-0.1)" title="Zoom out">🔍−</button>
    <button class="b" onclick="zoomReset()" title="Reset zoom"><span id="zl">100%</span></button>
    <button class="b" onclick="zoomStep(0.1)" title="Zoom in">🔍+</button>
    <span class="sp"></span>
    <a class="b" href="{data_uri}" download="{safe_name}.pptx">⬇ Download PPTX</a>
  </div>
  <div class="sw">{"".join(slide_parts)}</div>
</div>
<script>
(function(){{
  function fit(){{
    try{{
      var fe=window.frameElement;
      if(fe){{
        fe.style.height='720px';
        fe.style.minHeight='720px';
        fe.style.width='100%';
        fe.setAttribute('height','720');
      }}
    }}catch(e){{}}
    try{{window.parent.postMessage({{type:'ibm-docgen-resize',height:720}},'*');}}catch(e){{}}
  }}
  fit();setTimeout(fit,100);setTimeout(fit,500);setTimeout(fit,1500);
  window.addEventListener('load',fit);
  new MutationObserver(fit).observe(document.documentElement,{{attributes:true,childList:true,subtree:false}});
}})();
var cur=0,sl=document.querySelectorAll(".sl"),tot=sl.length,zm=1;
function applyZoom(){{sl.forEach(function(p){{p.style.transform='scale('+zm+')'}});var el=document.getElementById('zl');if(el)el.textContent=Math.round(zm*100)+'%'}}
function zoomStep(d){{zm=Math.max(0.5,Math.min(2,zm+d));applyZoom()}}
function zoomReset(){{zm=1;applyZoom()}}
function nav(d){{sl[cur].style.display="none";cur=Math.max(0,Math.min(tot-1,cur+d));
sl[cur].style.display="block";document.getElementById("sn").textContent="Slide "+(cur+1)+" / "+tot;
var sw=document.querySelector(".sw");if(sw)sw.scrollTop=0}}
document.addEventListener("keydown",function(e){{
if(e.key==="ArrowLeft")nav(-1);if(e.key==="ArrowRight")nav(1);
if((e.ctrlKey||e.metaKey)&&e.key==="=")e.preventDefault()||zoomStep(0.1);
if((e.ctrlKey||e.metaKey)&&e.key==="-")e.preventDefault()||zoomStep(-0.1);
if((e.ctrlKey||e.metaKey)&&e.key==="0")e.preventDefault()||zoomReset()}});
</script></body></html>"""
        return HTMLResponse(content=html, headers={"Content-Disposition": "inline"})
    def _ext(self, name: str) -> str:
        if "." not in name: return ""
        return "." + name.rsplit(".", 1)[-1].lower()
    def _classify(self, filename: str) -> str:
        f = (filename or "").lower()
        if "case_study" in f or "case-study" in f or "success" in f: return "case_study"
        if "architecture" in f or "solution" in f or "blueprint" in f: return "solution_brief"
        if "methodology" in f or "approach" in f or "framework" in f: return "methodology"
        if "offering" in f or "capability" in f or "portfolio" in f: return "capability"
        return "general"
    def _humanize(self, stem: str) -> str:
        s = re.sub(r"[_\-]+", " ", stem)
        s = re.sub(r"(\d+)", r" \1", s)
        return s.strip().title()
    def _html_esc(self, s: str) -> str:
        return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    async def _emit(self, emitter, msg, done=False):
        if emitter:
            try: await emitter({"type": "status", "data": {"description": msg, "done": done}})
            except Exception: pass
    def _start_heartbeat(self, emitter, eta_seconds=None, initial_phase=None): return None
    async def _stop_heartbeat(self, task): return None
    def _set_phase(self, phase, eta_seconds=None): pass
    def _eta_for(self, phase_key, scale=1.0): return 5
